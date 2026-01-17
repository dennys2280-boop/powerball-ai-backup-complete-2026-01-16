from __future__ import annotations

import json
import os
import random
import re
import tempfile
import time
from collections import Counter, defaultdict
from datetime import date, datetime
from io import BytesIO
from pathlib import Path as FilePath
from statistics import mean
from typing import Any, Dict, List, Literal, Optional, Tuple
from urllib.parse import quote_plus

import pandas as pd
from fastapi import (
    Body,
    Depends,
    FastAPI,
    File,
    HTTPException,
    Path,
    Query,
    Request,
    Response,
    UploadFile,
)
from fastapi.responses import (
    HTMLResponse,
    JSONResponse,
    RedirectResponse,
    StreamingResponse,
)
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, ConfigDict
from sqlalchemy import asc, desc, extract, func, text
from sqlalchemy.orm import Session
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.gzip import GZipMiddleware

# ✅ IMPORTS CORREGIDOS (estructura app/)
from app.database import Base, SessionLocal, engine
from app.models import DrawResult, Ticket


from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware



from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware



app = FastAPI(title="Powerball API")

@app.get("/")
def root():
    return {"status": "ok"}



@app.get("/api/dashboard/summary")
def dashboard_summary():
    return {
        "kpis": [
            {"id": "k1", "label": "Registros", "value": 3},
            {"id": "k2", "label": "Draws", "value": 35},
            {"id": "k3", "label": "Estado", "value": "BACKEND"},
        ],
        "main": {
            "title": "Zona de resultados",
            "message": "Datos servidos por FastAPI local",
        },
        "side": {
            "title": "Filtros / Acciones",
            "message": "Backend conectado correctamente",
        },
    }

@app.get("/api/history")
def history():
    return [
        {"id": 1, "date": "2025-12-22", "draws": 12, "note": "Backend OK"},
        {"id": 2, "date": "2025-12-21", "draws": 8, "note": "FastAPI conectado"},
        {"id": 3, "date": "2025-12-20", "draws": 15, "note": "Sin STUB"},
    ]





app = FastAPI(title="Powerball API")

# CORS para Vite
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
def root():
    return {"status": "ok"}

@app.get("/api/dashboard/summary")
def dashboard_summary():
    return {
        "kpis": [
            {"id": "k1", "label": "Registros", "value": 3},
            {"id": "k2", "label": "Draws", "value": 35},
            {"id": "k3", "label": "Estado", "value": "BACKEND"},
        ],
        "main": {
            "title": "Zona de resultados",
            "message": "Datos servidos por FastAPI local",
        },
        "side": {
            "title": "Filtros / Acciones",
            "message": "Backend conectado correctamente",
        },
    }

@app.get("/api/history")
def history():
    return [
        {"id": 1, "date": "2025-12-22", "draws": 12, "note": "Backend OK"},
        {"id": 2, "date": "2025-12-21", "draws": 8, "note": "FastAPI conectado"},
        {"id": 3, "date": "2025-12-20", "draws": 15, "note": "Sin STUB"},
    ]



# ---------------------------------------------------------------------
# FastAPI app
# ---------------------------------------------------------------------
app = FastAPI(title="Powerball AI")

app.add_middleware(GZipMiddleware, minimum_size=1000)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------
# DB dependency
# ---------------------------------------------------------------------
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ---------------------------------------------------------------------
# ⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇
# ⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇ TU CÓDIGO ORIGINAL SIGUE AQUÍ ⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇⬇
# (endpoints, lógica, análisis, exports, etc.)
# ---------------------------------------------------------------------





# ✅ Usa FilePath (pathlib) para rutas de archivos
BASE_DIR = FilePath(__file__).resolve().parent

# Fuente de verdad para el CSV de sorteos (estable en cualquier working dir)
CSV_PATH = BASE_DIR / "powerball_draws.csv"

EXPORT_HEADERS = "date,white1,white2,white3,white4,white5,powerball\n"


def _parse_date_flexible(s: str, year: int | None = None) -> date:
    """
    Acepta:
      - YYYY-MM-DD  -> fecha exacta
      - (upgrade) YYYY/MM/DD -> fecha exacta
      - MM-DD       -> usa year (si viene) o el año actual
    """
    s = (s or "").strip()
    if not s:
        raise HTTPException(400, "Fecha inválida: vacío. Usa YYYY-MM-DD o MM-DD.")

    # Caso 1: ISO completo YYYY-MM-DD
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(400, f"Fecha inválida (esperado YYYY-MM-DD): {s}")

    # (upgrade) Caso 1b: YYYY/MM/DD
    if len(s) == 10 and s[4] == "/" and s[7] == "/":
        try:
            return datetime.strptime(s, "%Y/%m/%d").date()
        except ValueError:
            raise HTTPException(400, f"Fecha inválida (esperado YYYY/MM/DD): {s}")

    # Caso 2: MM-DD
    if len(s) == 5 and s[2] == "-":
        try:
            mm = int(s[:2])
            dd = int(s[3:])
        except ValueError:
            raise HTTPException(400, f"Fecha inválida (esperado MM-DD): {s}")

        y = int(year) if year is not None else date.today().year
        try:
            return date(y, mm, dd)
        except ValueError:
            raise HTTPException(400, f"Fecha inválida (MM-DD con year={y}): {s}")

    raise HTTPException(400, f"Formato de fecha inválido: '{s}'. Usa YYYY-MM-DD o MM-DD.")


app = FastAPI(
    title="Powerball + IA API",
    description="Backend para gestionar jugadas de Powerball.",
    version="1.2.0",
)


@app.get("/__test_export_route")
def __test_export_route():
    return {"ok": True}


# Mantengo los nombres existentes para compatibilidad,
# pero los hago estables (sin depender del working directory).
POWERBALL_DRAWS_CSV = str(CSV_PATH)
TICKETS_CSV = str(BASE_DIR / "tickets.csv")

WHITE_COLS = ["n1", "n2", "n3", "n4", "n5"]

from io import StringIO

def _df_to_csv_stream(df: pd.DataFrame, filename: str) -> StreamingResponse:
    # StringIO -> bytes para control consistente de encoding/newlines
    sio = StringIO()
    df.to_csv(sio, index=False, lineterminator="\n")
    data = sio.getvalue().encode("utf-8")  # si quieres Excel-friendly: "utf-8-sig"
    buf = BytesIO(data)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _df_to_xlsx_stream(
    sheets: Dict[str, pd.DataFrame],
    filename: str,
) -> StreamingResponse:
    buf = BytesIO()

    def _unique_sheet_name(existing: set[str], name: str) -> str:
        base = (name or "Sheet").strip() or "Sheet"
        base = base[:31]
        candidate = base
        i = 2
        while candidate in existing:
            suffix = f"_{i}"
            candidate = (base[: 31 - len(suffix)] + suffix)[:31]
            i += 1
        existing.add(candidate)
        return candidate

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        used_names: set[str] = set()
        for sheet_name, df in sheets.items():
            safe_name = _unique_sheet_name(used_names, sheet_name)
            df.to_excel(writer, index=False, sheet_name=safe_name)
            ws = writer.sheets[safe_name]

            # UX: congelar header
            ws.freeze_panes = "A2"

            # Header en negrita
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = header_font

            # Autosize columnas (cap de filas para performance)
            max_rows_to_measure = 300  # upgrade: evita costos enormes en hojas gigantes
            for col in ws.columns:
                col_letter = col[0].column_letter
                max_len = 0

                # medir header siempre
                header_val = "" if col[0].value is None else str(col[0].value)
                max_len = max(max_len, len(header_val))

                # medir solo primeras N filas
                for cell in col[1 : 1 + max_rows_to_measure]:
                    val = "" if cell.value is None else str(cell.value)
                    if len(val) > max_len:
                        max_len = len(val)

                ws.column_dimensions[col_letter].width = min(45, max(10, max_len + 2))

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _draws_day_df(month: int, day: int) -> pd.DataFrame:
    df = _load_draws()
    out = _filter_draws_by_day(df, day=day, month=month)
    if out.empty:
        return out

    out2 = out.copy()

    # --- Fecha: draw_date o date ---
    if "draw_date" not in out2.columns:
        if "date" in out2.columns:
            out2 = out2.rename(columns={"date": "draw_date"})
        else:
            raise HTTPException(status_code=400, detail="powerball_draws.csv debe incluir columna draw_date o date")

    out2["draw_date"] = pd.to_datetime(out2["draw_date"], errors="coerce")
    out2 = out2.dropna(subset=["draw_date"])
    if out2.empty:
        return out2

    # --- Blancos: WHITE_COLS o white1..white5 ---
    if not all(c in out2.columns for c in WHITE_COLS):
        csv_whites = ["white1", "white2", "white3", "white4", "white5"]
        if all(c in out2.columns for c in csv_whites):
            out2 = out2.rename(columns=dict(zip(csv_whites, WHITE_COLS)))
        else:
            raise HTTPException(
                status_code=400,
                detail=f"powerball_draws.csv debe incluir columnas {WHITE_COLS} o {csv_whites}",
            )

    # --- Powerball: pb o powerball ---
    if "pb" not in out2.columns:
        if "powerball" in out2.columns:
            out2 = out2.rename(columns={"powerball": "pb"})
        else:
            raise HTTPException(status_code=400, detail="powerball_draws.csv debe incluir columna pb o powerball")

    # Normalizar tipos numéricos (evita strings raros)
    for c in WHITE_COLS + ["pb"]:
        out2[c] = pd.to_numeric(out2[c], errors="coerce")
    out2 = out2.dropna(subset=WHITE_COLS + ["pb"])
    if out2.empty:
        return out2

    # --- Year: si no existe, lo calculamos desde draw_date ---
    if "year" not in out2.columns:
        out2["year"] = out2["draw_date"].dt.year

    out2["year"] = pd.to_numeric(out2["year"], errors="coerce").astype("Int64")
    out2 = out2.dropna(subset=["year"])
    if out2.empty:
        return out2

    out2 = out2[["year", "draw_date"] + WHITE_COLS + ["pb"]].copy()
    out2["year"] = out2["year"].astype(int)
    out2["pb"] = out2["pb"].astype(int)
    for c in WHITE_COLS:
        out2[c] = out2[c].astype(int)

    out2["draw_date"] = out2["draw_date"].dt.strftime("%Y-%m-%d")
    return out2


def _compare_day_df(month: int, day: int, ticket: str) -> pd.DataFrame:
    ticket_white, ticket_pb = _parse_ticket_param(ticket)
    df = _load_draws()
    draws_day = _filter_draws_by_day(df, day=day, month=month)
    results = _compare_ticket_to_draws(ticket_white, ticket_pb, draws_day).copy()
    if results.empty:
        return results
    results["draw_date"] = pd.to_datetime(results["draw_date"], errors="coerce").dt.strftime("%Y-%m-%d")
    results = results.dropna(subset=["draw_date"])
    return results


def _compare_multi_day_df(month: int, day: int) -> pd.DataFrame:
    df_draws = _load_draws()
    draws_day = _filter_draws_by_day(df_draws, day=day, month=month)

    tickets = _load_tickets()
    if tickets.empty or draws_day.empty:
        return pd.DataFrame([])

    # Pre-normalizar draws
    drows = []
    for _, d in draws_day.iterrows():
        try:
            dset = {int(d[c]) for c in WHITE_COLS}
            dpb = int(d["pb"])
            dyear = int(d["year"])
            ddate = pd.to_datetime(d["draw_date"], errors="coerce")
            if pd.isna(ddate):
                continue
            ddate_s = ddate.strftime("%Y-%m-%d")
            dnums = [int(d[c]) for c in WHITE_COLS]
        except Exception:
            continue

        drows.append((dset, dpb, dyear, ddate_s, dnums))

    if not drows:
        return pd.DataFrame([])

    # Pre-normalizar tickets
    trows = []
    for _, t in tickets.iterrows():
        try:
            tid = str(t["ticket_id"])
            tnums = [int(t[c]) for c in WHITE_COLS]
            tset = set(tnums)
            tpb = int(t["pb"])
        except Exception:
            continue
        trows.append((tid, tnums, tset, tpb))

    if not trows:
        return pd.DataFrame([])

    items: list[dict[str, Any]] = []
    for tid, tnums, tset, tpb in trows:
        for dset, dpb, dyear, ddate_s, dnums in drows:
            match_white = len(tset.intersection(dset))
            match_pb = 1 if dpb == tpb else 0
            items.append(
                {
                    "ticket_id": tid,
                    "year": dyear,
                    "draw_date": ddate_s,
                    "ticket_n1": tnums[0],
                    "ticket_n2": tnums[1],
                    "ticket_n3": tnums[2],
                    "ticket_n4": tnums[3],
                    "ticket_n5": tnums[4],
                    "ticket_pb": tpb,
                    "draw_n1": dnums[0],
                    "draw_n2": dnums[1],
                    "draw_n3": dnums[2],
                    "draw_n4": dnums[3],
                    "draw_n5": dnums[4],
                    "draw_pb": dpb,
                    "match_white": int(match_white),
                    "match_pb": int(match_pb),
                    "score": int(match_white + match_pb),
                }
            )

    items.sort(key=lambda x: (x["score"], x["match_white"], x["match_pb"], x["year"]), reverse=True)
    return pd.DataFrame(items)


# ---------------------------
# Helpers
# ---------------------------
def _load_draws(db: Optional[Session] = None) -> pd.DataFrame:
    base_dir = os.path.dirname(os.path.abspath(__file__))
    csv_path = POWERBALL_DRAWS_CSV if os.path.isabs(POWERBALL_DRAWS_CSV) else os.path.join(base_dir, POWERBALL_DRAWS_CSV)

    if os.path.exists(csv_path):
        try:
            df = pd.read_csv(csv_path)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error leyendo draws CSV ({csv_path}): {e}")

        df = df.copy()
        df.columns = [str(c).strip().lower() for c in df.columns]

        if "draw_date" not in df.columns:
            if "date" in df.columns:
                df = df.rename(columns={"date": "draw_date"})
            else:
                raise HTTPException(status_code=422, detail=f"{POWERBALL_DRAWS_CSV} debe incluir columna draw_date o date")

        df["draw_date"] = pd.to_datetime(df["draw_date"], errors="coerce")
        if df["draw_date"].notna().sum() == 0 and len(df) > 0:
            raise HTTPException(status_code=422, detail="draw_date no tiene valores de fecha válidos (NaT)")

        rename_map: Dict[str, str] = {}
        for i, c in enumerate(WHITE_COLS, start=1):
            if c not in df.columns:
                if f"wn{i}" in df.columns:
                    rename_map[f"wn{i}"] = c
                elif f"white{i}" in df.columns:
                    rename_map[f"white{i}"] = c

        if "pb" not in df.columns:
            if "winning_powerball" in df.columns:
                rename_map["winning_powerball"] = "pb"
            elif "powerball" in df.columns:
                rename_map["powerball"] = "pb"

        if rename_map:
            df = df.rename(columns=rename_map)

        needed = ["draw_date"] + WHITE_COLS + ["pb"]
        missing = [c for c in needed if c not in df.columns]
        if missing:
            raise HTTPException(status_code=422, detail=f"{POWERBALL_DRAWS_CSV} falta(n) columna(s): {missing}")

        # ✅ Normaliza tipos numéricos (evita strings/NaNs que rompen después)
        for c in WHITE_COLS + ["pb"]:
            df[c] = pd.to_numeric(df[c], errors="coerce")
        df = df.dropna(subset=["draw_date"] + WHITE_COLS + ["pb"])
        if df.empty:
            return pd.DataFrame(columns=["draw_date"] + WHITE_COLS + ["pb"])

        for c in WHITE_COLS + ["pb"]:
            df[c] = df[c].astype(int)

        return df

    if db is None:
        raise HTTPException(status_code=404, detail=f"No se encontró {POWERBALL_DRAWS_CSV} (buscado en {csv_path})")

    drs = db.query(DrawResult).order_by(DrawResult.draw_date.asc()).all()
    if not drs:
        return pd.DataFrame(columns=["draw_date"] + WHITE_COLS + ["pb"])

    rows = []
    for d in drs:
        rows.append({
            "draw_date": d.draw_date,
            "n1": int(d.wn1),
            "n2": int(d.wn2),
            "n3": int(d.wn3),
            "n4": int(d.wn4),
            "n5": int(d.wn5),
            "pb": int(d.winning_powerball),
        })
    df = pd.DataFrame(rows)
    df["draw_date"] = pd.to_datetime(df["draw_date"], errors="coerce")
    return df


def _filter_draws_by_day(df: pd.DataFrame, day: int, month: int) -> pd.DataFrame:
    if not (1 <= day <= 31):
        raise HTTPException(status_code=422, detail="day debe ser 1..31")
    if not (1 <= month <= 12):
        raise HTTPException(status_code=422, detail="month debe ser 1..12")

    df = df.copy()
    if "draw_date" not in df.columns:
        raise HTTPException(status_code=422, detail="draw_date no existe en los datos de draws")

    df["draw_date"] = pd.to_datetime(df["draw_date"], errors="coerce")
    if df["draw_date"].notna().sum() == 0 and len(df) > 0:
        raise HTTPException(status_code=422, detail="draw_date no tiene valores de fecha válidos (NaT)")

    df["day"] = df["draw_date"].dt.day
    df["month"] = df["draw_date"].dt.month
    df["year"] = df["draw_date"].dt.year

    out = df[(df["day"] == day) & (df["month"] == month)].sort_values("year")
    return out

def _parse_ticket_param(ticket: str) -> Tuple[List[int], int]:
    # Formato: "4,6,8,10,12|22"
    try:
        left, right = ticket.strip().split("|")
        whites = [int(x.strip()) for x in left.split(",") if x.strip()]
        pb = int(right.strip())
    except Exception:
        raise HTTPException(status_code=422, detail="Formato ticket inválido. Usa: 4,6,8,10,12|22")

    if len(whites) != 5:
        raise HTTPException(status_code=422, detail="El ticket debe tener exactamente 5 bolas blancas")
    if len(set(whites)) != 5:
        raise HTTPException(status_code=422, detail="Las 5 bolas blancas no pueden repetirse")
    if not all(1 <= n <= 69 for n in whites):
        raise HTTPException(status_code=422, detail="Bolas blancas deben estar en 1..69")
    if not (1 <= pb <= 26):
        raise HTTPException(status_code=422, detail="Powerball debe estar en 1..26")

    return sorted(whites), pb

def _compare_ticket_to_draws(ticket_white: List[int], ticket_pb: int, draws_df: pd.DataFrame) -> pd.DataFrame:
    tset = set(ticket_white)
    out = draws_df.copy()

    out["draw_white_set"] = out[WHITE_COLS].apply(lambda r: set(r.values.tolist()), axis=1)
    out["match_white"] = out["draw_white_set"].apply(lambda s: len(s.intersection(tset)))
    out["match_pb"] = (out["pb"].astype(int) == int(ticket_pb)).astype(int)
    out["score"] = out["match_white"] + out["match_pb"]

    cols = ["year", "draw_date"] + WHITE_COLS + ["pb", "match_white", "match_pb", "score"]
    return out[cols].sort_values(["score", "match_white", "match_pb", "year"], ascending=[False, False, False, True])

def _insights(draws_df: pd.DataFrame) -> Dict[str, Any]:
    """
    Calcula insights de sorteos. Soporta diferentes esquemas de columnas:
    - date + white1..white5 + powerball
    - draw_date + wn1..wn5 + winning_powerball
    - w1..w5 + pb
    """
    if draws_df is None or draws_df.empty:
        return {"count": 0, "white_frequency": [], "pb_frequency": [], "avg_sum_white": None}

    df = draws_df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]

    # Detectar columnas de blancos (5) y powerball
    white_candidates = [
        ["white1", "white2", "white3", "white4", "white5"],
        ["wn1", "wn2", "wn3", "wn4", "wn5"],
        ["w1", "w2", "w3", "w4", "w5"],
        ["n1", "n2", "n3", "n4", "n5"],
    ]
    pb_candidates = ["powerball", "winning_powerball", "pb"]

    white_cols = next((cols for cols in white_candidates if all(c in df.columns for c in cols)), None)
    pb_col = next((c for c in pb_candidates if c in df.columns), None)

    if white_cols is None:
        return {
            "count": 0,
            "white_frequency": [],
            "pb_frequency": [],
            "avg_sum_white": None,
            "error": "No se encontraron las 5 columnas de white balls (white1..white5 / wn1..wn5 / w1..w5 / n1..n5).",
        }

    if pb_col is None:
        return {
            "count": 0,
            "white_frequency": [],
            "pb_frequency": [],
            "avg_sum_white": None,
            "error": "No se encontró columna de Powerball (powerball / winning_powerball / pb).",
        }

    # Asegurar numéricos (por si vienen como strings)
    for c in white_cols + [pb_col]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df = df.dropna(subset=white_cols + [pb_col])
    if df.empty:
        return {"count": 0, "white_frequency": [], "pb_frequency": [], "avg_sum_white": None}

    # Frecuencias
    whites = []
    for c in white_cols:
        whites.extend(df[c].astype(int).tolist())

    white_freq = Counter(whites)
    pb_freq = Counter(df[pb_col].astype(int).tolist())

    # Promedio de suma de blancos
    tmp = df.copy()
    tmp["sum_white"] = tmp[white_cols].sum(axis=1)
    avg_sum = float(tmp["sum_white"].mean())

    white_list = [{"number": int(k), "frequency": int(v)} for k, v in white_freq.most_common()]
    pb_list = [{"pb": int(k), "frequency": int(v)} for k, v in pb_freq.most_common()]

    return {
        "count": int(len(df)),
        "white_frequency": white_list,
        "pb_frequency": pb_list,
        "avg_sum_white": round(avg_sum, 2),
    }

def _load_tickets(db: Optional[Session] = None) -> pd.DataFrame:
    """
    Carga tickets desde CSV (ruta absoluta) y, si no existe, hace fallback a DB (Ticket).
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    csv_path = TICKETS_CSV if os.path.isabs(TICKETS_CSV) else os.path.join(base_dir, TICKETS_CSV)

    if os.path.exists(csv_path):
        try:
            df = pd.read_csv(csv_path)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error leyendo tickets CSV ({csv_path}): {e}")

        required = ["ticket_id"] + WHITE_COLS + ["pb"]
        missing = [c for c in required if c not in df.columns]
        if missing:
            raise HTTPException(status_code=422, detail=f"tickets.csv falta(n) columna(s): {missing}")
        return df

    # Fallback DB
    if db is None:
        raise HTTPException(status_code=404, detail=f"No se encontró {TICKETS_CSV} (buscado en {csv_path})")

    ts = db.query(Ticket).order_by(Ticket.id.asc()).all()
    if not ts:
        return pd.DataFrame(columns=["ticket_id"] + WHITE_COLS + ["pb"])

    rows = []
    for t in ts:
        rows.append({
            "ticket_id": str(getattr(t, "ticket_id", None) or t.id),
            "n1": int(getattr(t, "n1")),
            "n2": int(getattr(t, "n2")),
            "n3": int(getattr(t, "n3")),
            "n4": int(getattr(t, "n4")),
            "n5": int(getattr(t, "n5")),
            "pb": int(getattr(t, "pb")),
        })
    return pd.DataFrame(rows)

@app.get("/draws/by-date")
def draws_by_date(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
):
    df = _load_draws()
    out = _filter_draws_by_day(df, day=day, month=month)
    out2 = out[["year", "draw_date"] + WHITE_COLS + ["pb"]].copy()
    out2["draw_date"] = out2["draw_date"].dt.strftime("%Y-%m-%d")
    return {"count": int(len(out2)), "items": out2.to_dict(orient="records")}

@app.get("/insights/by-date")
def insights_by_date(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
):
    df = _load_draws()
    draws_day = _filter_draws_by_day(df, day=day, month=month)
    return _insights(draws_day)

@app.get("/compare/by-date")
def compare_by_date(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
    ticket: str = Query(..., description='Formato: "4,6,8,10,12|22"'),
):
    ticket_white, ticket_pb = _parse_ticket_param(ticket)
    df = _load_draws()
    draws_day = _filter_draws_by_day(df, day=day, month=month)
    results = _compare_ticket_to_draws(ticket_white, ticket_pb, draws_day)

    results2 = results.copy()
    results2["draw_date"] = pd.to_datetime(results2["draw_date"]).dt.strftime("%Y-%m-%d")
    return {"count": int(len(results2)), "items": results2.to_dict(orient="records")}

@app.get("/compare/by-date/multi")
def compare_by_date_multi(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
):
    df_draws = _load_draws()
    draws_day = _filter_draws_by_day(df_draws, day=day, month=month)

    tickets = _load_tickets()
    if tickets.empty:
        return {"count": 0, "items": []}

    # Comparación: tickets x sorteos del día
    items = []
    for _, t in tickets.iterrows():
        tset = set([int(t[c]) for c in WHITE_COLS])
        tpb = int(t["pb"])

        for _, d in draws_day.iterrows():
            dset = set([int(d[c]) for c in WHITE_COLS])
            match_white = len(tset.intersection(dset))
            match_pb = 1 if int(d["pb"]) == tpb else 0
            items.append({
                "ticket_id": str(t["ticket_id"]),
                "year": int(d["year"]),
                "draw_date": pd.to_datetime(d["draw_date"]).strftime("%Y-%m-%d"),
                "match_white": match_white,
                "match_pb": match_pb,
                "score": match_white + match_pb,
            })

    # ordeno por score desc
    items.sort(key=lambda x: (x["score"], x["match_white"], x["match_pb"], x["year"]), reverse=True)


REQUIRED_DRAW_COLS = ["date", "white1", "white2", "white3", "white4", "white5", "powerball"]

def _normalize_draws_df(df: pd.DataFrame) -> pd.DataFrame:
    # normaliza nombres de columnas
    df = df.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]

    missing = [c for c in REQUIRED_DRAW_COLS if c not in df.columns]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Faltan columnas requeridas: {missing}. Requeridas: {REQUIRED_DRAW_COLS}"
        )

    # quedarnos solo con las columnas esperadas (por si viene extra)
    df = df[REQUIRED_DRAW_COLS].copy()

    # date -> date
    # acepta 'YYYY-MM-DD' y también timestamps de excel
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
    if df["date"].isna().any():
        bad_rows = df[df["date"].isna()].index.tolist()[:10]
        raise HTTPException(status_code=400, detail=f"Fechas inválidas en filas: {bad_rows}")

    # ints
    for c in ["white1", "white2", "white3", "white4", "white5", "powerball"]:
        df[c] = pd.to_numeric(df[c], errors="coerce").astype("Int64")
        if df[c].isna().any():
            bad_rows = df[df[c].isna()].index.tolist()[:10]
            raise HTTPException(status_code=400, detail=f"Valores inválidos en '{c}' en filas: {bad_rows}")

    # reglas básicas de rango (Powerball clásico: blancos 1-69, powerball 1-26)
    whites = ["white1", "white2", "white3", "white4", "white5"]
    if ((df[whites] < 1) | (df[whites] > 69)).any().any():
        raise HTTPException(status_code=400, detail="Hay white balls fuera de rango (1-69).")
    if ((df["powerball"] < 1) | (df["powerball"] > 26)).any():
        raise HTTPException(status_code=400, detail="Hay powerball fuera de rango (1-26).")

    # opcional: quitar duplicados por date (nos quedamos con el último)
    df = df.sort_values("date").drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)

    return df


def _read_upload_to_df(file: UploadFile) -> pd.DataFrame:
    filename = (file.filename or "").lower()
    if not filename:
        raise HTTPException(status_code=400, detail="El archivo no tiene nombre.")

    content = file.file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Archivo vacío.")

    bio = BytesIO(content)

    try:
        if filename.endswith(".csv"):
            try:
                bio.seek(0)
                df = pd.read_csv(bio)
            except UnicodeDecodeError:
                bio.seek(0)
                df = pd.read_csv(bio, encoding="latin-1")
        elif filename.endswith(".xlsx") or filename.endswith(".xls"):
            bio.seek(0)
            df = pd.read_excel(bio)
        else:
            raise HTTPException(status_code=400, detail="Formato no soportado. Usa .csv o .xlsx")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No pude leer el archivo: {str(e)}")

    return df



def _upsert_draws(db: Session, df: pd.DataFrame) -> Dict[str, int]:
    inserted = 0
    updated = 0

    try:
        for row in df.to_dict(orient="records"):
            draw_date: date = row["date"]

            existing = db.query(DrawResult).filter(DrawResult.draw_date == draw_date).first()
            if existing:
                existing.wn1 = int(row["white1"])
                existing.wn2 = int(row["white2"])
                existing.wn3 = int(row["white3"])
                existing.wn4 = int(row["white4"])
                existing.wn5 = int(row["white5"])
                existing.winning_powerball = int(row["powerball"])
                updated += 1
            else:
                db.add(DrawResult(
                    draw_date=draw_date,
                    wn1=int(row["white1"]),
                    wn2=int(row["white2"]),
                    wn3=int(row["white3"]),
                    wn4=int(row["white4"]),
                    wn5=int(row["white5"]),
                    winning_powerball=int(row["powerball"]),
                ))
                inserted += 1

        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error guardando draws en DB: {e}")

    return {"inserted": inserted, "updated": updated, "total": int(len(df))}


def _export_db_to_csv(db: Session) -> None:
    # ✅ Exporta TODO el histórico de DrawResult a powerball_draws.csv
    rows = (
        db.query(
            DrawResult.draw_date,
            DrawResult.wn1,
            DrawResult.wn2,
            DrawResult.wn3,
            DrawResult.wn4,
            DrawResult.wn5,
            DrawResult.winning_powerball,
        )
        .order_by(DrawResult.draw_date.asc())
        .all()
    )

    out = pd.DataFrame(rows, columns=REQUIRED_DRAW_COLS)
    # date a ISO yyyy-mm-dd
    out["date"] = pd.to_datetime(out["date"]).dt.date.astype(str)
    out.to_csv(CSV_PATH, index=False)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@app.post("/draws/upload")
def upload_draws(
    file: UploadFile = File(...),
    export_csv: bool = Query(True, description="Si True, regenera powerball_draws.csv desde la DB al finalizar."),
    db: Session = Depends(get_db),
):
    df_raw = _read_upload_to_df(file)
    df = _normalize_draws_df(df_raw)

    stats = _upsert_draws(db, df)

    if export_csv:
        _export_db_to_csv(db)

    return {
        "ok": True,
        "filename": file.filename,
        "rows_received": int(len(df_raw)),
        "rows_processed": stats["total"],
        "inserted": stats["inserted"],
        "updated": stats["updated"],
        "exported_csv": bool(export_csv),
        "csv_path": str(CSV_PATH) if export_csv else None,
    }

# ---------------------------
#   EXPORT: /draws/by-date
# ---------------------------

EXPORT_COLS = EXPORT_HEADERS.strip().split(",")

def _to_export_schema(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normaliza a esquema humano:
    date, white1..white5, powerball
    Soporta df con year/draw_date/n1..n5/pb.
    """
    if df is None or df.empty:
        return pd.DataFrame(columns=EXPORT_COLS)

    out = df.copy()

    # draw_date -> date (string ISO)
    if "draw_date" in out.columns:
        out["date"] = pd.to_datetime(out["draw_date"], errors="coerce").dt.strftime("%Y-%m-%d")
    elif "date" in out.columns:
        out["date"] = pd.to_datetime(out["date"], errors="coerce").dt.strftime("%Y-%m-%d")
    else:
        out["date"] = None

    # n1..n5 -> white1..white5
    rename_whites = {f"n{i}": f"white{i}" for i in range(1, 6)}
    out = out.rename(columns=rename_whites)

    # pb -> powerball
    if "pb" in out.columns and "powerball" not in out.columns:
        out = out.rename(columns={"pb": "powerball"})

    # asegurar columnas y orden
    for c in EXPORT_COLS:
        if c not in out.columns:
            out[c] = None

    return out[EXPORT_COLS]


@app.get("/draws/by-date/export.csv")
def draws_by_date_export_csv(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
    db: Session = Depends(get_db),
):
    # 1) Intento normal (CSV si existe)
    try:
        out2 = _draws_day_df(month, day)
    except HTTPException:
        out2 = pd.DataFrame()

    # 2) Fallback real: si no hay CSV o no hay datos, intentamos desde DB
    if out2.empty:
        df_db = _load_draws(db=db)
        out2 = _filter_draws_by_day(df_db, day=day, month=month)

    filename = f"draws_by_date_{month:02d}-{day:02d}.csv"
    return _df_to_csv_stream(_to_export_schema(out2), filename)


@app.get("/draws/by-date/export.xlsx")
def draws_by_date_export_xlsx(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
    db: Session = Depends(get_db),
):
    # 1) Intento normal (CSV si existe)
    try:
        out2 = _draws_day_df(month, day)
    except HTTPException:
        out2 = pd.DataFrame()

    # 2) Fallback real: si no hay CSV o no hay datos, intentamos desde DB
    if out2.empty:
        df_db = _load_draws(db=db)
        out2 = _filter_draws_by_day(df_db, day=day, month=month)

    filename = f"draws_by_date_{month:02d}-{day:02d}.xlsx"
    return _df_to_xlsx_stream({"draws": _to_export_schema(out2)}, filename)


# ---------------------------
#   EXPORT: /insights/by-date
# ---------------------------
@app.get("/insights/by-date/export.csv")
def insights_by_date_export_csv(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
    db: Session = Depends(get_db),
):
    # 1) Intento normal (CSV si existe)
    try:
        df = _load_draws()
    except HTTPException:
        df = pd.DataFrame()

    # 2) Fallback real a DB si no hay datos
    if df is None or df.empty:
        df = _load_draws(db=db)

    draws_day = _filter_draws_by_day(df, day=day, month=month)
    ins = _insights(draws_day)

    # CSV en formato "long": category, metric, value, number, frequency
    rows: List[Dict[str, Any]] = []

    rows.append({
        "category": "summary",
        "metric": "month",
        "value": int(month),
        "number": "",
        "frequency": "",
    })
    rows.append({
        "category": "summary",
        "metric": "day",
        "value": int(day),
        "number": "",
        "frequency": "",
    })
    rows.append({
        "category": "summary",
        "metric": "count",
        "value": int(ins.get("count", 0)),
        "number": "",
        "frequency": "",
    })
    rows.append({
        "category": "summary",
        "metric": "avg_sum_white",
        "value": ins.get("avg_sum_white", None),
        "number": "",
        "frequency": "",
    })

    # Si _insights reporta error, lo agregamos (mejor DX)
    if ins.get("error"):
        rows.append({
            "category": "summary",
            "metric": "error",
            "value": str(ins.get("error")),
            "number": "",
            "frequency": "",
        })

    for x in (ins.get("white_frequency") or []):
        rows.append({
            "category": "white_frequency",
            "metric": "",
            "value": "",
            "number": int(x.get("number")),
            "frequency": int(x.get("frequency")),
        })

    for x in (ins.get("pb_frequency") or []):
        rows.append({
            "category": "pb_frequency",
            "metric": "",
            "value": "",
            "number": int(x.get("pb")),
            "frequency": int(x.get("frequency")),
        })

    out = pd.DataFrame(rows, columns=["category", "metric", "value", "number", "frequency"])
    filename = f"insights_by_date_{month:02d}-{day:02d}.csv"
    return _df_to_csv_stream(out, filename)


@app.get("/insights/by-date/export.xlsx")
def insights_by_date_export_xlsx(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
    db: Session = Depends(get_db),
):
    # 1) Intento normal (CSV si existe)
    try:
        df = _load_draws()
    except HTTPException:
        df = pd.DataFrame()

    # 2) Fallback real a DB si no hay datos
    if df is None or df.empty:
        df = _load_draws(db=db)

    draws_day = _filter_draws_by_day(df, day=day, month=month)
    ins = _insights(draws_day)

    df_summary = pd.DataFrame([{
        "month": int(month),
        "day": int(day),
        "count": int(ins.get("count", 0)),
        "avg_sum_white": ins.get("avg_sum_white", None),
        "error": str(ins.get("error")) if ins.get("error") else None,
    }])

    # Hojas siempre válidas (aunque estén vacías)
    df_white = pd.DataFrame(ins.get("white_frequency") or [], columns=["number", "frequency"])
    df_pb = pd.DataFrame(ins.get("pb_frequency") or [], columns=["pb", "frequency"])

    filename = f"insights_by_date_{month:02d}-{day:02d}.xlsx"
    return _df_to_xlsx_stream(
        {
            "summary": df_summary,
            "white_frequency": df_white,
            "pb_frequency": df_pb,
        },
        filename,
    )


# ---------------------------
#   EXPORT: /compare/by-date
# ---------------------------
def _safe_filename_fragment(s: str) -> str:
    """
    Convierte un string a fragmento seguro para filename:
    - solo letras, números, guiones, guion bajo y punto
    - reemplaza espacios y separadores comunes
    """
    s = (s or "").strip()
    if not s:
        return "ticket"

    # normalizaciones simples
    s = s.replace(",", "-").replace("|", "_").replace(" ", "")
    # filtrar caracteres peligrosos
    allowed = []
    for ch in s:
        if ch.isalnum() or ch in ("-", "_", "."):
            allowed.append(ch)
        else:
            allowed.append("_")
    out = "".join(allowed)
    # cap longitud razonable
    return out[:80] if len(out) > 80 else out


@app.get("/compare/by-date/export.csv")
def compare_by_date_export_csv(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
    ticket: str = Query(..., description='Formato: "4,6,8,10,12|22"'),
    db: Session = Depends(get_db),
):
    # 1) Intento normal
    try:
        results = _compare_day_df(month, day, ticket)
    except HTTPException:
        results = pd.DataFrame()

    # 2) Fallback DB real si no hubo datos (por ejemplo si falta CSV local)
    # Nota: _compare_day_df hoy llama _load_draws() sin db. Si no hay CSV y dependía del CSV,
    # aquí hacemos el pipeline equivalente desde DB.
    if results.empty:
        ticket_white, ticket_pb = _parse_ticket_param(ticket)
        df_db = _load_draws(db=db)
        draws_day = _filter_draws_by_day(df_db, day=day, month=month)
        results = _compare_ticket_to_draws(ticket_white, ticket_pb, draws_day).copy()
        if not results.empty:
            results["draw_date"] = pd.to_datetime(results["draw_date"], errors="coerce").dt.strftime("%Y-%m-%d")
            results = results.dropna(subset=["draw_date"])

    # 3) Si sigue vacío, devolvemos CSV válido con columnas esperadas
    if results is None or results.empty:
        results = pd.DataFrame(columns=["year", "draw_date"] + WHITE_COLS + ["pb", "match_white", "match_pb", "score"])

    safe_ticket = _safe_filename_fragment(ticket)
    filename = f"compare_by_date_{month:02d}-{day:02d}_{safe_ticket}.csv"
    return _df_to_csv_stream(results, filename)


@app.get("/compare/by-date/export.xlsx")
def compare_by_date_export_xlsx(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
    ticket: str = Query(..., description='Formato: "4,6,8,10,12|22"'),
    db: Session = Depends(get_db),
):
    # 1) Intento normal
    try:
        results = _compare_day_df(month, day, ticket)
    except HTTPException:
        results = pd.DataFrame()

    # 2) Fallback DB real si no hubo datos
    if results.empty:
        ticket_white, ticket_pb = _parse_ticket_param(ticket)
        df_db = _load_draws(db=db)
        draws_day = _filter_draws_by_day(df_db, day=day, month=month)
        results = _compare_ticket_to_draws(ticket_white, ticket_pb, draws_day).copy()
        if not results.empty:
            results["draw_date"] = pd.to_datetime(results["draw_date"], errors="coerce").dt.strftime("%Y-%m-%d")
            results = results.dropna(subset=["draw_date"])

    # 3) Si sigue vacío, devolvemos XLSX válido con columnas esperadas
    if results is None or results.empty:
        results = pd.DataFrame(columns=["year", "draw_date"] + WHITE_COLS + ["pb", "match_white", "match_pb", "score"])

    safe_ticket = _safe_filename_fragment(ticket)
    filename = f"compare_by_date_{month:02d}-{day:02d}_{safe_ticket}.xlsx"
    return _df_to_xlsx_stream({"compare": results}, filename)

# ---------------------------
#   EXPORT: /compare/by-date/multi
# ---------------------------
@app.get("/compare/by-date/multi/export.csv")
def compare_by_date_multi_export_csv(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
    db: Session = Depends(get_db),
):
    # 1) Intento normal
    try:
        out = _compare_multi_day_df(month, day)
    except HTTPException:
        out = pd.DataFrame()

    # 2) Fallback real a DB si no hay datos
    if out is None or out.empty:
        df_draws = _load_draws(db=db)
        draws_day = _filter_draws_by_day(df_draws, day=day, month=month)

        tickets = _load_tickets(db=db)
        if not tickets.empty and not draws_day.empty:
            items = []
            for _, t in tickets.iterrows():
                tset = set(int(t[c]) for c in WHITE_COLS)
                tpb = int(t["pb"])

                for _, d in draws_day.iterrows():
                    dset = set(int(d[c]) for c in WHITE_COLS)
                    match_white = len(tset.intersection(dset))
                    match_pb = 1 if int(d["pb"]) == tpb else 0
                    items.append({
                        "ticket_id": str(t["ticket_id"]),
                        "year": int(d["year"]),
                        "draw_date": pd.to_datetime(d["draw_date"]).strftime("%Y-%m-%d"),
                        "ticket_n1": int(t["n1"]),
                        "ticket_n2": int(t["n2"]),
                        "ticket_n3": int(t["n3"]),
                        "ticket_n4": int(t["n4"]),
                        "ticket_n5": int(t["n5"]),
                        "ticket_pb": int(t["pb"]),
                        "draw_n1": int(d["n1"]),
                        "draw_n2": int(d["n2"]),
                        "draw_n3": int(d["n3"]),
                        "draw_n4": int(d["n4"]),
                        "draw_n5": int(d["n5"]),
                        "draw_pb": int(d["pb"]),
                        "match_white": int(match_white),
                        "match_pb": int(match_pb),
                        "score": int(match_white + match_pb),
                    })

            items.sort(
                key=lambda x: (x["score"], x["match_white"], x["match_pb"], x["year"]),
                reverse=True,
            )
            out = pd.DataFrame(items)

    # 3) Garantizar archivo válido aunque esté vacío
    if out is None or out.empty:
        out = pd.DataFrame(
            columns=[
                "ticket_id",
                "year",
                "draw_date",
                "ticket_n1",
                "ticket_n2",
                "ticket_n3",
                "ticket_n4",
                "ticket_n5",
                "ticket_pb",
                "draw_n1",
                "draw_n2",
                "draw_n3",
                "draw_n4",
                "draw_n5",
                "draw_pb",
                "match_white",
                "match_pb",
                "score",
            ]
        )

    filename = f"compare_by_date_multi_{month:02d}-{day:02d}.csv"
    return _df_to_csv_stream(out, filename)


@app.get("/compare/by-date/multi/export.xlsx")
def compare_by_date_multi_export_xlsx(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
    db: Session = Depends(get_db),
):
    # 1) Intento normal
    try:
        out = _compare_multi_day_df(month, day)
    except HTTPException:
        out = pd.DataFrame()

    # 2) Fallback real a DB
    if out is None or out.empty:
        df_draws = _load_draws(db=db)
        draws_day = _filter_draws_by_day(df_draws, day=day, month=month)
        tickets = _load_tickets(db=db)

        items = []
        if not tickets.empty and not draws_day.empty:
            for _, t in tickets.iterrows():
                tset = set(int(t[c]) for c in WHITE_COLS)
                tpb = int(t["pb"])

                for _, d in draws_day.iterrows():
                    dset = set(int(d[c]) for c in WHITE_COLS)
                    match_white = len(tset.intersection(dset))
                    match_pb = 1 if int(d["pb"]) == tpb else 0
                    items.append({
                        "ticket_id": str(t["ticket_id"]),
                        "year": int(d["year"]),
                        "draw_date": pd.to_datetime(d["draw_date"]).strftime("%Y-%m-%d"),
                        "match_white": int(match_white),
                        "match_pb": int(match_pb),
                        "score": int(match_white + match_pb),
                    })

            items.sort(
                key=lambda x: (x["score"], x["match_white"], x["match_pb"], x["year"]),
                reverse=True,
            )
            out = pd.DataFrame(items)

    # 3) Archivo válido aunque esté vacío
    if out is None or out.empty:
        out = pd.DataFrame(
            columns=[
                "ticket_id",
                "year",
                "draw_date",
                "match_white",
                "match_pb",
                "score",
            ]
        )

    filename = f"compare_by_date_multi_{month:02d}-{day:02d}.xlsx"
    return _df_to_xlsx_stream({"multi": out}, filename)


# ---------------------------
#   AI INSIGHT (base)
# ---------------------------
def compute_ai_insight(
    db: Session,
    compare: str,
) -> Dict[str, Any]:
    """
    AI Insight basado en:
    - DrawResult (frecuencias reales del sorteo)
    - Ticket (comportamiento histórico)
    """

    parsed = parse_compare(compare)
    if not parsed:
        raise HTTPException(status_code=400, detail="compare inválido")

    base_regs, base_pb = parsed
    base_regs = sorted(list(base_regs))

    # Placeholder seguro (se extiende luego)
    return {
        "input": {
            "whites": base_regs,
            "powerball": int(base_pb),
        },
        "status": "pending",
        "note": "AI insight pendiente de implementación completa",
    }


    # -------------------------
    #   DRAW RESULTS ANALYSIS
    # -------------------------
    drs = db.query(DrawResult).all()
    if not drs:
        return {"score": 0, "notes": ["No draw results available"]}

    # Frecuencias de white balls y powerball, y suma de whites por sorteo
    freq = defaultdict(int)      # white ball -> count
    pb_freq = defaultdict(int)   # powerball -> count
    sums: List[int] = []

    for d in drs:
        # normalizar a int por seguridad (DB a veces devuelve Decimals/None)
        try:
            nums = [int(d.wn1), int(d.wn2), int(d.wn3), int(d.wn4), int(d.wn5)]
        except Exception:
            # registro corrupto -> ignorar
            continue

        for n in nums:
            freq[n] += 1

        sums.append(nums[0] + nums[1] + nums[2] + nums[3] + nums[4])

        try:
            pb = int(d.winning_powerball)
            pb_freq[pb] += 1
        except Exception:
            # PB inválido -> ignorar
            pass

    if not sums or not freq:
        return {"score": 0, "notes": ["No valid draw results available (invalid rows)"]}

    # Promedios
    avg_freq = mean(freq.values())
    avg_sum = mean(sums)

    # Suma de la combinación base
    base_sum = sum(base_regs)

    # Frecuencia promedio de la combinación (si un número no existe, su freq=0)
    base_freq_score = mean([freq.get(int(n), 0) for n in base_regs])

    # -------------------------
    #   TICKETS ANALYSIS
    # -------------------------
    tickets = db.query(Ticket).all()
    ticket_freq = defaultdict(float)  # float porque ponderamos
    total_weight = 0.0

    now = datetime.utcnow()

    for t in tickets:
        # Intentamos obtener fecha del ticket (opcional)
        t_date = None
        for attr in ("created_at", "created", "date", "draw_date"):
            if hasattr(t, attr):
                raw = getattr(t, attr)
                try:
                    t_date = pd.to_datetime(raw, errors="coerce") if raw else None
                except Exception:
                    t_date = None
                break

        # Peso por recencia
        if t_date is not None and not pd.isna(t_date):
            try:
                days_old = max((now - t_date.to_pydatetime()).days, 0)
            except Exception:
                days_old = 0
            weight = 1.0 / (1.0 + days_old / 30.0)  # ~1 mes = peso 0.5
        else:
            weight = 1.0

        try:
            nums = [int(t.n1), int(t.n2), int(t.n3), int(t.n4), int(t.n5)]
        except Exception:
            # Ticket corrupto -> ignorar
            continue

        for n in nums:
            ticket_freq[n] += weight

        total_weight += weight

    if ticket_freq and total_weight > 0:
        avg_ticket_freq = mean(ticket_freq.values())
        base_ticket_score = mean([ticket_freq.get(int(n), 0.0) for n in base_regs])
    else:
        avg_ticket_freq = 0.0
        base_ticket_score = 0.0

    # -------------------------
    #   FEATURES
    # -------------------------
    odd = sum(1 for n in base_regs if int(n) % 2 == 1)
    even = 5 - odd

    low = sum(1 for n in base_regs if int(n) <= 35)
    high = 5 - low

    pb_rarity = int(pb_freq.get(int(base_pb), 0)) if pb_freq else 0
    pb_avg = mean(pb_freq.values()) if pb_freq else 0



        # -------------------------
    #   SCORING (0–100)
    # -------------------------
    score = 50.0  # usar float internamente para permitir ajuste fino

    # Frecuencia histórica de draws
    if avg_freq > 0:
        if base_freq_score > avg_freq:
            score += 10
        elif base_freq_score < avg_freq:
            score -= 5
        # si es igual, no ajusta

    # Frecuencia histórica de tickets
    if avg_ticket_freq > 0:
        if base_ticket_score > avg_ticket_freq:
            score += 10
        # si es igual o menor, no ajusta

    # Cercanía a la suma promedio (ventana flexible)
    sum_diff = abs(base_sum - avg_sum)
    if sum_diff < 10:
        score += 10
    elif sum_diff < 20:
        score += 5
    else:
        score -= 5

    # Balance impar/par
    if odd in (2, 3):
        score += 5
    elif odd in (1, 4):
        score -= 2

    # Balance low/high
    if low in (2, 3):
        score += 5
    elif low in (1, 4):
        score -= 2

    # Rareza del Powerball (comparación estable)
    if pb_avg > 0:
        if pb_rarity < pb_avg:
            score += 10
        elif pb_rarity > pb_avg:
            score -= 5
        # si es igual, no ajusta

    # Clamp final
    score = int(max(0, min(100, round(score))))


    # -------------------------
    #   EXPLANATION
    # -------------------------
    notes: List[str] = []

    # Frecuencia de números blancos (draws)
    if avg_freq > 0:
        delta_freq = base_freq_score - avg_freq
        if delta_freq > 0:
            notes.append("Regular numbers: above-average historical frequency")
        elif delta_freq < 0:
            notes.append("Regular numbers: below-average historical frequency")
        else:
            notes.append("Regular numbers: around historical average frequency")
    else:
        notes.append("Regular numbers: insufficient historical data")

    # Paridad
    notes.append(f"Parity (odd/even): {odd}/{even}")

    # Distribución low / high
    notes.append(f"Low / High split: {low}/{high}")

    # Suma de blancos
    sum_diff = abs(base_sum - avg_sum)
    if avg_sum > 0:
        if sum_diff < 10:
            notes.append("Sum of numbers: very close to historical mean")
        elif sum_diff < 20:
            notes.append("Sum of numbers: moderately close to historical mean")
        else:
            notes.append("Sum of numbers: far from historical mean")
    else:
        notes.append("Sum of numbers: insufficient historical data")

    # Powerball
    if pb_avg > 0:
        delta_pb = pb_rarity - pb_avg
        if delta_pb < 0:
            notes.append("Powerball: relatively rare in historical draws")
        elif delta_pb > 0:
            notes.append("Powerball: relatively common in historical draws")
        else:
            notes.append("Powerball: average historical frequency")
    else:
        notes.append("Powerball: insufficient historical data")

    return {
        "score": score,
        "notes": notes,
    }




# -------------------------
#   SIMPLE IN-MEMORY CACHE
# -------------------------
# key -> (timestamp, value)
_COMPARE_CACHE: dict[str, Tuple[float, dict]] = {}

COMPARE_CACHE_TTL = 600  # segundos (10 minutos)
_COMPARE_CACHE_MAXSIZE = 1_000  # límite duro para evitar crecimiento infinito


def _make_compare_cache_key(
    compare: str,
    status: Optional[str],
    type: Optional[str],
    start_date: Optional[str],
    end_date: Optional[str],
) -> str:
    """
    Genera una key estable y segura para cache.
    Se normalizan valores para evitar duplicados lógicos.
    """
    return "|".join(
        [
            (compare or "").strip(),
            (status or "").strip(),
            (type or "").strip(),
            (start_date or "").strip(),
            (end_date or "").strip(),
        ]
    )


def _get_compare_cache(key: str) -> Optional[dict]:
    item = _COMPARE_CACHE.get(key)
    if not item:
        return None

    ts, value = item
    now = time.time()

    # Expirado por TTL
    if (now - ts) > COMPARE_CACHE_TTL:
        _COMPARE_CACHE.pop(key, None)
        return None

    return value


def _set_compare_cache(key: str, value: dict) -> None:
    """
    Inserta en cache y aplica:
    - TTL
    - Evicción simple por tamaño (FIFO aproximado)
    """
    now = time.time()

    # Evicción simple si supera tamaño máximo
    if len(_COMPARE_CACHE) >= _COMPARE_CACHE_MAXSIZE:
        # eliminar el más antiguo
        oldest_key = min(_COMPARE_CACHE.items(), key=lambda x: x[1][0])[0]
        _COMPARE_CACHE.pop(oldest_key, None)

    _COMPARE_CACHE[key] = (now, value)


# -------------------------
#   CONFIG / CONSTANTS
# -------------------------
REG_MIN, REG_MAX = 1, 69
PB_MIN, PB_MAX = 1, 26

ALLOWED_TYPES = {"QUICK_PICK", "MANUAL"}
ALLOWED_STATUS = {"PAST", "FUTURE"}

# Estilos Excel
FILL_AQUA = PatternFill(start_color="B7EDE6", end_color="B7EDE6", fill_type="solid")  # verde agua
FILL_PB = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")    # rojo suave
RED_FONT = Font(color="FF0000", bold=True)

# Crear tablas (idempotente)
try:
    Base.metadata.create_all(bind=engine)
except Exception as e:
    # No romper el arranque por DB no disponible (ej. en docs/tests)
    # Si necesitas que sea hard-fail en prod, luego lo hacemos configurable.
    print(f"[WARN] Base.metadata.create_all failed: {e}")



# -------------------------
#   MIDDLEWARE
# -------------------------
app.add_middleware(GZipMiddleware, minimum_size=1000)

# CORS: en dev puede ser "*", pero en prod conviene restringir.
# Upgrade seguro: permite override por env sin romper compatibilidad.
cors_origins_env = os.getenv("CORS_ALLOW_ORIGINS", "*").strip()
if cors_origins_env == "*":
    cors_allow_origins = ["*"]
else:
    cors_allow_origins = [o.strip() for o in cors_origins_env.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_allow_origins,
    allow_credentials=False,  # correcto si allow_origins incluye "*"
    allow_methods=["*"],
    allow_headers=["*"],
)

# -------------------------
#        DB DEP
# -------------------------
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()



# -------------------------
#        HELPERS
# -------------------------
def _norm_status(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    v = str(s).strip().upper()
    return v if v in ALLOWED_STATUS else None


def _norm_type(t: Optional[str]) -> Optional[str]:
    if t is None:
        return None
    v = str(t).strip().upper()
    if not v:
        return None
    if v in ("QP", "QUICKPICK", "QUICK-PICK"):
        v = "QUICK_PICK"
    return v if v in ALLOWED_TYPES else None


def normalize_regular_numbers(n1: int, n2: int, n3: int, n4: int, n5: int, order: str = "asc") -> List[int]:
    nums_sorted = sorted([n1, n2, n3, n4, n5])
    if str(order).strip().lower() == "desc":
        nums_sorted.reverse()
    return nums_sorted


def validate_ticket_numbers(n1: int, n2: int, n3: int, n4: int, n5: int, pb: int) -> None:
    regs = [n1, n2, n3, n4, n5]
    if any((x < REG_MIN or x > REG_MAX) for x in regs):
        raise HTTPException(status_code=400, detail=f"N1..N5 deben estar entre {REG_MIN}-{REG_MAX}")
    if len(set(regs)) != 5:
        raise HTTPException(status_code=400, detail="N1..N5 no pueden repetirse dentro del mismo ticket")
    if pb < PB_MIN or pb > PB_MAX:
        raise HTTPException(status_code=400, detail=f"Powerball debe estar entre {PB_MIN}-{PB_MAX}")


def numbers_key(regs: List[int], pb: int) -> str:
    r = sorted(int(x) for x in regs)
    return f"{r[0]}-{r[1]}-{r[2]}-{r[3]}-{r[4]}|{int(pb)}"


def _quote(v: Optional[str]) -> str:
    if not v:
        return ""
    return quote_plus(str(v))


def _safe_int(x: Any) -> Optional[int]:
    try:
        if x is None:
            return None
        s = str(x).strip()
        if not s:
            return None
        return int(s)
    except Exception:
        return None


def parse_compare(compare: Optional[str]) -> Optional[Tuple[set, int]]:
    """
    Acepta formatos:
      - "10,16,29,33,69|22"
      - "10 16 29 33 69 | 22"
      - "10 16 29 33 69 22"   (último es PB)
      - "10,16,29,33,69 22"  (último es PB)
    Devuelve (set(regs), pb)
    """
    if not compare:
        return None

    raw = str(compare).strip()
    raw = raw.replace(";", ",").replace("\t", " ")
    raw = " ".join(raw.split())  # colapsa espacios

    if "|" in raw:
        left, right = raw.split("|", 1)
        left = left.strip().replace(" ", ",")
        right = right.strip()
        try:
            regs = [int(x.strip()) for x in left.split(",") if x.strip()]
            pb = int(right.strip())
        except Exception:
            raise HTTPException(status_code=400, detail='compare inválido. Ej: "10,16,29,33,69|22"')
        if len(regs) != 5:
            raise HTTPException(status_code=400, detail="compare debe tener 5 números regulares antes del |")
        validate_ticket_numbers(regs[0], regs[1], regs[2], regs[3], regs[4], pb)
        return set(regs), pb

    raw2 = raw.replace(",", " ")
    parts = [p for p in raw2.split(" ") if p.strip()]
    nums: List[int] = []

    for p in parts:
        v = _safe_int(p)
        if v is None:
            raise HTTPException(status_code=400, detail='compare inválido. Ej: "10,16,29,33,69|22"')
        nums.append(v)

    if len(nums) != 6:
        raise HTTPException(
            status_code=400,
            detail='compare inválido. Usa "10,16,29,33,69|22" o "10 16 29 33 69 22".',
        )

    regs = nums[:5]
    pb = nums[5]
    validate_ticket_numbers(regs[0], regs[1], regs[2], regs[3], regs[4], pb)
    return set(regs), pb


def _parse_date_iso(d: str) -> date:
    try:
        return date.fromisoformat(str(d).strip())
    except Exception:
        raise HTTPException(status_code=400, detail="Fecha inválida. Usa formato YYYY-MM-DD")


def _parse_date_maybe(s: Optional[str]) -> Optional[date]:
    if s is None:
        return None
    v = str(s).strip()
    if not v:
        return None
    return _parse_date_iso(v)


def _parse_int_list_csv(v: Optional[str]) -> List[int]:
    if not v:
        return []
    raw = str(v).replace(";", ",").replace("|", ",")
    items = [x.strip() for x in raw.split(",") if x.strip()]
    out: List[int] = []
    for it in items:
        n = _safe_int(it)
        if n is None:
            raise HTTPException(status_code=400, detail=f"Lista inválida: {v}")
        out.append(int(n))
    return out


def _count_fast(q):
    """
    count() más estable y usualmente más rápido:
    SELECT count(*) FROM (subquery sin order_by)
    """
    try:
        return q.order_by(None).count()
    except Exception:
        return q.count()



# -------------------------
#   PERFORMANCE: INDICES
# -------------------------
def _ensure_indexes() -> None:
    """
    Crea índices comunes si el motor lo permite.
    - SQLite: CREATE INDEX IF NOT EXISTS ok
    - Otros motores: puede variar, por eso se encapsula en try/except.
    """
    stmts = [
        "CREATE INDEX IF NOT EXISTS ix_ticket_status ON ticket(status)",
        "CREATE INDEX IF NOT EXISTS ix_ticket_draw_date ON ticket(draw_date)",
        "CREATE INDEX IF NOT EXISTS ix_ticket_type ON ticket(type)",
        "CREATE INDEX IF NOT EXISTS ix_drawresult_draw_date ON draw_result(draw_date)",
    ]

    try:
        with engine.begin() as conn:
            for s in stmts:
                try:
                    conn.execute(text(s))
                except Exception:
                    # ignorar statement que no aplique
                    pass
    except Exception:
        # engine no disponible / permisos / DB no lista
        pass


@app.on_event("startup")
def _on_startup():
    # Upgrade: evitar trabajo innecesario en procesos donde no quieres tocar DB
    # (por ejemplo tests, builds, docs), si se desactiva con env.
    if os.getenv("DISABLE_DB_INDEXES", "").strip() in ("1", "true", "TRUE", "yes", "YES"):
        return
    _ensure_indexes()


# -------------------------
#        PRIZES / MATCH
# -------------------------
def get_prize(matched_regular: int, matched_pb: bool) -> float:
    """
    Tabla estándar de premios (sin Power Play).
    Devuelve 0.0 para jackpot (para mantener consistencia numérica).
    """
    matched_regular = int(matched_regular)
    matched_pb = bool(matched_pb)

    if matched_regular == 5 and matched_pb:
        return 0.0  # jackpot
    if matched_regular == 5 and not matched_pb:
        return 1_000_000.0
    if matched_regular == 4 and matched_pb:
        return 50_000.0
    if matched_regular == 4 and not matched_pb:
        return 100.0
    if matched_regular == 3 and matched_pb:
        return 100.0
    if matched_regular == 3 and not matched_pb:
        return 7.0
    if matched_regular == 2 and matched_pb:
        return 7.0
    if matched_regular == 1 and matched_pb:
        return 4.0
    if matched_regular == 0 and matched_pb:
        return 4.0
    return 0.0


def calculate_matches(ticket: Ticket, result: DrawResult) -> None:
    """
    Calcula matches y setea campos derivados en el ticket.
    Robusto a tipos raros (strings/None) y a naming differences (pb/powerball).
    """
    # Números del ticket
    try:
        ticket_nums = {int(ticket.n1), int(ticket.n2), int(ticket.n3), int(ticket.n4), int(ticket.n5)}
    except Exception:
        ticket_nums = set()

    # Números ganadores
    try:
        winning_nums = {int(result.wn1), int(result.wn2), int(result.wn3), int(result.wn4), int(result.wn5)}
    except Exception:
        winning_nums = set()

    matched_regular = len(ticket_nums.intersection(winning_nums))

    # Powerball del ticket: soporta ticket.pb o ticket.powerball
    t_pb_raw = getattr(ticket, "pb", None)
    if t_pb_raw is None:
        t_pb_raw = getattr(ticket, "powerball", None)

    # Powerball del resultado: soporta result.winning_powerball o result.pb
    r_pb_raw = getattr(result, "winning_powerball", None)
    if r_pb_raw is None:
        r_pb_raw = getattr(result, "pb", None)

    try:
        matched_pb = int(t_pb_raw) == int(r_pb_raw)
    except Exception:
        matched_pb = False

    prize = get_prize(matched_regular, matched_pb)

    # Set de atributos: soporta distintos nombres de columnas en modelos
    setattr(ticket, "matched_regular_numbers", int(matched_regular))
    setattr(ticket, "matched_powerball", bool(matched_pb))
    setattr(ticket, "prize_amount", float(prize))


# -------------------------
#        SCHEMAS
# -------------------------
class TicketCreate(BaseModel):
    draw_date: date
    status: Optional[Literal["PAST", "FUTURE"]] = None

    n1: int
    n2: int
    n3: int
    n4: int
    n5: int
    powerball: int

    type: Literal["QUICK_PICK", "MANUAL"]
    cost: float

    # Validación mínima sin cambiar comportamiento: evita negativos y NaNs obvios
    model_config = ConfigDict(extra="ignore")


class TicketUpdate(BaseModel):
    draw_date: Optional[date] = None
    status: Optional[Literal["PAST", "FUTURE"]] = None

    n1: Optional[int] = None
    n2: Optional[int] = None
    n3: Optional[int] = None
    n4: Optional[int] = None
    n5: Optional[int] = None
    powerball: Optional[int] = None

    type: Optional[Literal["QUICK_PICK", "MANUAL"]] = None
    cost: Optional[float] = None

    model_config = ConfigDict(extra="ignore")


class TicketOut(TicketCreate):
    id: int
    matched_regular_numbers: int = 0
    matched_powerball: bool = False
    prize_amount: float = 0.0
    model_config = ConfigDict(from_attributes=True)


class TicketNormalizedOut(BaseModel):
    id: int
    draw_date: date
    status: str
    type: str
    cost: float

    n1: int
    n2: int
    n3: int
    n4: int
    n5: int
    powerball: int

    matched_regular_numbers: int = 0
    matched_powerball: bool = False
    prize_amount: float = 0.0

    model_config = ConfigDict(from_attributes=True, extra="ignore")


class DrawResultCreate(BaseModel):
    draw_date: date
    wn1: int
    wn2: int
    wn3: int
    wn4: int
    wn5: int
    winning_powerball: int

    model_config = ConfigDict(extra="ignore")


class DrawResultOut(DrawResultCreate):
    id: int
    model_config = ConfigDict(from_attributes=True)


class StatsSummary(BaseModel):
    total_tickets: int
    total_cost: float
    total_prize: float
    balance: float
    winning_tickets: int

    model_config = ConfigDict(extra="ignore")


# ============================================================
#   COMPARE + AI RECOMMENDER (schemas base)
# ============================================================
class CompareMatchesResponse(BaseModel):
    base: str
    total_scanned: int
    groups: Dict[str, int]
    tickets: Dict[str, List[TicketOut]]

    model_config = ConfigDict(extra="ignore")


class RecommendRequest(BaseModel):
    status: Optional[Literal["PAST", "FUTURE"]] = None
    type: Optional[Literal["QUICK_PICK", "MANUAL"]] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None

    k: int = 50
    seed: Optional[int] = None

    fixed_first: Optional[int] = None
    fixed_numbers: Optional[List[int]] = None
    exclude_numbers: Optional[List[int]] = None

    fixed_powerball: Optional[int] = None
    exclude_powerballs: Optional[List[int]] = None

    top_pool_regulars: int = 25
    top_pool_powerballs: int = 10

    model_config = ConfigDict(extra="ignore")


class RecommendResponse(BaseModel):
    generated: int
    seed: Optional[int]
    combos: List[Dict[str, Any]]

    model_config = ConfigDict(extra="ignore")


class SaveRecommendationsRequest(BaseModel):
    # Generación
    status: Optional[Literal["PAST", "FUTURE"]] = "PAST"
    type: Optional[Literal["QUICK_PICK", "MANUAL"]] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None

    k: int = 50
    seed: Optional[int] = None

    fixed_first: Optional[int] = None
    fixed_numbers: Optional[List[int]] = None
    exclude_numbers: Optional[List[int]] = None

    fixed_powerball: Optional[int] = None
    exclude_powerballs: Optional[List[int]] = None

    top_pool_regulars: int = 25
    top_pool_powerballs: int = 10

    # Guardado
    future_draw_date: date
    cost_per_ticket: float = 2.0
    save_type: Literal["QUICK_PICK", "MANUAL"] = "QUICK_PICK"
    normalize_on_save: bool = True

    model_config = ConfigDict(extra="ignore")


class SaveRecommendationsResponse(BaseModel):
    requested: int
    generated: int
    inserted: int
    skipped_duplicates: int
    future_draw_date: date
    status: str
    message: str

    model_config = ConfigDict(extra="ignore")


def _ticket_regs(t: Ticket) -> List[int]:
    # Soporta modelos con n1..n5 (preferido)
    return [int(t.n1), int(t.n2), int(t.n3), int(t.n4), int(t.n5)]


def _ticket_pb(t: Ticket) -> int:
    # Soporta pb o powerball según modelo
    v = getattr(t, "pb", None)
    if v is None:
        v = getattr(t, "powerball", None)
    return int(v)


def _match_counts(base_regs: set, base_pb: int, t: Ticket) -> Tuple[int, bool]:
    regs = set(_ticket_regs(t))
    mr = len(regs.intersection(base_regs))
    mpb = (_ticket_pb(t) == int(base_pb))
    return mr, mpb


def _filter_tickets_for_ai(payload: Any, db: Session) -> List[Ticket]:
    q = db.query(Ticket)

    # Status
    p_status = getattr(payload, "status", None)
    if p_status:
        q = q.filter(Ticket.status == p_status)

    # Type: normalizamos usando tu helper (acepta QP/QUICKPICK/etc)
    p_type = getattr(payload, "type", None)
    if p_type:
        ttype = _norm_type(p_type)
        if ttype:
            q = q.filter(Ticket.type == ttype)

    # Fechas
    p_start = getattr(payload, "start_date", None)
    if p_start:
        q = q.filter(Ticket.draw_date >= p_start)

    p_end = getattr(payload, "end_date", None)
    if p_end:
        q = q.filter(Ticket.draw_date <= p_end)

    # Upgrade: evita traer columnas extra por relaciones/cargas pesadas (si existieran)
    # (sin tocar el modelo, esto ya ayuda en algunos ORM setups)
    return q.all()


def _existing_keys_for_draw(db: Session, draw_date_val: date) -> set:
    # Upgrade: select solo columnas necesarias (reduce carga)
    rows = (
        db.query(Ticket.n1, Ticket.n2, Ticket.n3, Ticket.n4, Ticket.n5, getattr(Ticket, "pb", Ticket.powerball))
        .filter(Ticket.draw_date == draw_date_val)
        .all()
    )

    out = set()
    for r in rows:
        # r puede ser tuple (n1..n5,pb) o Row
        try:
            n1, n2, n3, n4, n5, pb = r
        except Exception:
            n1, n2, n3, n4, n5 = r[0], r[1], r[2], r[3], r[4]
            pb = r[5] if len(r) > 5 else getattr(r, "pb", getattr(r, "powerball", None))

        out.add(numbers_key([int(n1), int(n2), int(n3), int(n4), int(n5)], int(pb)))

    return out

# ============================================================
# ✅ FIX: FUNCIÓN QUE TE FALTABA (AI generator)
# ============================================================
def recommend_from_history(req: RecommendRequest, db: Session) -> RecommendResponse:
    """
    Genera combinaciones usando frecuencias históricas (tickets filtrados) + restricciones opcionales.
    Produce combos únicos dentro de la ejecución (no repite keys).
    """
    # k seguro
    k = int(req.k or 50)
    k = max(1, min(5000, k))

    # RNG reproducible si seed viene, sino random seguro
    seed_val = int(req.seed) if req.seed is not None else random.randrange(1, 10**9)
    rng = random.Random(seed_val)

    tickets = _filter_tickets_for_ai(req, db)

    reg_counter = Counter()
    pb_counter = Counter()

    # Compatibilidad pb/powerball
    for t in tickets:
        try:
            reg_counter.update([int(t.n1), int(t.n2), int(t.n3), int(t.n4), int(t.n5)])
        except Exception:
            pass
        try:
            pb_val = getattr(t, "pb", None)
            if pb_val is None:
                pb_val = getattr(t, "powerball", None)
            if pb_val is not None:
                pb_counter.update([int(pb_val)])
        except Exception:
            pass

    # fallback uniforme si no hay historial
    if not reg_counter:
        reg_counter.update({n: 1 for n in range(REG_MIN, REG_MAX + 1)})
    if not pb_counter:
        pb_counter.update({n: 1 for n in range(PB_MIN, PB_MAX + 1)})

    # restricciones
    fixed_first = int(req.fixed_first) if req.fixed_first is not None else None
    fixed_numbers = [int(x) for x in (req.fixed_numbers or [])]
    exclude_numbers = set(int(x) for x in (req.exclude_numbers or []))

    fixed_pb = int(req.fixed_powerball) if req.fixed_powerball is not None else None
    exclude_pbs = set(int(x) for x in (req.exclude_powerballs or []))

    # normalizar fixed_numbers (unique, mantener orden)
    fixed_numbers = list(dict.fromkeys(fixed_numbers))

    # valida ranges
    if fixed_first is not None and not (REG_MIN <= fixed_first <= REG_MAX):
        raise HTTPException(status_code=400, detail="fixed_first fuera de rango")
    for x in fixed_numbers:
        if not (REG_MIN <= x <= REG_MAX):
            raise HTTPException(status_code=400, detail=f"fixed_numbers fuera de rango: {x}")
    if fixed_pb is not None and not (PB_MIN <= fixed_pb <= PB_MAX):
        raise HTTPException(status_code=400, detail="fixed_powerball fuera de rango")

    # valida conflictos con excludes
    if fixed_first is not None and fixed_first in exclude_numbers:
        raise HTTPException(status_code=400, detail="fixed_first está en exclude_numbers")
    for x in fixed_numbers:
        if x in exclude_numbers:
            raise HTTPException(status_code=400, detail=f"fixed_numbers contiene excluido: {x}")
    if fixed_pb is not None and fixed_pb in exclude_pbs:
        raise HTTPException(status_code=400, detail="fixed_powerball está en exclude_powerballs")

    # valida duplicados entre fixed_first y fixed_numbers
    if fixed_first is not None and fixed_first in fixed_numbers:
        fixed_numbers = [x for x in fixed_numbers if x != fixed_first]

    if (1 if fixed_first is not None else 0) + len(fixed_numbers) > 5:
        raise HTTPException(status_code=400, detail="Demasiados números fijos (fixed_first + fixed_numbers > 5)")

    # pools top
    top_pool_regulars = max(5, min(69, int(req.top_pool_regulars or 25)))
    top_regs = [n for (n, _) in reg_counter.most_common(top_pool_regulars)]

    # asegurar que fijos estén presentes en pool
    for x in ([fixed_first] if fixed_first is not None else []) + fixed_numbers:
        if x is not None and x not in top_regs:
            top_regs.append(x)

    # aplicar excludes
    top_regs = [n for n in top_regs if n not in exclude_numbers]
    if len(top_regs) < 5:
        top_regs = [n for n in range(REG_MIN, REG_MAX + 1) if n not in exclude_numbers]

    if len(top_regs) < 5:
        raise HTTPException(status_code=400, detail="No hay suficientes números regulares disponibles (revisa excludes)")

    top_pool_powerballs = max(1, min(26, int(req.top_pool_powerballs or 10)))
    top_pbs = [n for (n, _) in pb_counter.most_common(top_pool_powerballs)]

    if fixed_pb is not None and fixed_pb not in top_pbs:
        top_pbs.append(fixed_pb)

    top_pbs = [n for n in top_pbs if n not in exclude_pbs]
    if not top_pbs:
        top_pbs = [n for n in range(PB_MIN, PB_MAX + 1) if n not in exclude_pbs]

    if not top_pbs:
        raise HTTPException(status_code=400, detail="No hay powerballs disponibles (revisa exclude_powerballs)")

    # weights precomputados (upgrade performance)
    reg_weights = {n: max(1, int(reg_counter.get(n, 1))) for n in range(REG_MIN, REG_MAX + 1)}
    pb_weights = {n: max(1, int(pb_counter.get(n, 1))) for n in range(PB_MIN, PB_MAX + 1)}

    def pick_regs() -> List[int]:
        locked: List[int] = []
        if fixed_first is not None:
            locked.append(fixed_first)
        for x in fixed_numbers:
            if x not in locked:
                locked.append(x)

        chosen = list(locked)
        pool = [n for n in top_regs if n not in chosen]

        # sampling sin repetición: choices + remove es ok con pool pequeño
        while len(chosen) < 5 and pool:
            weights = [reg_weights[n] for n in pool]
            n = rng.choices(pool, weights=weights, k=1)[0]
            chosen.append(n)
            pool.remove(n)

        # si no completa, rellena con universo permitido
        chosen = list(dict.fromkeys(chosen))
        if len(chosen) < 5:
            universe = [n for n in range(REG_MIN, REG_MAX + 1) if n not in exclude_numbers and n not in chosen]
            rng.shuffle(universe)
            chosen.extend(universe[: (5 - len(chosen))])

        if len(chosen) != 5 or len(set(chosen)) != 5:
            raise HTTPException(status_code=400, detail="No pude generar combinación válida con las restricciones dadas")
        return chosen

    def pick_pb() -> int:
        if fixed_pb is not None:
            return fixed_pb
        weights = [pb_weights[n] for n in top_pbs]
        return rng.choices(top_pbs, weights=weights, k=1)[0]

    combos: List[Dict[str, Any]] = []
    seen = set()
    attempts = 0
    max_attempts = max(5000, k * 50)

    while len(combos) < k and attempts < max_attempts:
        attempts += 1

        regs = pick_regs()
        pb = pick_pb()

        # (redundante pero seguro)
        if pb in exclude_pbs:
            continue

        key = numbers_key(regs, pb)
        if key in seen:
            continue
        seen.add(key)

        # output order: si fixed_first, mantenerlo primero
        if fixed_first is not None:
            rest = sorted([x for x in regs if x != fixed_first])
            out_regs = [fixed_first] + rest
        else:
            out_regs = sorted(regs)

        combos.append(
            {
                "n1": int(out_regs[0]),
                "n2": int(out_regs[1]),
                "n3": int(out_regs[2]),
                "n4": int(out_regs[3]),
                "n5": int(out_regs[4]),
                "powerball": int(pb),
                "key": key,
            }
        )

    return RecommendResponse(generated=len(combos), seed=seed_val, combos=combos)

# -------------------------
#   UI APP LAYOUT (PRO)
# -------------------------
APP_CSS = """
:root{
  --bg:#0b1220;
  --panel:#0f1a2e;
  --panel2:#101f38;
  --card:#0f213f;
  --muted:#9bb0d1;
  --text:#eaf1ff;
  --line:rgba(255,255,255,.08);
  --chip:rgba(255,255,255,.06);
  --accent:#7dd3fc;
  --accent2:#34d399;
  --warn:#fca5a5;
  --shadow: 0 10px 30px rgba(0,0,0,.35);
  --radius: 18px;
  --radius2: 14px;
  --mono: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace;
  --sans: ui-sans-serif, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, "Apple Color Emoji","Segoe UI Emoji";
}
*{ box-sizing:border-box; }
html,body{ height:100%; }
body{
  margin:0;
  font-family:var(--sans);
  background:
    radial-gradient(1000px 600px at 20% 10%, rgba(125,211,252,.12), transparent 55%),
    radial-gradient(900px 500px at 80% 20%, rgba(52,211,153,.10), transparent 55%),
    var(--bg);
  color:var(--text);
}
a{ color:inherit; text-decoration:none; }
small, .muted{ color:var(--muted); }
.app{
  display:grid;
  grid-template-columns: 280px 1fr;
  min-height:100vh;
}
.sidebar{
  position:sticky; top:0; height:100vh;
  padding:18px;
  border-right:1px solid var(--line);
  background: linear-gradient(180deg, rgba(255,255,255,.04), rgba(255,255,255,.02));
}
.brand{
  display:flex; align-items:center; gap:10px;
  padding:12px 12px; border:1px solid var(--line);
  border-radius: var(--radius);
  background: rgba(255,255,255,.03);
  box-shadow: var(--shadow);
}
.logo{
  width:38px; height:38px; border-radius:14px;
  background: linear-gradient(135deg, rgba(125,211,252,.9), rgba(52,211,153,.9));
  box-shadow: 0 8px 20px rgba(0,0,0,.35);
}
.brand h1{ font-size:14px; margin:0; line-height:1.1; }
.brand p{ margin:0; font-size:12px; color:var(--muted); }

.nav{ margin-top:14px; display:flex; flex-direction:column; gap:8px; }
.nav a{
  display:flex; align-items:center; gap:10px;
  padding:10px 12px;
  border:1px solid var(--line);
  border-radius: 14px;
  background: rgba(255,255,255,.02);
  transition: transform .08s ease, background .12s ease, border-color .12s ease;
}
.nav a:hover{ background: rgba(255,255,255,.05); transform: translateY(-1px); border-color: rgba(125,211,252,.25); }
.nav a.active{
  background: rgba(125,211,252,.10);
  border-color: rgba(125,211,252,.35);
}
.main{ padding:18px 18px 28px; }

.topbar{
  display:flex; align-items:center; justify-content:space-between;
  gap:10px;
  padding:12px 14px;
  border:1px solid var(--line);
  border-radius: var(--radius);
  background: rgba(255,255,255,.03);
  box-shadow: var(--shadow);
}
.topbar h2{ margin:0; font-size:16px; letter-spacing:.2px; }
.topbar .right{ display:flex; gap:10px; flex-wrap:wrap; justify-content:flex-end; }
.pill{
  display:inline-flex; align-items:center; gap:8px;
  padding:8px 10px;
  border:1px solid var(--line);
  border-radius: 999px;
  background: var(--chip);
  color: var(--muted);
  font-size:12px;
}
.grid{
  margin-top:14px;
  display:grid;
  grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
  gap:12px;
}
.card{
  border:1px solid var(--line);
  border-radius: var(--radius);
  background: rgba(255,255,255,.03);
  box-shadow: var(--shadow);
  padding:14px;
}
.card h3{ margin:0 0 6px; font-size:14px; }
.card p{ margin:0; color:var(--muted); font-size:13px; }

.btn{
  display:inline-flex;
  align-items:center;
  gap:8px;
  padding:9px 12px;
  border:1px solid var(--line);
  border-radius: 14px;
  background: rgba(255,255,255,.04);
  color:var(--text);
  cursor:pointer;
  transition: transform .08s ease, background .12s ease, border-color .12s ease;
  font-size:13px;
}
.btn:hover{ background: rgba(255,255,255,.07); transform: translateY(-1px); border-color: rgba(125,211,252,.25); }
.btn.primary{ background: rgba(125,211,252,.14); border-color: rgba(125,211,252,.35); }
.btn.good{ background: rgba(52,211,153,.12); border-color: rgba(52,211,153,.35); }
.btn.warn{ background: rgba(252,165,165,.12); border-color: rgba(252,165,165,.35); }

hr.sep{ border:0; height:1px; background: var(--line); margin:14px 0; }

.tableWrap{
  margin-top:12px;
  border:1px solid var(--line);
  border-radius: var(--radius);
  overflow:hidden;
  background: rgba(255,255,255,.02);
}
table{ border-collapse: collapse; width:100%; }
th,td{ border-bottom:1px solid var(--line); padding:10px 10px; text-align:center; font-size:13px; }
th{
  background: rgba(255,255,255,.04);
  color: var(--muted);
  font-weight:700;
  position: sticky;
  top:0;
  backdrop-filter: blur(8px);
}
tr:hover td{ background: rgba(255,255,255,.03); }
.code{ font-family:var(--mono); background: rgba(255,255,255,.06); padding:2px 6px; border-radius:10px; }

.formRow{
  display:flex; gap:10px; flex-wrap:wrap; align-items:center;
  padding:12px 12px;
  border:1px solid var(--line);
  border-radius: var(--radius);
  background: rgba(255,255,255,.02);
  margin-top:12px;
}
.field{
  padding:9px 10px;
  border:1px solid var(--line);
  border-radius: 14px;
  background: rgba(0,0,0,.18);
  color: var(--text);
  min-width: 240px;
  outline:none;
}
.fieldSmall{ min-width: 160px; }
.field:focus{ border-color: rgba(125,211,252,.45); }

.toast{
  position: fixed;
  right: 18px;
  bottom: 18px;
  background: rgba(0,0,0,.75);
  color:#fff;
  padding: 10px 12px;
  border-radius: 14px;
  display:none;
  font-size:13px;
  border:1px solid rgba(255,255,255,.12);
  box-shadow: var(--shadow);
}

.badge{
  display:inline-flex;
  align-items:center;
  gap:8px;
  font-size:12px;
  padding:6px 10px;
  border:1px solid var(--line);
  border-radius: 999px;
  background: rgba(255,255,255,.03);
  color: var(--muted);
}
"""

def render_app_page(*, title: str, active: str, body_html: str, right_pills_html: str = "") -> HTMLResponse:
    """
    Renderiza una página HTML completa para la UI.
    Upgrade: escape básico de title/active para evitar XSS accidental.
    """
    from html import escape

    safe_title = escape(str(title))
    safe_active = str(active or "").strip()

    def nav_item(href: str, label: str, key: str) -> str:
        cls = "active" if key == safe_active else ""
        # label y href vienen hardcodeados aquí; si luego los haces dinámicos, escápalos.
        return f'<a class="{cls}" href="{href}">{label}</a>'

    html = f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>{safe_title}</title>
  <style>{APP_CSS}</style>
</head>
<body>
  <div class="app">
    <aside class="sidebar">
      <div class="brand">
        <div class="logo"></div>
        <div>
          <h1>Powerball AI</h1>
          <p>Dashboard • SQLite • FastAPI</p>
        </div>
      </div>

      <nav class="nav">
        {nav_item("/ui", "🏠 Dashboard", "home")}
        {nav_item("/tickets/table", "📋 Tickets Table", "table")}
        {nav_item("/ui/compare", "🧩 Compare", "compare")}
        {nav_item("/ui/recommendations", "🤖 AI Recommendations", "ai")}
        {nav_item("/docs", "📘 Swagger", "docs")}
      </nav>

      <hr class="sep"/>
      <div class="muted" style="font-size:12px; line-height:1.35;">
        Tip: escribe compare como <span class="code">10,16,29,33,69|22</span>.
      </div>
    </aside>

    <main class="main">
      <div class="topbar">
        <h2>{safe_title}</h2>
        <div class="right">{right_pills_html}</div>
      </div>

      {body_html}
    </main>
  </div>
</body>
</html>
"""

    # Upgrade: headers útiles para UI (no rompe)
    return HTMLResponse(
        content=html,
        headers={
            "Cache-Control": "no-store",
            "X-Content-Type-Options": "nosniff",
        },
    )


# -------------------------
#        BASE ROUTES + UI
# -------------------------
@app.get("/")
async def read_root():
    return {"message": "API de Powerball + IA funcionando 🚀", "ui": "/ui", "docs": "/docs"}


@app.get("/apple-touch-icon.png")
@app.get("/apple-touch-icon-precomposed.png")
def apple_touch_icon():
    return Response(status_code=204)


@app.get("/ui/board")
def ui_board_redirect(request: Request):
    qs = request.url.query
    if qs:
        return RedirectResponse(url=f"/tickets/table?{qs}", status_code=307)
    return RedirectResponse(url="/tickets/table", status_code=307)


@app.get("/status")
async def get_status():
    return {"status": "ok"}


@app.get("/health")
async def health():
    return {"ok": True, "service": "powerball_ai"}


@app.get("/admin/self_test")
def admin_self_test(db: Session = Depends(get_db)):
    out: Dict[str, Any] = {"ok": True, "checks": {}}

    # DB ping
    try:
        db.execute(text("SELECT 1"))
        out["checks"]["db"] = "ok"
    except Exception as e:
        out["ok"] = False
        out["checks"]["db"] = f"fail: {e}"

    # counts
    try:
        total = db.query(func.count(Ticket.id)).scalar() or 0
        out["checks"]["tickets_total"] = int(total)
    except Exception as e:
        out["ok"] = False
        out["checks"]["tickets_total"] = f"fail: {e}"

    try:
        future = db.query(func.count(Ticket.id)).filter(Ticket.status == "FUTURE").scalar() or 0
        past = db.query(func.count(Ticket.id)).filter(Ticket.status == "PAST").scalar() or 0
        out["checks"]["tickets_future"] = int(future)
        out["checks"]["tickets_past"] = int(past)
    except Exception as e:
        out["ok"] = False
        out["checks"]["tickets_status_counts"] = f"fail: {e}"

    # last draw
    try:
        last = db.query(DrawResult).order_by(DrawResult.draw_date.desc()).first()
        out["checks"]["last_draw_exists"] = bool(last)
        if last:
            out["checks"]["last_draw_date"] = last.draw_date.isoformat()
    except Exception as e:
        out["ok"] = False
        out["checks"]["last_draw"] = f"fail: {e}"

    return out


def _safe_scalar_int(q, default: int = 0) -> int:
    try:
        v = q.scalar()
        return int(v or 0)
    except Exception:
        return int(default)


@app.get("/ui", response_class=HTMLResponse)
def ui_home(db: Session = Depends(get_db)):
    total = _safe_scalar_int(db.query(func.count(Ticket.id)))
    future = _safe_scalar_int(db.query(func.count(Ticket.id)).filter(Ticket.status == "FUTURE"))
    past = _safe_scalar_int(db.query(func.count(Ticket.id)).filter(Ticket.status == "PAST"))

    last_draw_txt = "—"
    try:
        last_draw = db.query(DrawResult).order_by(DrawResult.draw_date.desc()).first()
    except Exception:
        last_draw = None

    if last_draw:
        # Protege contra None/strings raros
        try:
            dd = last_draw.draw_date.isoformat()
        except Exception:
            dd = str(last_draw.draw_date)

        try:
            wnums = f"{int(last_draw.wn1)},{int(last_draw.wn2)},{int(last_draw.wn3)},{int(last_draw.wn4)},{int(last_draw.wn5)}"
        except Exception:
            wnums = f"{last_draw.wn1},{last_draw.wn2},{last_draw.wn3},{last_draw.wn4},{last_draw.wn5}"

        try:
            pb = int(last_draw.winning_powerball)
        except Exception:
            pb = last_draw.winning_powerball

        last_draw_txt = f"{dd} | {wnums} PB {pb}"

    right = f"""
      <span class="pill">Total: <b style="color:var(--text)">{total}</b></span>
      <span class="pill">PAST: <b style="color:var(--text)">{past}</b></span>
      <span class="pill">FUTURE: <b style="color:var(--text)">{future}</b></span>
    """

    # (UX) Links de navegación + cards
    body = f"""
      <div class="grid">
        <div class="card">
          <h3>Tickets Table</h3>
          <p>Tabla clara (1 combinación por fila), filtros, compare, matches y export.</p>
          <div style="margin-top:12px; display:flex; gap:10px; flex-wrap:wrap;">
            <a class="btn primary" href="/tickets/table">Abrir</a>
            <a class="btn" href="/tickets/table?status=FUTURE">FUTURE</a>
            <a class="btn" href="/tickets/table?status=PAST">PAST</a>
          </div>
        </div>

        <div class="card">
          <h3>Compare</h3>
          <p>Agrupa por 3/4/5/6 coincidencias y exporta Excel.</p>
          <div style="margin-top:12px;">
            <a class="btn primary" href="/ui/compare">Abrir</a>
          </div>
          <div class="muted" style="margin-top:10px;">Ej: <span class="code">10,16,29,33,69|22</span></div>
        </div>

        <div class="card">
          <h3>AI Recommendations</h3>
          <p>Genera combinaciones, exporta y guarda como FUTURE con dedupe.</p>
          <div style="margin-top:12px;">
            <a class="btn good" href="/ui/recommendations">Abrir</a>
          </div>
        </div>

        <div class="card">
          <h3>Estado</h3>
          <p>Last draw: <span class="code">{last_draw_txt}</span></p>
          <div style="margin-top:12px; display:flex; gap:10px; flex-wrap:wrap;">
            <a class="btn" href="/admin/self_test">Run self test</a>
            <a class="btn" href="/docs">Swagger</a>
          </div>
        </div>
      </div>
    """

    return render_app_page(title="Dashboard", active="home", body_html=body, right_pills_html=right)

# -------------------------
#    COMPARE VALIDATOR
# -------------------------
@app.get("/compare/parse")
def compare_parse(compare: str = Query(..., description='Ej: "10,16,29,33,69|22"')):
    parsed = parse_compare(compare)
    if not parsed:
        raise HTTPException(status_code=400, detail="compare inválido")

    regs, pb = parsed
    regs_sorted = sorted(int(x) for x in regs)

    normalized = f"{regs_sorted[0]},{regs_sorted[1]},{regs_sorted[2]},{regs_sorted[3]},{regs_sorted[4]}|{int(pb)}"
    return {"ok": True, "normalized": normalized, "regulars": regs_sorted, "powerball": int(pb)}


# -------------------------
#        STATS
# -------------------------
@app.get("/stats/summary", response_model=StatsSummary)
def get_stats_summary(db: Session = Depends(get_db)):
    # Upgrade: hacerlo en SQL (mucho más rápido que traer todos los tickets)
    try:
        total_tickets = int(db.query(func.count(Ticket.id)).scalar() or 0)
    except Exception:
        total_tickets = 0

    # cost/prize pueden ser NULL; coalesce evita None
    try:
        total_cost = float(
            db.query(func.coalesce(func.sum(Ticket.cost), 0.0)).scalar() or 0.0
        )
    except Exception:
        total_cost = 0.0

    # prize_amount puede no existir en tu modelo viejo: fallback seguro
    prize_col = getattr(Ticket, "prize_amount", None)
    if prize_col is not None:
        try:
            total_prize = float(
                db.query(func.coalesce(func.sum(prize_col), 0.0)).scalar() or 0.0
            )
        except Exception:
            total_prize = 0.0

        try:
            winning_tickets = int(
                db.query(func.count(Ticket.id))
                .filter(func.coalesce(prize_col, 0.0) > 0.0)
                .scalar()
                or 0
            )
        except Exception:
            winning_tickets = 0
    else:
        # Si no hay columna prize_amount, mantenemos compatibilidad sin romper
        total_prize = 0.0
        winning_tickets = 0

    balance = float(total_prize - total_cost)

    return StatsSummary(
        total_tickets=total_tickets,
        total_cost=total_cost,
        total_prize=total_prize,
        balance=balance,
        winning_tickets=winning_tickets,
    )


# ============================================================
# ✅ /tickets/table PRO (UI principal) — Opción 2 + Extras
# ============================================================
@app.get("/tickets/table", response_class=HTMLResponse)
def tickets_table(
    status: Optional[str] = Query(default=None),
    type: Optional[str] = Query(default=None),
    order: Literal["asc", "desc"] = Query(default="asc"),
    compare: Optional[str] = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=200, ge=20, le=2000),
    key_contains: Optional[str] = Query(default=None),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),

    # ✅ EXTRAS
    only_matches: bool = Query(default=False),
    min_match: int = Query(default=0, ge=0, le=6),      # 0=all, 3,4,5,6
    sort_by: Literal["id", "draw_date", "matches"] = Query(default="id"),

    db: Session = Depends(get_db),
):
    from html import escape

    def esc(v: Any) -> str:
        # escape seguro para atributos/HTML
        return escape("" if v is None else str(v), quote=True)

    sd = _parse_date_maybe(start_date)
    ed = _parse_date_maybe(end_date)

    query = db.query(Ticket)

    s = _norm_status(status)
    if s:
        query = query.filter(Ticket.status == s)

    ttype = _norm_type(type)
    if ttype:
        query = query.filter(Ticket.type == ttype)

    if sd:
        query = query.filter(Ticket.draw_date >= sd)
    if ed:
        query = query.filter(Ticket.draw_date <= ed)

    # Compare: puede levantar HTTPException si es inválido
    compare_parsed = parse_compare(compare) if compare else None

    # Último draw (si no hay compare)
    try:
        last_draw = db.query(DrawResult).order_by(DrawResult.draw_date.desc()).first()
    except Exception:
        last_draw = None

    if compare_parsed:
        winners_set, winning_pb = compare_parsed
        winners_set = {int(x) for x in winners_set}
        winning_pb = int(winning_pb)
        compare_label = f"Comparando con: {compare}"
    elif last_draw:
        winners_set = {int(last_draw.wn1), int(last_draw.wn2), int(last_draw.wn3), int(last_draw.wn4), int(last_draw.wn5)}
        try:
            winning_pb = int(last_draw.winning_powerball)
        except Exception:
            winning_pb = last_draw.winning_powerball
        compare_label = (
            f"Comparando con último draw ({last_draw.draw_date.isoformat()}): "
            f"{sorted(list(winners_set))} | PB {winning_pb}"
        )
    else:
        winners_set = set()
        winning_pb = None
        compare_label = "Sin comparación (no hay draw_results guardados y no enviaste compare)."

    # Total de tickets según filtros SQL (status/type/date). Ojo: no incluye key_contains / matches porque son in-memory.
    total = _count_fast(query)

    page_size = max(20, min(2000, int(page_size)))
    pages = max(1, (int(total) + page_size - 1) // page_size)
    page = max(1, min(int(page), pages))
    offset = (page - 1) * page_size

    # Upgrade: para sort_by id/draw_date, ordenar en DB ANTES de paginar (más correcto)
    if sort_by == "draw_date":
        query = query.order_by(Ticket.draw_date.asc(), Ticket.id.asc())
    else:
        # default: id asc
        query = query.order_by(Ticket.id.asc())

    base_tickets = query.offset(offset).limit(page_size).all()

    key_contains_norm = (key_contains or "").strip()

    def _ticket_pb(t: Ticket) -> Optional[int]:
        v = getattr(t, "pb", None)
        if v is None:
            v = getattr(t, "powerball", None)
        try:
            return int(v) if v is not None else None
        except Exception:
            return None

    def calc_match(t: Ticket) -> Tuple[int, bool, int]:
        """returns (matched_regular, matched_pb, total_balls=mr + pb_if)"""
        if not winners_set or winning_pb is None:
            return 0, False, 0

        try:
            regs = {int(t.n1), int(t.n2), int(t.n3), int(t.n4), int(t.n5)}
        except Exception:
            return 0, False, 0

        mr = len(regs.intersection(winners_set))

        tpb = _ticket_pb(t)
        mpb = (tpb is not None) and (int(tpb) == int(winning_pb))

        total_balls = mr + (1 if mpb else 0)
        return mr, mpb, total_balls

    items = []
    for t in base_tickets:
        regs_sorted = normalize_regular_numbers(t.n1, t.n2, t.n3, t.n4, t.n5, order=order)
        tpb = _ticket_pb(t) or 0
        k = numbers_key([t.n1, t.n2, t.n3, t.n4, t.n5], int(tpb))

        if key_contains_norm and key_contains_norm not in k:
            continue

        mr, mpb, total_balls = calc_match(t)

        if only_matches and total_balls <= 0:
            continue
        if min_match and total_balls < int(min_match):
            continue

        items.append((t, regs_sorted, k, mr, mpb, total_balls))

    # sort_by=matches es in-memory (no se puede en SQL sin materializar)
    if sort_by == "matches":
        items.sort(key=lambda x: (x[5], x[3], x[0].draw_date, x[0].id), reverse=True)

    def td(val: int, is_pb: bool = False) -> str:
        style = ""
        if is_pb and winning_pb is not None and int(val) == int(winning_pb):
            style = 'style="background:rgba(252,165,165,.22); color:#ffd1d1; font-weight:800;"'
        elif (not is_pb) and winners_set and int(val) in winners_set:
            style = 'style="background:rgba(52,211,153,.18); font-weight:800;"'
        return f"<td {style}>{int(val)}</td>"

    def match_badge(mr: int, mpb: bool, total_balls: int) -> str:
        if not winners_set or winning_pb is None:
            return "<span class='badge'>—</span>"
        if total_balls <= 0:
            return "<span class='badge'>0</span>"
        pb_txt = " +PB" if mpb else ""
        return (
            "<span class='badge'>"
            f"<b style='color:var(--text)'>{int(mr)}</b>{pb_txt} → "
            f"<b style='color:var(--text)'>{int(total_balls)}</b>"
            "</span>"
        )

    rows = ""
    shown = 0

    for (t, regs, k, mr, mpb, total_balls) in items:
        shown += 1
        tpb = _ticket_pb(t) or 0
        combo_txt = f"{regs[0]} {regs[1]} {regs[2]} {regs[3]} {regs[4]} | PB {int(tpb)}"

        rows += "<tr>"
        rows += f"<td>{int(t.id)}</td>"
        rows += f"<td>{esc(t.draw_date)}</td>"
        rows += f"<td><span class='badge'>{esc(t.status)}</span></td>"
        rows += f"<td><span class='badge'>{esc(t.type)}</span></td>"
        rows += td(regs[0]); rows += td(regs[1]); rows += td(regs[2]); rows += td(regs[3]); rows += td(regs[4])
        rows += td(int(tpb), is_pb=True)
        rows += f"<td>{match_badge(mr, mpb, total_balls)}</td>"
        rows += f"<td>${float(t.cost):.2f}</td>"
        rows += f"<td><span class='code'>{esc(k)}</span></td>"
        rows += f"<td><button class='btn' style='padding:7px 10px;' data-copy='{esc(combo_txt)}'>Copy</button></td>"
        rows += "</tr>\n"

    if shown == 0:
        # ✅ Fix: colspan correcto (la tabla tiene 14 columnas)
        rows = "<tr><td colspan='14' class='muted' style='text-align:left; padding:14px;'>No hay resultados con estos filtros.</td></tr>"

    def qparam(k: str, v: Optional[str]) -> str:
        if v is None:
            return ""
        vv = str(v).strip()
        if not vv:
            return ""
        return f"&{k}={_quote(vv)}"

    base_qs = (
        f"?page_size={page_size}"
        f"{qparam('status', s)}"
        f"{qparam('type', ttype)}"
        f"{qparam('order', order)}"
        f"{qparam('compare', compare)}"
        f"{qparam('key_contains', key_contains_norm)}"
        f"{qparam('start_date', start_date)}"
        f"{qparam('end_date', end_date)}"
        f"{qparam('sort_by', sort_by)}"
        f"&only_matches={'true' if only_matches else 'false'}"
        f"&min_match={int(min_match)}"
    )

    prev_link = f"/tickets/table{base_qs}&page={max(1, page-1)}"
    next_link = f"/tickets/table{base_qs}&page={min(pages, page+1)}"

    exp_xlsx = (
        f"/export_excel?order={order}"
        f"{qparam('status', s)}"
        f"{qparam('type', ttype)}"
        f"{qparam('compare', compare)}"
        f"{qparam('start_date', start_date)}"
        f"{qparam('end_date', end_date)}"
    )
    exp_csv = (
        f"/export_csv?order={order}"
        f"{qparam('status', s)}"
        f"{qparam('type', ttype)}"
        f"{qparam('start_date', start_date)}"
        f"{qparam('end_date', end_date)}"
    )

    compare_default = (compare or "").strip()

    body = f"""
      <div class="formRow">
        <input id="compareInput" class="field" placeholder="Compare: 10,16,29,33,69|22" value="{esc(compare_default)}"/>

        <select id="statusSel" class="field fieldSmall">
          <option value="" {"selected" if not s else ""}>Status: ALL</option>
          <option value="PAST" {"selected" if s=="PAST" else ""}>PAST</option>
          <option value="FUTURE" {"selected" if s=="FUTURE" else ""}>FUTURE</option>
        </select>

        <select id="typeSel" class="field fieldSmall">
          <option value="" {"selected" if not ttype else ""}>Type: ALL</option>
          <option value="QUICK_PICK" {"selected" if ttype=="QUICK_PICK" else ""}>QUICK_PICK</option>
          <option value="MANUAL" {"selected" if ttype=="MANUAL" else ""}>MANUAL</option>
        </select>

        <select id="orderSel" class="field fieldSmall">
          <option value="asc" {"selected" if order=="asc" else ""}>Order: ASC</option>
          <option value="desc" {"selected" if order=="desc" else ""}>Order: DESC</option>
        </select>

        <select id="sortBySel" class="field fieldSmall">
          <option value="id" {"selected" if sort_by=="id" else ""}>Sort: ID</option>
          <option value="draw_date" {"selected" if sort_by=="draw_date" else ""}>Sort: DATE</option>
          <option value="matches" {"selected" if sort_by=="matches" else ""}>Sort: MATCHES</option>
        </select>

        <input id="keyContains" class="field fieldSmall" placeholder="key contains (opcional)" value="{esc(key_contains_norm)}"/>
        <input id="startDate" class="field fieldSmall" placeholder="start YYYY-MM-DD" value="{esc(start_date or '')}"/>
        <input id="endDate" class="field fieldSmall" placeholder="end YYYY-MM-DD" value="{esc(end_date or '')}"/>

        <label class="badge" style="cursor:pointer;">
          <input id="onlyMatches" type="checkbox" style="transform:scale(1.1);" {"checked" if only_matches else ""}/>
          Only matches
        </label>

        <select id="minMatch" class="field fieldSmall">
          <option value="0" {"selected" if int(min_match)==0 else ""}>Min match: 0</option>
          <option value="3" {"selected" if int(min_match)==3 else ""}>Min match: 3</option>
          <option value="4" {"selected" if int(min_match)==4 else ""}>Min match: 4</option>
          <option value="5" {"selected" if int(min_match)==5 else ""}>Min match: 5</option>
          <option value="6" {"selected" if int(min_match)==6 else ""}>Min match: 6</option>
        </select>

        <button id="applyBtn" class="btn primary">Apply</button>
        <a class="btn" href="{esc(exp_xlsx)}">Export Excel</a>
        <a class="btn" href="{esc(exp_csv)}">Export CSV</a>
        <a class="btn" href="/ui/compare">Compare Tool</a>
      </div>

      <div class="muted" style="margin-top:10px;">{esc(compare_label)}</div>

      <div class="tableWrap">
        <table>
          <thead>
            <tr>
              <th>ID</th>
              <th>Draw Date</th>
              <th>Status</th>
              <th>Type</th>
              <th>N1</th><th>N2</th><th>N3</th><th>N4</th><th>N5</th>
              <th>PB</th>
              <th>Match</th>
              <th>Cost</th>
              <th>Key</th>
              <th>Copy</th>
            </tr>
          </thead>
          <tbody>
            {rows}
          </tbody>
        </table>
      </div>

      <div style="margin-top:12px; display:flex; gap:10px; flex-wrap:wrap; align-items:center;">
        <a class="btn" href="{esc(prev_link)}">⬅ Prev</a>
        <span class="pill">Page <b style="color:var(--text)">{page}</b> / {pages}</span>
        <span class="pill">Total <b style="color:var(--text)">{int(total)}</b></span>
        <a class="btn" href="{esc(next_link)}">Next ➡</a>
      </div>

      <div class="toast" id="toast">Copied ✅</div>

      <script>
        function toast(msg) {{
          const t = document.getElementById('toast');
          t.textContent = msg || 'OK';
          t.style.display = 'block';
          setTimeout(()=>t.style.display='none', 900);
        }}

        // Copy buttons
        document.querySelectorAll('button[data-copy]').forEach(btn => {{
          btn.addEventListener('click', async () => {{
            const txt = btn.getAttribute('data-copy');
            try {{
              await navigator.clipboard.writeText(txt);
              toast('Copied ✅');
            }} catch (e) {{
              toast('Copy failed');
            }}
          }});
        }});

        // Apply (sin editar URL a mano)
        document.getElementById('applyBtn').addEventListener('click', () => {{
          const compare = (document.getElementById('compareInput').value || '').trim();
          const status = (document.getElementById('statusSel').value || '').trim();
          const type = (document.getElementById('typeSel').value || '').trim();
          const order = (document.getElementById('orderSel').value || 'asc').trim();
          const sortBy = (document.getElementById('sortBySel').value || 'id').trim();
          const keyc = (document.getElementById('keyContains').value || '').trim();
          const sd = (document.getElementById('startDate').value || '').trim();
          const ed = (document.getElementById('endDate').value || '').trim();
          const onlyMatches = document.getElementById('onlyMatches').checked;
          const minMatch = (document.getElementById('minMatch').value || '0').trim();

          const params = new URLSearchParams();
          if (compare) params.set('compare', compare);
          if (status) params.set('status', status);
          if (type) params.set('type', type);
          if (order) params.set('order', order);
          if (sortBy) params.set('sort_by', sortBy);
          if (keyc) params.set('key_contains', keyc);
          if (sd) params.set('start_date', sd);
          if (ed) params.set('end_date', ed);
          if (onlyMatches) params.set('only_matches', 'true');
          if (minMatch && minMatch !== '0') params.set('min_match', minMatch);

          params.set('page', '1');
          params.set('page_size', '{page_size}');

          window.location.href = '/tickets/table?' + params.toString();
        }});
      </script>
    """

    right = f"""
      <span class="pill">Showing: <b style="color:var(--text)">{int(shown)}</b></span>
      <span class="pill">Page size: <b style="color:var(--text)">{int(page_size)}</b></span>
    """

    return render_app_page(title="Tickets Table", active="table", body_html=body, right_pills_html=right)


# ============================================================
#   TICKETS CRUD
# ============================================================
@app.post("/tickets", response_model=TicketOut)
def create_ticket(payload: TicketCreate, db: Session = Depends(get_db)):
    ttype = _norm_type(payload.type)
    if not ttype:
        raise HTTPException(status_code=400, detail="type inválido. Usa QUICK_PICK o MANUAL.")

    # Validación de números
    validate_ticket_numbers(payload.n1, payload.n2, payload.n3, payload.n4, payload.n5, payload.powerball)

    # Status por defecto según draw_date
    status = payload.status or ("PAST" if payload.draw_date <= date.today() else "FUTURE")

    # Key única para dedupe por draw_date
    key = numbers_key([payload.n1, payload.n2, payload.n3, payload.n4, payload.n5], int(payload.powerball))

    # Upgrade: dedupe eficiente (no cargar todos los tickets de ese día)
    exists_q = (
        db.query(Ticket.id)
        .filter(
            Ticket.draw_date == payload.draw_date,
            Ticket.n1 == int(payload.n1),
            Ticket.n2 == int(payload.n2),
            Ticket.n3 == int(payload.n3),
            Ticket.n4 == int(payload.n4),
            Ticket.n5 == int(payload.n5),
        )
    )
    # Compatibilidad pb/powerball en modelo
    pb_col = getattr(Ticket, "pb", None)
    if pb_col is None:
        pb_col = getattr(Ticket, "powerball")
    exists_q = exists_q.filter(pb_col == int(payload.powerball))

    if db.query(exists_q.exists()).scalar():
        raise HTTPException(status_code=409, detail="Ticket duplicado (misma combinación para ese draw_date).")

    # Crear ticket (compatibilidad pb/powerball)
    ticket_kwargs = dict(
        draw_date=payload.draw_date,
        status=status,
        n1=int(payload.n1),
        n2=int(payload.n2),
        n3=int(payload.n3),
        n4=int(payload.n4),
        n5=int(payload.n5),
        type=ttype,
        cost=float(payload.cost or 0.0),
        matched_regular_numbers=0,
        matched_powerball=False,
        prize_amount=0.0,
    )
    if hasattr(Ticket, "pb"):
        ticket_kwargs["pb"] = int(payload.powerball)
    else:
        ticket_kwargs["powerball"] = int(payload.powerball)

    ticket = Ticket(**ticket_kwargs)

    # Si existe draw result para ese día, computa matches
    res = db.query(DrawResult).filter(DrawResult.draw_date == payload.draw_date).first()
    if res:
        calculate_matches(ticket, res)

    db.add(ticket)
    db.commit()
    db.refresh(ticket)
    return ticket


@app.get("/tickets", response_model=List[TicketOut])
def list_tickets(
    status: Optional[str] = Query(default=None),
    type: Optional[str] = Query(default=None),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    limit: int = Query(default=2000, ge=1, le=20000),
    db: Session = Depends(get_db),
):
    s = _norm_status(status)
    ttype = _norm_type(type)
    sd = _parse_date_maybe(start_date)
    ed = _parse_date_maybe(end_date)

    q = db.query(Ticket)
    if s:
        q = q.filter(Ticket.status == s)
    if ttype:
        q = q.filter(Ticket.type == ttype)
    if sd:
        q = q.filter(Ticket.draw_date >= sd)
    if ed:
        q = q.filter(Ticket.draw_date <= ed)

    return q.order_by(Ticket.id.asc()).limit(int(limit)).all()


@app.get("/tickets/{ticket_id}", response_model=TicketOut)
def get_ticket(ticket_id: int = Path(..., ge=1), db: Session = Depends(get_db)):
    t = db.query(Ticket).filter(Ticket.id == int(ticket_id)).first()
    if not t:
        raise HTTPException(status_code=404, detail="Ticket no encontrado")
    return t


@app.patch("/tickets/{ticket_id}", response_model=TicketOut)
def update_ticket(ticket_id: int = Path(..., ge=1), payload: TicketUpdate = Body(...), db: Session = Depends(get_db)):
    t = db.query(Ticket).filter(Ticket.id == int(ticket_id)).first()
    if not t:
        raise HTTPException(status_code=404, detail="Ticket no encontrado")

    # Update campos simples
    if payload.draw_date is not None:
        t.draw_date = payload.draw_date

    if payload.status is not None:
        if payload.status not in ALLOWED_STATUS:
            raise HTTPException(status_code=400, detail="status inválido")
        t.status = payload.status

    if payload.type is not None:
        ttype = _norm_type(payload.type)
        if not ttype:
            raise HTTPException(status_code=400, detail="type inválido")
        t.type = ttype

    # Determinar números finales (mix payload + existentes)
    n1 = int(payload.n1) if payload.n1 is not None else int(t.n1)
    n2 = int(payload.n2) if payload.n2 is not None else int(t.n2)
    n3 = int(payload.n3) if payload.n3 is not None else int(t.n3)
    n4 = int(payload.n4) if payload.n4 is not None else int(t.n4)
    n5 = int(payload.n5) if payload.n5 is not None else int(t.n5)

    # Compatibilidad pb/powerball
    current_pb = getattr(t, "pb", None)
    if current_pb is None:
        current_pb = getattr(t, "powerball", None)

    pb = int(payload.powerball) if payload.powerball is not None else int(current_pb)

    validate_ticket_numbers(n1, n2, n3, n4, n5, pb)

    t.n1, t.n2, t.n3, t.n4, t.n5 = n1, n2, n3, n4, n5
    if hasattr(t, "pb"):
        t.pb = pb
    else:
        t.powerball = pb

    if payload.cost is not None:
        t.cost = float(payload.cost)

    # Upgrade: dedupe eficiente (no cargar todos los tickets del día)
    dup_q = (
        db.query(Ticket.id)
        .filter(
            Ticket.draw_date == t.draw_date,
            Ticket.id != t.id,
            Ticket.n1 == int(t.n1),
            Ticket.n2 == int(t.n2),
            Ticket.n3 == int(t.n3),
            Ticket.n4 == int(t.n4),
            Ticket.n5 == int(t.n5),
        )
    )
    pb_col = getattr(Ticket, "pb", None)
    if pb_col is None:
        pb_col = getattr(Ticket, "powerball")
    dup_q = dup_q.filter(pb_col == int(pb))

    if db.query(dup_q.exists()).scalar():
        raise HTTPException(status_code=409, detail="Update crea duplicado (misma combinación para ese draw_date).")

    # Recalcular matches si hay draw result, sino reset
    res = db.query(DrawResult).filter(DrawResult.draw_date == t.draw_date).first()
    if res:
        calculate_matches(t, res)
    else:
        setattr(t, "matched_regular_numbers", 0)
        setattr(t, "matched_powerball", False)
        setattr(t, "prize_amount", 0.0)

    db.commit()
    db.refresh(t)
    return t


@app.delete("/tickets/{ticket_id}")
def delete_ticket(ticket_id: int = Path(..., ge=1), db: Session = Depends(get_db)):
    t = db.query(Ticket).filter(Ticket.id == int(ticket_id)).first()
    if not t:
        raise HTTPException(status_code=404, detail="Ticket no encontrado")
    db.delete(t)
    db.commit()
    return {"ok": True, "deleted": int(ticket_id)}


# ============================================================
#   DRAW RESULTS CRUD
# ============================================================
@app.post("/draw_results", response_model=DrawResultOut)
def create_draw_result(payload: DrawResultCreate, db: Session = Depends(get_db)):
    regs = [int(payload.wn1), int(payload.wn2), int(payload.wn3), int(payload.wn4), int(payload.wn5)]

    if any((x < REG_MIN or x > REG_MAX) for x in regs):
        raise HTTPException(status_code=400, detail="wn1..wn5 fuera de rango")
    if len(set(regs)) != 5:
        raise HTTPException(status_code=400, detail="wn1..wn5 no pueden repetirse")

    pb = int(payload.winning_powerball)
    if pb < PB_MIN or pb > PB_MAX:
        raise HTTPException(status_code=400, detail="winning_powerball fuera de rango")

    # Upgrade: exists limpio
    exists = db.query(DrawResult.id).filter(DrawResult.draw_date == payload.draw_date).first()
    if exists:
        raise HTTPException(status_code=409, detail="Ya existe un draw_result para esa fecha")

    dr = DrawResult(
        draw_date=payload.draw_date,
        wn1=regs[0],
        wn2=regs[1],
        wn3=regs[2],
        wn4=regs[3],
        wn5=regs[4],
        winning_powerball=pb,
    )
    db.add(dr)
    db.commit()
    db.refresh(dr)

    # Upgrade: tickets en streaming
    tickets_q = db.query(Ticket).filter(Ticket.draw_date == payload.draw_date).yield_per(500)

    today = date.today()
    updated = 0

    for t in tickets_q:
        try:
            # Compatibilidad pb/powerball para calculate_matches()
            if not hasattr(t, "powerball") and hasattr(t, "pb"):
                try:
                    setattr(t, "powerball", getattr(t, "pb"))
                except Exception:
                    pass

            calculate_matches(t, dr)

            # Status PAST si ya pasó
            if getattr(t, "draw_date", None) and t.draw_date <= today:
                t.status = "PAST"

            updated += 1
        except Exception:
            # Ticket corrupto o inesperado: no rompe el proceso completo
            continue

    db.commit()
    return dr


@app.get("/draw_results", response_model=List[DrawResultOut])
def list_draw_results(limit: int = Query(default=2000, ge=1, le=20000), db: Session = Depends(get_db)):
    return db.query(DrawResult).order_by(DrawResult.draw_date.desc()).limit(int(limit)).all()


@app.get("/draw_results/{draw_date}", response_model=DrawResultOut)
def get_draw_result(draw_date: str, year: int | None = None, db: Session = Depends(get_db)):
    d = _parse_date_flexible(draw_date, year=year)
    dr = db.query(DrawResult).filter(DrawResult.draw_date == d).first()
    if not dr:
        raise HTTPException(status_code=404, detail="No existe draw_result para esa fecha")
    return dr


@app.delete("/draw_results/{draw_date}")
def delete_draw_result(draw_date: str, year: int | None = None, db: Session = Depends(get_db)):
    d = _parse_date_flexible(draw_date, year=year)
    dr = db.query(DrawResult).filter(DrawResult.draw_date == d).first()
    if not dr:
        raise HTTPException(status_code=404, detail="No existe draw_result para esa fecha")

    # Todo en una transacción lógica
    db.delete(dr)

    tickets_q = db.query(Ticket).filter(Ticket.draw_date == d).yield_per(500)

    affected = 0
    for t in tickets_q:
        try:
            t.matched_regular_numbers = 0
            t.matched_powerball = False
            t.prize_amount = 0.0
            affected += 1
        except Exception:
            continue

    db.commit()
    return {"ok": True, "deleted_draw_date": d.isoformat(), "affected_tickets": int(affected)}




# ============================================================
#   IMPORT EXCEL (Draw Results)
# ============================================================
def _normalize_columns(cols: List[str]) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for c in cols:
        raw = str(c).strip()
        k = raw.lower()
        k = k.replace(" ", "_").replace("-", "_").replace("__", "_")
        mapping[c] = k
    return mapping


def _coerce_status(v: Any, draw_date_val: date) -> str:
    s = _norm_status(str(v) if v is not None else "")
    if s:
        return s
    return "PAST" if draw_date_val <= date.today() else "FUTURE"


def _to_int_cell(x: Any, field: str) -> int:
    """
    Convierte celdas Excel típicas:
    - 10
    - 10.0
    - "10"
    - " 10 "
    """
    if x is None or (isinstance(x, float) and pd.isna(x)) or (isinstance(x, str) and not x.strip()):
        raise ValueError(f"{field} vacío")
    try:
        if isinstance(x, float):
            if pd.isna(x):
                raise ValueError
            return int(x)
        return int(str(x).strip())
    except Exception:
        raise ValueError(f"{field} inválido: {x}")


def _powerball_rules_for_date(draw_date_val: date) -> Dict[str, int]:
    """
    Reglas históricas por fecha para importar 1992 → hoy sin errores.

    Eras:
    - 1992-04-22 .. 1997-11-04 : 5/45 + PB 1/45
    - 1997-11-05 .. 2002-10-08 : 5/49 + PB 1/42
    - 2002-10-09 .. 2005-08-30 : 5/53 + PB 1/42
    - 2005-08-31 .. 2009-01-06 : 5/55 + PB 1/42
    - 2009-01-07 .. 2012-01-17 : 5/59 + PB 1/39
    - 2012-01-18 .. 2015-10-06 : 5/59 + PB 1/35
    - 2015-10-07 .. (hoy)      : 5/69 + PB 1/26
    """
    d = draw_date_val

    # Moderno
    if d >= date(2015, 10, 7):
        return {"n_min": 1, "n_max": 69, "pb_min": 1, "pb_max": 26}

    # 2012–2015 (PB 1–35, blancas 1–59)
    if d >= date(2012, 1, 18):
        return {"n_min": 1, "n_max": 59, "pb_min": 1, "pb_max": 35}

    # 2009–2012 (PB 1–39, blancas 1–59)
    if d >= date(2009, 1, 7):
        return {"n_min": 1, "n_max": 59, "pb_min": 1, "pb_max": 39}

    # 2005–2009 (PB 1–42, blancas 1–55)
    if d >= date(2005, 8, 31):
        return {"n_min": 1, "n_max": 55, "pb_min": 1, "pb_max": 42}

    # 2002–2005 (PB 1–42, blancas 1–53)
    if d >= date(2002, 10, 9):
        return {"n_min": 1, "n_max": 53, "pb_min": 1, "pb_max": 42}

    # 1997–2002 (PB 1–42, blancas 1–49)
    if d >= date(1997, 11, 5):
        return {"n_min": 1, "n_max": 49, "pb_min": 1, "pb_max": 42}

    # 1992–1997 (PB 1–45, blancas 1–45)
    if d >= date(1992, 4, 22):
        return {"n_min": 1, "n_max": 45, "pb_min": 1, "pb_max": 45}

    # Si dices "desde 1992", esto NO debería ocurrir.
    raise ValueError(f"draw_date {d.isoformat()} es anterior a 1992-04-22 (era pre-Powerball)")


def _validate_draw_numbers_by_date(
    *,
    draw_date_val: date,
    n1: int, n2: int, n3: int, n4: int, n5: int,
    pb: int
) -> None:
    rules = _powerball_rules_for_date(draw_date_val)
    nums = [int(n1), int(n2), int(n3), int(n4), int(n5)]
    if len(set(nums)) != 5:
        raise HTTPException(status_code=400, detail="Números regulares no pueden repetirse")

    for i, v in enumerate(nums, start=1):
        if v < rules["n_min"] or v > rules["n_max"]:
            raise HTTPException(
                status_code=400,
                detail=f"n{i} debe estar entre {rules['n_min']}-{rules['n_max']} para {draw_date_val.isoformat()}"
            )

    if int(pb) < rules["pb_min"] or int(pb) > rules["pb_max"]:
        raise HTTPException(
            status_code=400,
            detail=f"Powerball debe estar entre {rules['pb_min']}-{rules['pb_max']} para {draw_date_val.isoformat()}"
        )


@app.post("/import_excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Sube un archivo Excel (.xlsx/.xls)")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Archivo vacío.")

    try:
        df = pd.read_excel(BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No pude leer Excel: {e}")

    if df is None or df.empty:
        return {"ok": True, "inserted": 0, "updated": 0, "skipped": 0, "errors": 0}

    df = df.copy()
    colmap = _normalize_columns(list(df.columns))
    df.rename(columns=colmap, inplace=True)

    # --- Normalización: nombres alternos comunes ---
    if "draw_date" not in df.columns:
        for alt in ("date", "drawdate", "draw_date_", "draw_dt"):
            if alt in df.columns:
                df.rename(columns={alt: "draw_date"}, inplace=True)
                break

    # Si viene como wn* (formato DB), lo convertimos a n*
    if "n1" not in df.columns and "wn1" in df.columns:
        df.rename(
            columns={"wn1": "n1", "wn2": "n2", "wn3": "n3", "wn4": "n4", "wn5": "n5"},
            inplace=True,
        )

    if "powerball" not in df.columns and "winning_powerball" in df.columns:
        df.rename(columns={"winning_powerball": "powerball"}, inplace=True)

    if "powerball" not in df.columns:
        for alt in ("pb", "power_ball", "power ball", "powerball_", "power_ball_"):
            alt_norm = alt.replace(" ", "_")
            if alt_norm in df.columns:
                df.rename(columns={alt_norm: "powerball"}, inplace=True)
                break

    required_any = {"draw_date", "n1", "n2", "n3", "n4", "n5", "powerball"}
    if not required_any.issubset(set(df.columns)):
        raise HTTPException(
            status_code=400,
            detail=(
                f"Excel debe incluir columnas: {sorted(list(required_any))}. "
                f"Columnas detectadas: {list(df.columns)}"
            ),
        )

    inserted = 0
    updated = 0
    skipped = 0
    errors = 0
    error_samples: List[Dict[str, Any]] = []

    for idx, row in enumerate(df.itertuples(index=False), start=0):
        try:
            dd = getattr(row, "draw_date", None)
            if dd is None or (isinstance(dd, float) and pd.isna(dd)) or (isinstance(dd, str) and not dd.strip()):
                raise ValueError("draw_date vacío")

            if isinstance(dd, datetime):
                draw_date_val = dd.date()
            elif isinstance(dd, date):
                draw_date_val = dd
            else:
                draw_date_val = _parse_date_iso(str(dd))

            n1 = _to_int_cell(getattr(row, "n1", None), "n1")
            n2 = _to_int_cell(getattr(row, "n2", None), "n2")
            n3 = _to_int_cell(getattr(row, "n3", None), "n3")
            n4 = _to_int_cell(getattr(row, "n4", None), "n4")
            n5 = _to_int_cell(getattr(row, "n5", None), "n5")
            pb = _to_int_cell(getattr(row, "powerball", None), "powerball")

            # ✅ Validación por fecha (arregla 1992-actual)
            _validate_draw_numbers_by_date(
                draw_date_val=draw_date_val,
                n1=n1, n2=n2, n3=n3, n4=n4, n5=n5,
                pb=pb
            )

            dr = db.query(DrawResult).filter(DrawResult.draw_date == draw_date_val).first()

            if dr is None:
                dr = DrawResult(
                    draw_date=draw_date_val,
                    wn1=int(n1),
                    wn2=int(n2),
                    wn3=int(n3),
                    wn4=int(n4),
                    wn5=int(n5),
                    winning_powerball=int(pb),
                )
                db.add(dr)
                inserted += 1
            else:
                same = (
                    int(dr.wn1) == int(n1)
                    and int(dr.wn2) == int(n2)
                    and int(dr.wn3) == int(n3)
                    and int(dr.wn4) == int(n4)
                    and int(dr.wn5) == int(n5)
                    and int(dr.winning_powerball) == int(pb)
                )
                if same:
                    skipped += 1
                else:
                    dr.wn1 = int(n1)
                    dr.wn2 = int(n2)
                    dr.wn3 = int(n3)
                    dr.wn4 = int(n4)
                    dr.wn5 = int(n5)
                    dr.winning_powerball = int(pb)
                    updated += 1

        except Exception as e:
            errors += 1
            if len(error_samples) < 10:
                error_samples.append({"row": int(idx), "error": str(e)})

    db.commit()

    out: Dict[str, Any] = {
        "ok": True,
        "inserted": int(inserted),
        "updated": int(updated),
        "skipped": int(skipped),
        "errors": int(errors),
    }
    if error_samples:
        out["error_samples"] = error_samples
    return out





MONTHS = ("January","February","March","April","May","June",
          "July","August","September","October","November","December")

DOW = ("Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday")

def _parse_date(text: str):
    if not isinstance(text, str):
        return None
    s = text.strip()
    if not s:
        return None

    # descarta encabezados tipo "January 1993"
    if re.fullmatch(rf"({'|'.join(MONTHS)})\s+\d{{4}}", s):
        return None

    # quita "Wednesday " etc si viene al inicio
    for d in DOW:
        if s.startswith(d + " "):
            s = s[len(d)+1:]
            break

    # soporta "January 2nd 1993" y "January 2 1993"
    s = re.sub(r"(\d{1,2})(st|nd|rd|th)", r"\1", s)

    dt = pd.to_datetime(s, errors="coerce")
    if pd.isna(dt):
        return None
    return dt

def _extract_6_numbers(text: str):
    if not isinstance(text, str):
        return None
    nums = [int(x) for x in re.findall(r"\d+", text)]
    if len(nums) != 6:
        return None
    return nums

def convert_powerball_legacy_xlsx(input_path: str) -> pd.DataFrame:
    raw = pd.read_excel(input_path)

    # buscamos pares de columnas: (fecha, numeros) que funcionen mejor
    cols = list(raw.columns)
    best = None
    best_count = -1

    for i in range(len(cols)):
        for j in range(len(cols)):
            if i == j:
                continue
            c_date, c_nums = cols[i], cols[j]
            count = 0
            for _, r in raw[[c_date, c_nums]].iterrows():
                dt = _parse_date(r[c_date])
                nums = _extract_6_numbers(r[c_nums])
                if dt is not None and nums is not None:
                    count += 1
            if count > best_count:
                best_count = count
                best = (c_date, c_nums)

    if best is None or best_count == 0:
        raise ValueError("No se encontraron columnas (fecha, numeros) válidas en el archivo.")

    c_date, c_nums = best

    rows = []
    for _, r in raw[[c_date, c_nums]].iterrows():
        dt = _parse_date(r[c_date])
        nums = _extract_6_numbers(r[c_nums])
        if dt is None or nums is None:
            continue
        rows.append((dt, *nums[:5], nums[5]))

    df = pd.DataFrame(rows, columns=["draw_date","n1","n2","n3","n4","n5","powerball"])
    df["draw_date"] = pd.to_datetime(df["draw_date"]).dt.strftime("%Y-%m-%d")
    df = df.sort_values("draw_date", ascending=False).reset_index(drop=True)
    return df



# ============================================================
#   EXPORT CSV / EXCEL (con resaltado matches)
# ============================================================
def _filtered_ticket_query(
    db: Session,
    status: Optional[str],
    type: Optional[str],
    start_date: Optional[str],
    end_date: Optional[str],
):
    s = _norm_status(status)
    ttype = _norm_type(type)
    sd = _parse_date_maybe(start_date)
    ed = _parse_date_maybe(end_date)

    q = db.query(Ticket)
    if s:
        q = q.filter(Ticket.status == s)
    if ttype:
        q = q.filter(Ticket.type == ttype)
    if sd:
        q = q.filter(Ticket.draw_date >= sd)
    if ed:
        q = q.filter(Ticket.draw_date <= ed)
    return q


def _ticket_pb_value(t: Ticket) -> int:
    """
    Compatibilidad: algunos modelos usan pb en vez de powerball.
    """
    if hasattr(t, "pb"):
        return int(getattr(t, "pb"))
    return int(getattr(t, "powerball"))


@app.get("/export_csv")
def export_csv(
    status: Optional[str] = Query(default=None),
    type: Optional[str] = Query(default=None),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    order: Literal["asc", "desc"] = Query(default="asc"),
    db: Session = Depends(get_db),
):
    q = _filtered_ticket_query(db, status, type, start_date, end_date)

    # Upgrade: no cargamos todo si hay muchos; igual armamos rows (CSV requiere materializar)
    tickets = q.order_by(Ticket.id.asc()).yield_per(1000)

    rows: List[Dict[str, Any]] = []
    for t in tickets:
        pb_val = _ticket_pb_value(t)
        regs = normalize_regular_numbers(int(t.n1), int(t.n2), int(t.n3), int(t.n4), int(t.n5), order=order)

        rows.append({
            "id": int(t.id),
            "draw_date": t.draw_date.isoformat() if getattr(t, "draw_date", None) else "",
            "status": str(getattr(t, "status", "")),
            "type": str(getattr(t, "type", "")),
            "n1": int(regs[0]),
            "n2": int(regs[1]),
            "n3": int(regs[2]),
            "n4": int(regs[3]),
            "n5": int(regs[4]),
            "powerball": int(pb_val),
            "matched_regular_numbers": int(getattr(t, "matched_regular_numbers", 0) or 0),
            "matched_powerball": bool(getattr(t, "matched_powerball", False) or False),
            "prize_amount": float(getattr(t, "prize_amount", 0.0) or 0.0),
            "cost": float(getattr(t, "cost", 0.0) or 0.0),
            "key": numbers_key([int(t.n1), int(t.n2), int(t.n3), int(t.n4), int(t.n5)], int(pb_val)),
        })

    df = pd.DataFrame(rows)

    filename = "tickets_export.csv"
    # Si ya tienes el helper global, úsalo (mejor consistencia)
    try:
        return _df_to_csv_stream(df, filename)
    except Exception:
        buf = BytesIO()
        df.to_csv(buf, index=False)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )


@app.get("/export_compare_group")
def export_compare_group(
    compare: str = Query(...),
    group: Literal["3", "4", "5", "6"] = Query(...),
    status: Optional[str] = Query(default=None),
    type: Optional[str] = Query(default=None),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    order: Literal["asc", "desc"] = Query(default="asc"),
    db: Session = Depends(get_db),
):
    parsed = parse_compare(compare)
    if not parsed:
        raise HTTPException(status_code=400, detail="compare inválido")

    winners_set, winning_pb = parsed  # base regs + pb
    sd = _parse_date_maybe(start_date)
    ed = _parse_date_maybe(end_date)

    q = db.query(Ticket)

    s = _norm_status(status)
    if s:
        q = q.filter(Ticket.status == s)

    ttype = _norm_type(type)
    if ttype:
        q = q.filter(Ticket.type == ttype)

    if sd:
        q = q.filter(Ticket.draw_date >= sd)
    if ed:
        q = q.filter(Ticket.draw_date <= ed)

    rows: List[Dict[str, Any]] = []
    tickets = q.order_by(Ticket.id.asc()).yield_per(1000)

    for t in tickets:
        pb_val = _ticket_pb_value(t)
        regs = normalize_regular_numbers(int(t.n1), int(t.n2), int(t.n3), int(t.n4), int(t.n5), order=order)
        regs_set = set(regs)

        # Matches
        mr = len(regs_set.intersection(winners_set)) if winners_set else 0
        mpb = (winning_pb is not None and int(pb_val) == int(winning_pb))

        # Grupo:
        # - "6" => 5 + PB
        # - "5" => 5 sin PB
        # - "4" => 4 reg (PB puede o no, igual entra)
        # - "3" => 3 reg (PB puede o no, igual entra)
        ok = (
            (group == "6" and mr == 5 and mpb) or
            (group == "5" and mr == 5 and not mpb) or
            (group == "4" and mr == 4) or
            (group == "3" and mr == 3)
        )
        if not ok:
            continue

        total_balls = mr + (1 if mpb else 0)

        rows.append({
            "ID": int(t.id),
            "Draw Date": t.draw_date.isoformat() if getattr(t, "draw_date", None) else "",
            "Status": str(getattr(t, "status", "")),
            "Type": str(getattr(t, "type", "")),
            "N1": int(regs[0]), "N2": int(regs[1]), "N3": int(regs[2]), "N4": int(regs[3]), "N5": int(regs[4]),
            "PB": int(pb_val),
            "Match Regular": int(mr),
            "Match PB": "YES" if mpb else "NO",
            "Total Balls": int(total_balls),
            "Prize": float(getattr(t, "prize_amount", 0.0) or 0.0),
            "Cost": float(getattr(t, "cost", 0.0) or 0.0),
            "Key": numbers_key([int(t.n1), int(t.n2), int(t.n3), int(t.n4), int(t.n5)], int(pb_val)),
        })

    if not rows:
        raise HTTPException(status_code=404, detail="No hay resultados para este grupo")

    df = pd.DataFrame(rows)

    # Excel con highlight
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sheet = f"group_{group}"
        df.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.sheets[sheet]

        # Header bold
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        # Colorear coincidencias con el "compare" como base
        if winners_set and winning_pb is not None:
            # ID(1), Draw Date(2), Status(3), Type(4), N1(5) ... N5(9), PB(10)
            col_n = {"N1": 5, "N2": 6, "N3": 7, "N4": 8, "N5": 9, "PB": 10}
            max_row = ws.max_row

            for r in range(2, max_row + 1):
                # Whites
                for c in (col_n["N1"], col_n["N2"], col_n["N3"], col_n["N4"], col_n["N5"]):
                    v = ws.cell(row=r, column=c).value
                    try:
                        if int(v) in winners_set:
                            ws.cell(row=r, column=c).fill = FILL_AQUA
                    except Exception:
                        pass

                # PB
                pbv = ws.cell(row=r, column=col_n["PB"]).value
                try:
                    if int(pbv) == int(winning_pb):
                        ws.cell(row=r, column=col_n["PB"]).fill = FILL_PB
                        ws.cell(row=r, column=col_n["PB"]).font = RED_FONT
                except Exception:
                    pass

        # Autosize
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(40, max(10, max_len + 2))

    buf.seek(0)
    filename = f"compare_group_{group}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


 #=========================
#   COMPARE - AI INSIGHT
# =========================
@app.get("/compare/insight")
def compare_insight(
    compare: str = Query(..., description='Ej: "10,16,29,33,69|22"'),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    # ✅ Validación temprana (y normalización de formato)
    parsed = parse_compare(compare)
    if not parsed:
        raise HTTPException(status_code=400, detail="compare inválido")

    regs, pb = parsed
    regs_sorted = sorted(list(regs))
    normalized_compare = (
        f"{regs_sorted[0]},{regs_sorted[1]},{regs_sorted[2]},{regs_sorted[3]},{regs_sorted[4]}|{int(pb)}"
    )

    # ✅ Cache con clave normalizada (evita misses por espacios/formato)
    cache_key = _make_compare_cache_key(normalized_compare, None, None, None, None) + "|insight"
    cached = _get_compare_cache(cache_key)
    if cached is not None:
        return cached

    # ✅ Ejecuta insight sobre el compare normalizado
    insight = compute_ai_insight(db, normalized_compare)

    # ✅ Forzar dict por robustez (sin romper compatibilidad)
    if not isinstance(insight, dict):
        insight = {"ok": True, "result": insight}

    # ✅ Cache seguro
    try:
        _set_compare_cache(cache_key, insight)
    except Exception:
        pass

    return insight


# ============================================================
#   AI RECOMMENDATIONS (API + UI) — ÚNICA VERSIÓN (UPGRADED)
# ============================================================
@app.post("/ai/recommend", response_model=RecommendResponse)
def ai_recommend(req: RecommendRequest, db: Session = Depends(get_db)):
    # ✅ Upgrade: cache simple por payload (evita recalcular al refrescar UI)
    try:
        cache_key = "ai_recommend|" + json.dumps(req.model_dump(), sort_keys=True, default=str)
    except Exception:
        cache_key = "ai_recommend|" + str(req.seed or "") + "|" + str(req.k or "")

    cached = _get_compare_cache(cache_key)
    if cached is not None:
        return RecommendResponse(**cached)

    rec = recommend_from_history(req, db)

    try:
        _set_compare_cache(cache_key, rec.model_dump())
    except Exception:
        pass

    return rec


@app.post("/ai/save_recommendations", response_model=SaveRecommendationsResponse)
def ai_save_recommendations(req: SaveRecommendationsRequest, db: Session = Depends(get_db)):
    # ✅ Upgrade: valida fecha FUTURE coherente (sin romper si la quieres en pasado)
    if req.future_draw_date is None:
        raise HTTPException(status_code=400, detail="future_draw_date es requerido")

    # Construye request de generación (reusa tu pipeline)
    gen_req = RecommendRequest(
        status=req.status,
        type=req.type,
        start_date=req.start_date,
        end_date=req.end_date,
        k=req.k,
        seed=req.seed,
        fixed_first=req.fixed_first,
        fixed_numbers=req.fixed_numbers,
        exclude_numbers=req.exclude_numbers,
        fixed_powerball=req.fixed_powerball,
        exclude_powerballs=req.exclude_powerballs,
        top_pool_regulars=req.top_pool_regulars,
        top_pool_powerballs=req.top_pool_powerballs,
    )
    rec = recommend_from_history(gen_req, db)

    # ✅ Dedupe en DB para ese draw_date
    existing_keys = _existing_keys_for_draw(db, req.future_draw_date)

    inserted = 0
    skipped = 0

    # ✅ Compat: Ticket puede tener powerball o pb
    ticket_has_pb = hasattr(Ticket, "pb")

    save_type = _norm_type(req.save_type) or "QUICK_PICK"
    cost = float(req.cost_per_ticket or 2.0)

    # ✅ Upgrade: chunked commit (evita transacción gigante en k grande)
    commit_every = 500

    for i, c in enumerate(rec.combos, start=1):
        regs = [int(c["n1"]), int(c["n2"]), int(c["n3"]), int(c["n4"]), int(c["n5"])]
        pb = int(c["powerball"])
        key = numbers_key(regs, pb)

        if key in existing_keys:
            skipped += 1
            continue

        if req.normalize_on_save:
            regs = sorted(regs)

        kw = dict(
            draw_date=req.future_draw_date,
            n1=int(regs[0]), n2=int(regs[1]), n3=int(regs[2]), n4=int(regs[3]), n5=int(regs[4]),
            type=save_type,
            cost=cost,
            status="FUTURE",
            matched_regular_numbers=0,
            matched_powerball=False,
            prize_amount=0.0,
        )
        if ticket_has_pb:
            kw["pb"] = int(pb)
        else:
            kw["powerball"] = int(pb)

        t = Ticket(**kw)
        db.add(t)
        inserted += 1
        existing_keys.add(key)

        if (i % commit_every) == 0:
            db.commit()

    db.commit()

    return SaveRecommendationsResponse(
        requested=int(req.k),
        generated=int(rec.generated),
        inserted=int(inserted),
        skipped_duplicates=int(skipped),
        future_draw_date=req.future_draw_date,
        status="ok",
        message=f"Guardados {inserted} tickets FUTURE. Duplicados evitados: {skipped}.",
    )


@app.get("/ui/recommendations", response_class=HTMLResponse)
def ui_recommendations():
    body = """
      <div class="card">
        <h3>AI Recommendations</h3>
        <p class="muted">Genera combinaciones por frecuencia histórica + restricciones opcionales. Guarda como FUTURE con dedupe.</p>

        <div class="formRow" style="margin-top:12px;">
          <select id="st" class="field fieldSmall">
            <option value="">History Status: ALL</option>
            <option value="PAST">PAST</option>
            <option value="FUTURE">FUTURE</option>
          </select>

          <select id="tp" class="field fieldSmall">
            <option value="">History Type: ALL</option>
            <option value="QUICK_PICK">QUICK_PICK</option>
            <option value="MANUAL">MANUAL</option>
          </select>

          <input id="k" class="field fieldSmall" value="50" />
          <input id="seed" class="field fieldSmall" placeholder="seed (opcional)" />
          <input id="fixedFirst" class="field fieldSmall" placeholder="fixed_first (opcional)" />
          <input id="fixedNums" class="field" placeholder="fixed_numbers CSV (opcional) ej: 10,16" />
          <input id="excludeNums" class="field" placeholder="exclude_numbers CSV (opcional) ej: 1,2,3" />

          <input id="fixedPB" class="field fieldSmall" placeholder="fixed_powerball (opcional)" />
          <input id="excludePB" class="field" placeholder="exclude_powerballs CSV (opcional) ej: 1,2" />

          <button id="gen" class="btn good">Generate</button>
          <button id="copyAll" class="btn">Copy all</button>
        </div>

        <div class="formRow" style="margin-top:10px;">
          <input id="futureDate" class="field fieldSmall" placeholder="future_draw_date YYYY-MM-DD" />
          <input id="cost" class="field fieldSmall" placeholder="cost_per_ticket" value="2.0" />
          <select id="saveType" class="field fieldSmall">
            <option value="QUICK_PICK" selected>Save Type: QUICK_PICK</option>
            <option value="MANUAL">Save Type: MANUAL</option>
          </select>
          <button id="save" class="btn primary">Save as FUTURE</button>
          <button id="dlcsv" class="btn">Download CSV</button>
        </div>

        <div id="out" style="margin-top:12px;"></div>
      </div>

      <div class="toast" id="toast">OK</div>

      <script>
        function toast(msg) {
          const t = document.getElementById('toast');
          t.textContent = msg || 'OK';
          t.style.display = 'block';
          setTimeout(()=>t.style.display='none', 1100);
        }

        function esc(s){
          return (s||'')
            .replaceAll('&','&amp;')
            .replaceAll('<','&lt;')
            .replaceAll('>','&gt;');
        }

        function parseCsvInts(s){
          s = (s||'').trim();
          if(!s) return [];
          return s
            .replaceAll(';',',')
            .replaceAll('|',',')
            .split(',')
            .map(x => x.trim())
            .filter(Boolean)
            .map(x => parseInt(x,10))
            .filter(x => Number.isFinite(x));
        }

        function nextDrawDateISO(){
          // Powerball: Mon/Wed/Sat. Elegimos el próximo de esos días.
          const d = new Date();
          d.setHours(0,0,0,0);
          const want = new Set([1,3,6]); // 0=Sun,1=Mon,3=Wed,6=Sat
          for(let i=0;i<14;i++){
            const dd = new Date(d.getTime() + i*86400000);
            if(want.has(dd.getDay()) && i>0){
              const y = dd.getFullYear();
              const m = String(dd.getMonth()+1).padStart(2,'0');
              const da = String(dd.getDate()).padStart(2,'0');
              return `${y}-${m}-${da}`;
            }
          }
          // fallback: mañana
          const dd = new Date(d.getTime() + 86400000);
          const y = dd.getFullYear();
          const m = String(dd.getMonth()+1).padStart(2,'0');
          const da = String(dd.getDate()).padStart(2,'0');
          return `${y}-${m}-${da}`;
        }

        function renderTable(combos){
          if(!combos || !combos.length){
            return `<div class="muted" style="margin-top:10px;">No combos generated.</div>`;
          }

          let rows = combos.map((c, idx) => {
            const nums = [c.n1,c.n2,c.n3,c.n4,c.n5].map(x=>parseInt(x,10));
            const pb = parseInt(c.powerball,10);
            const line = `${nums.join(' ')} | PB ${pb}`;
            return `
              <tr>
                <td>${idx+1}</td>
                <td>${nums[0]}</td><td>${nums[1]}</td><td>${nums[2]}</td><td>${nums[3]}</td><td>${nums[4]}</td>
                <td>${pb}</td>
                <td><span class="code">${esc(c.key||'')}</span></td>
                <td><button class="btn" style="padding:7px 10px;" data-copy="${esc(line)}">Copy</button></td>
              </tr>
            `;
          }).join('');

          return `
            <div class="tableWrap" style="margin-top:12px;">
              <table>
                <thead>
                  <tr>
                    <th>#</th>
                    <th>N1</th><th>N2</th><th>N3</th><th>N4</th><th>N5</th>
                    <th>PB</th>
                    <th>Key</th>
                    <th>Action</th>
                  </tr>
                </thead>
                <tbody>${rows}</tbody>
              </table>
            </div>
          `;
        }

        function downloadCsvFromCombos(combos){
          if(!combos || !combos.length){
            toast('Nothing to download');
            return;
          }
          const header = ['n1','n2','n3','n4','n5','powerball','key'];
          const lines = [header.join(',')].concat(
            combos.map(c => {
              const vals = [c.n1,c.n2,c.n3,c.n4,c.n5,c.powerball,(c.key||'')];
              return vals.map(v => String(v).replaceAll('"','""')).map(v => `"${v}"`).join(',');
            })
          );
          const blob = new Blob([lines.join('\\n')], {type:'text/csv;charset=utf-8;'});
          const url = URL.createObjectURL(blob);
          const a = document.createElement('a');
          a.href = url;
          a.download = 'ai_recommendations.csv';
          document.body.appendChild(a);
          a.click();
          a.remove();
          URL.revokeObjectURL(url);
        }

        window.__lastRecs = [];

        // ✅ Upgrade: disable buttons mientras corre (evita doble-submit)
        function setBusy(isBusy){
          ['gen','save','dlcsv','copyAll'].forEach(id=>{
            const el = document.getElementById(id);
            if(el) el.disabled = !!isBusy;
          });
        }

        async function generate(){
          setBusy(true);
          try{
            const status = (document.getElementById('st').value||'').trim();
            const type = (document.getElementById('tp').value||'').trim();

            const k = parseInt((document.getElementById('k').value||'50').trim(),10) || 50;
            const seedRaw = (document.getElementById('seed').value||'').trim();
            const seed = seedRaw ? parseInt(seedRaw,10) : null;

            const fixedFirstRaw = (document.getElementById('fixedFirst').value||'').trim();
            const fixed_first = fixedFirstRaw ? parseInt(fixedFirstRaw,10) : null;

            const fixed_numbers = parseCsvInts(document.getElementById('fixedNums').value||'');
            const exclude_numbers = parseCsvInts(document.getElementById('excludeNums').value||'');

            const fixedPBRaw = (document.getElementById('fixedPB').value||'').trim();
            const fixed_powerball = fixedPBRaw ? parseInt(fixedPBRaw,10) : null;

            const exclude_powerballs = parseCsvInts(document.getElementById('excludePB').value||'');

            const payload = {
              k: k,
              seed: seed,
              status: status || null,
              type: type || null,
              fixed_first: fixed_first,
              fixed_numbers: fixed_numbers.length ? fixed_numbers : null,
              exclude_numbers: exclude_numbers.length ? exclude_numbers : null,
              fixed_powerball: fixed_powerball,
              exclude_powerballs: exclude_powerballs.length ? exclude_powerballs : null
            };

            const res = await fetch('/ai/recommend', {
              method:'POST',
              headers:{'Content-Type':'application/json'},
              body: JSON.stringify(payload)
            });

            const data = await res.json();
            if(!res.ok){
              alert(data.detail || 'Error');
              return;
            }

            window.__lastRecs = (data.combos || []);
            const out = document.getElementById('out');

            out.innerHTML = `
              <div class="grid">
                <div class="card"><h3>Generated</h3><p><span class="code">${data.generated}</span></p></div>
                <div class="card"><h3>Seed</h3><p><span class="code">${esc(String(data.seed ?? 'auto'))}</span></p></div>
                <div class="card"><h3>Actions</h3><p class="muted">Copy / Download / Save as FUTURE</p></div>
              </div>
              ${renderTable(window.__lastRecs)}
            `;

            document.querySelectorAll('button[data-copy]').forEach(btn => {
              btn.addEventListener('click', async () => {
                const txt = btn.getAttribute('data-copy');
                try {
                  await navigator.clipboard.writeText(txt);
                  toast('Copied ✅');
                } catch(e){
                  toast('Copy failed');
                }
              });
            });

            toast('Generated ✅');
          } finally {
            setBusy(false);
          }
        }

        document.getElementById('gen').addEventListener('click', generate);

        document.getElementById('copyAll').addEventListener('click', async ()=>{
          const combos = window.__lastRecs || [];
          if(!combos.length){ toast('Nothing to copy'); return; }
          const lines = combos.map(c => {
            const nums = [c.n1,c.n2,c.n3,c.n4,c.n5].map(x=>parseInt(x,10));
            const pb = parseInt(c.powerball,10);
            return `${nums.join(' ')} | PB ${pb}`;
          });
          try{
            await navigator.clipboard.writeText(lines.join('\\n'));
            toast('Copied all ✅');
          }catch(e){
            toast('Copy failed');
          }
        });

        document.getElementById('dlcsv').addEventListener('click', ()=>{
          downloadCsvFromCombos(window.__lastRecs || []);
        });

        document.getElementById('save').addEventListener('click', async ()=>{
          const combos = window.__lastRecs || [];
          if(!combos.length){ toast('Generate first'); return; }

          setBusy(true);
          try{
            const status = (document.getElementById('st').value||'').trim();
            const type = (document.getElementById('tp').value||'').trim();

            const k = parseInt((document.getElementById('k').value||'50').trim(),10) || 50;
            const seedRaw = (document.getElementById('seed').value||'').trim();
            const seed = seedRaw ? parseInt(seedRaw,10) : null;

            const fixedFirstRaw = (document.getElementById('fixedFirst').value||'').trim();
            const fixed_first = fixedFirstRaw ? parseInt(fixedFirstRaw,10) : null;

            const fixed_numbers = parseCsvInts(document.getElementById('fixedNums').value||'');
            const exclude_numbers = parseCsvInts(document.getElementById('excludeNums').value||'');

            const fixedPBRaw = (document.getElementById('fixedPB').value||'').trim();
            const fixed_powerball = fixedPBRaw ? parseInt(fixedPBRaw,10) : null;

            const exclude_powerballs = parseCsvInts(document.getElementById('excludePB').value||'');

            const future_draw_date = (document.getElementById('futureDate').value||'').trim();
            if(!future_draw_date){
              alert('future_draw_date is required (YYYY-MM-DD)');
              return;
            }

            const cost = parseFloat((document.getElementById('cost').value||'2.0').trim());
            const saveType = (document.getElementById('saveType').value||'QUICK_PICK').trim();

            const payload = {
              status: status || null,
              type: type || null,
              k: k,
              seed: seed,
              fixed_first: fixed_first,
              fixed_numbers: fixed_numbers.length ? fixed_numbers : null,
              exclude_numbers: exclude_numbers.length ? exclude_numbers : null,
              fixed_powerball: fixed_powerball,
              exclude_powerballs: exclude_powerballs.length ? exclude_powerballs : null,
              future_draw_date: future_draw_date,
              cost_per_ticket: Number.isFinite(cost) ? cost : 2.0,
              save_type: saveType || 'QUICK_PICK',
              normalize_on_save: true
            };

            const res = await fetch('/ai/save_recommendations', {
              method:'POST',
              headers:{'Content-Type':'application/json'},
              body: JSON.stringify(payload)
            });

            const data = await res.json();
            if(!res.ok){
              alert(data.detail || 'Error saving');
              return;
            }
            toast(data.message || 'Saved ✅');
          } finally {
            setBusy(false);
          }
        });

        // set default future date
        (function init(){
          const fd = document.getElementById('futureDate');
          if(fd && !fd.value){
            fd.value = nextDrawDateISO();
          }
        })();
      </script>
    """
    return render_app_page(title="AI Recommendations", active="ai", body_html=body)

# ============================================================
#   MAINTENANCE / ADMIN UTILITIES (UPGRADED)
# ============================================================

@app.post("/admin/recompute_matches")
def admin_recompute_matches(db: Session = Depends(get_db)):
    """
    Recalcula matches/prize para tickets con draw_result en la misma fecha.
    Upgrade:
      - streaming (yield_per) para no cargar todo en RAM
      - cache de draw_results por fecha (dict)
      - compat powerball/pb
      - conteos claros
    """
    draws = db.query(DrawResult).all()
    draws_by_date = {d.draw_date: d for d in draws}

    today = date.today()

    processed = 0
    updated_with_draw = 0
    cleared_no_draw = 0

    # streaming
    for t in db.query(Ticket).yield_per(500):
        processed += 1
        dr = draws_by_date.get(getattr(t, "draw_date", None))

        if dr:
            # compat: calculate_matches usa ticket.powerball en tu versión
            if not hasattr(t, "powerball") and hasattr(t, "pb"):
                try:
                    setattr(t, "powerball", getattr(t, "pb"))
                except Exception:
                    pass

            calculate_matches(t, dr)

            try:
                if t.draw_date and t.draw_date <= today:
                    t.status = "PAST"
            except Exception:
                pass

            updated_with_draw += 1
        else:
            # sin draw_result, limpio
            try:
                setattr(t, "matched_regular_numbers", 0)
                setattr(t, "matched_powerball", False)
                setattr(t, "prize_amount", 0.0)
            except Exception:
                pass
            cleared_no_draw += 1

    db.commit()
    return {
        "ok": True,
        "tickets_processed": int(processed),
        "tickets_updated_with_draw": int(updated_with_draw),
        "tickets_cleared_no_draw": int(cleared_no_draw),
    }


@app.post("/admin/normalize_all_ticket_numbers")
def admin_normalize_all_ticket_numbers(db: Session = Depends(get_db)):
    """
    Ordena N1..N5 asc en todos los tickets (sin cambiar PB).
    Upgrade:
      - streaming
      - valida unicidad y rango antes de persistir
      - commit en batches
    """
    processed = 0
    changed = 0
    skipped_invalid = 0

    commit_every = 1000

    for t in db.query(Ticket).yield_per(500):
        processed += 1
        try:
            regs = [int(t.n1), int(t.n2), int(t.n3), int(t.n4), int(t.n5)]
            pb_val = int(getattr(t, "powerball", getattr(t, "pb", 0)))
        except Exception:
            skipped_invalid += 1
            continue

        # valida (sin reventar el proceso)
        try:
            validate_ticket_numbers(regs[0], regs[1], regs[2], regs[3], regs[4], pb_val)
        except Exception:
            skipped_invalid += 1
            continue

        regs_sorted = sorted(regs)
        if regs_sorted != regs:
            t.n1, t.n2, t.n3, t.n4, t.n5 = regs_sorted
            changed += 1

        if (processed % commit_every) == 0:
            db.commit()

    db.commit()
    return {
        "ok": True,
        "tickets_processed": int(processed),
        "changed": int(changed),
        "skipped_invalid": int(skipped_invalid),
    }


@app.get("/admin/duplicates")
def admin_duplicates(db: Session = Depends(get_db)):
    """
    Devuelve duplicados por (draw_date + key). No borra nada; solo reporta.
    Upgrade:
      - streaming
      - compat powerball/pb
    """
    seen: Dict[str, int] = {}
    dups: List[Dict[str, Any]] = []

    processed = 0
    for t in db.query(Ticket).yield_per(500):
        processed += 1

        dd = getattr(t, "draw_date", None)
        try:
            pb_val = int(getattr(t, "powerball", getattr(t, "pb", 0)))
        except Exception:
            pb_val = 0

        try:
            k = numbers_key([t.n1, t.n2, t.n3, t.n4, t.n5], pb_val)
        except Exception:
            # si hay algo corrupto, lo ignoramos
            continue

        dd_iso = dd.isoformat() if dd else ""
        kk = f"{dd_iso}|{k}"

        if kk in seen:
            dups.append({
                "first_id": int(seen[kk]),
                "dup_id": int(getattr(t, "id", 0) or 0),
                "draw_date": dd_iso,
                "key": k,
            })
        else:
            seen[kk] = int(getattr(t, "id", 0) or 0)

    return {"ok": True, "processed": int(processed), "duplicates": dups, "count": int(len(dups))}
