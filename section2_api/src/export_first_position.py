from __future__ import annotations

import csv
import json
import os
import random
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ------------------------
# Paths / Defaults
# ------------------------

PROJECT_ROOT = Path(__file__).resolve().parents[1]  # .../powerball_api
DATA_DIR = PROJECT_ROOT / "data"
SQLITE_PATH = DATA_DIR / "powerball.db"
CSV_PATH = DATA_DIR / "powerball.csv"
DB_PATH = SQLITE_PATH  # canonical sqlite for this project


# ------------------------
# EXPORT — First position
# ------------------------

def export_by_first_position(
    limit: int = 69,
    source: str = "auto",
) -> Dict[str, Any]:
    """
    Devuelve un resumen de frecuencia del "primer número" (white1/ball1)
    de los sorteos de Powerball.

    Estrategia de carga (source="auto"):
      1) DATABASE_URL (si existe)
      2) SQLite en ./data/powerball.db
      3) CSV en ./data/powerball.csv
    """
    if limit <= 0:
        limit = 69

    rows, used_source = _load_draw_rows(source)

    if not rows:
        return {
            "status": "error",
            "error": "NO_DATA",
            "message": (
                "No encontré datos para exportar. Probé DATABASE_URL, SQLite y CSV.\n"
                "Opciones:\n"
                " - Coloca SQLite en ./data/powerball.db con una tabla 'draws'\n"
                " - O coloca CSV en ./data/powerball.csv con columna white1/ball1\n"
                " - O define DATABASE_URL para una DB."
            ),
            "meta": {
                "expected_sqlite": str(SQLITE_PATH),
                "expected_csv": str(CSV_PATH),
                "database_url_set": bool(os.getenv("DATABASE_URL")),
            },
        }

    counts: Dict[int, int] = {}
    total = 0

    for (white1,) in rows:
        if white1 is None:
            continue
        try:
            n = int(white1)
        except Exception:
            continue
        counts[n] = counts.get(n, 0) + 1
        total += 1

    data = []
    for number in sorted(counts.keys()):
        c = counts[number]
        pct = round((c / total) * 100, 4) if total else 0.0
        data.append({"number": number, "count": c, "pct": pct})

    data_sorted = sorted(data, key=lambda x: x["count"], reverse=True)[:limit]

    return {
        "status": "ok",
        "source": used_source,
        "total_draws": total,
        "data": data_sorted,
        "meta": {
            "limit": limit,
            "note": "Frecuencia del primer número blanco (white1/ball1).",
        },
    }


# ------------------------
# Load pipeline (auto)
# ------------------------

def _load_draw_rows(source: str) -> Tuple[List[Tuple[Optional[int]]], str]:
    """
    Retorna lista de tuplas (white1,) y un string con la fuente usada.
    """
    source = (source or "auto").lower().strip()

    if source in ("auto", "db", "database", "database_url"):
        db_url = os.getenv("DATABASE_URL")
        if db_url:
            try:
                rows = _load_from_database_url(db_url)
                if rows:
                    return rows, "DATABASE_URL"
            except Exception:
                pass

    if source in ("auto", "sqlite"):
        if SQLITE_PATH.exists():
            try:
                rows = _load_from_sqlite(SQLITE_PATH)
                if rows:
                    return rows, f"sqlite:{SQLITE_PATH}"
            except Exception:
                pass

    if source in ("auto", "csv"):
        if CSV_PATH.exists():
            try:
                rows = _load_from_csv(CSV_PATH)
                if rows:
                    return rows, f"csv:{CSV_PATH}"
            except Exception:
                pass

    return [], "none"


def _sqlite_columns(conn: sqlite3.Connection, table: str) -> List[str]:
    cur = conn.cursor()
    cur.execute(f"PRAGMA table_info({table});")
    return [row[1] for row in cur.fetchall()]


def _load_from_sqlite(path: Path) -> List[Tuple[Optional[int]]]:
    """
    Espera una tabla 'draws' con alguna columna de primer número:
      - white1  (preferida)
      - ball1
      - n1
    """
    conn = sqlite3.connect(str(path))
    try:
        conn.row_factory = sqlite3.Row
        cols = _sqlite_columns(conn, "draws")

        first_col = None
        for candidate in ("white1", "ball1", "n1"):
            if candidate in cols:
                first_col = candidate
                break

        if not first_col:
            raise RuntimeError(
                f"SQLite: tabla 'draws' no tiene white1/ball1/n1. Tiene: {sorted(cols)}"
            )

        cur = conn.cursor()
        cur.execute(f"SELECT {first_col} AS white1 FROM draws WHERE {first_col} IS NOT NULL;")
        rows = [(r["white1"],) for r in cur.fetchall()]
        return rows
    finally:
        conn.close()


def _load_from_csv(path: Path) -> List[Tuple[Optional[int]]]:
    """
    CSV esperado con header y columna:
      - white1 (preferida) o ball1 o n1
    """
    with path.open("r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return []

        cols = {c.strip().lower(): c for c in reader.fieldnames}
        first_key = None
        for candidate in ("white1", "ball1", "n1"):
            if candidate in cols:
                first_key = cols[candidate]
                break

        if not first_key:
            raise RuntimeError(f"CSV: no tiene columnas white1/ball1/n1. Tiene: {reader.fieldnames}")

        out: List[Tuple[Optional[int]]] = []
        for row in reader:
            raw = row.get(first_key)
            if raw is None or raw == "":
                continue
            try:
                out.append((int(raw),))
            except Exception:
                continue
        return out


def _load_from_database_url(db_url: str) -> List[Tuple[Optional[int]]]:
    """
    Soporta Postgres con psycopg2 o psycopg (si está instalado).
    Espera tabla 'draws' con columna white1/ball1/n1.
    """
    parsed = urlparse(db_url)
    scheme = (parsed.scheme or "").lower()

    if scheme.startswith("postgres"):
        try:
            import psycopg2  # type: ignore
            return _load_from_postgres_psycopg2(db_url)
        except Exception:
            try:
                import psycopg  # type: ignore
                return _load_from_postgres_psycopg(db_url)
            except Exception as e:
                raise RuntimeError(
                    "DATABASE_URL es Postgres pero no tienes instalado psycopg2 o psycopg."
                ) from e

    raise RuntimeError(f"DATABASE_URL scheme no soportado: {scheme}")


def _load_from_postgres_psycopg2(db_url: str) -> List[Tuple[Optional[int]]]:
    import psycopg2  # type: ignore

    conn = psycopg2.connect(db_url)
    try:
        cur = conn.cursor()
        for col in ("white1", "ball1", "n1"):
            try:
                cur.execute(f"SELECT {col} AS white1 FROM draws WHERE {col} IS NOT NULL;")
                rows = cur.fetchall()
                if rows:
                    return [(r[0],) for r in rows]
            except Exception:
                continue
        raise RuntimeError("Postgres: no encuentro columnas white1/ball1/n1 en draws.")
    finally:
        conn.close()


def _load_from_postgres_psycopg(db_url: str) -> List[Tuple[Optional[int]]]:
    import psycopg  # type: ignore

    with psycopg.connect(db_url) as conn:
        with conn.cursor() as cur:
            for col in ("white1", "ball1", "n1"):
                try:
                    cur.execute(f"SELECT {col} AS white1 FROM draws WHERE {col} IS NOT NULL;")
                    rows = cur.fetchall()
                    if rows:
                        return [(r[0],) for r in rows]
                except Exception:
                    continue
    raise RuntimeError("Postgres: no encuentro columnas white1/ball1/n1 en draws.")


# ------------------------
# HISTORY — Read-only
# ------------------------

def list_draws_by_position(position: int, number: int, limit: int = 5000) -> Dict[str, Any]:
    """
    position: 1..5  (white1..white5)
    number:   número a filtrar
    """
    if position not in (1, 2, 3, 4, 5):
        return {"status": "error", "error": "INVALID_POSITION", "message": "position debe ser 1..5"}

    col = f"white{position}"

    conn = sqlite3.connect(str(DB_PATH))
    try:
        cur = conn.cursor()
        cur.execute(
            f"""
            SELECT draw_date, white1, white2, white3, white4, white5, powerball
            FROM draws
            WHERE {col} = ?
            ORDER BY draw_date ASC
            LIMIT ?
            """,
            (int(number), int(limit)),
        )
        rows = cur.fetchall()

        data: List[Dict[str, Any]] = []
        for r in rows:
            data.append({
                "draw_date": r[0],
                "white1": r[1],
                "white2": r[2],
                "white3": r[3],
                "white4": r[4],
                "white5": r[5],
                "powerball": r[6],
            })

        return {
            "status": "ok",
            "position": position,
            "number": int(number),
            "count": len(data),
            "data": data,
            "meta": {"limit": int(limit), "db": str(DB_PATH)},
        }
    finally:
        conn.close()


def _format_lines(item: Dict[str, Any]) -> str:
    return (
        f"{item['draw_date']} | "
        f"{item['white1']}-{item['white2']}-{item['white3']}-{item['white4']}-{item['white5']} "
        f"PB:{item['powerball']}"
    )


def list_draws_filtered(
    white1: Optional[int] = None,
    white2: Optional[int] = None,
    white3: Optional[int] = None,
    white4: Optional[int] = None,
    white5: Optional[int] = None,
    powerball: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    complete: bool = True,
    sort: str = "draw_date",
    direction: str = "asc",
    limit: int = 5000,
    output: str = "json",  # "json" | "lines"
) -> Dict[str, Any]:

    where: List[str] = []
    params: List[Any] = []

    def add_eq(col: str, val: Optional[int]):
        if val is None:
            return
        where.append(f"{col} = ?")
        params.append(int(val))

    add_eq("white1", white1)
    add_eq("white2", white2)
    add_eq("white3", white3)
    add_eq("white4", white4)
    add_eq("white5", white5)
    add_eq("powerball", powerball)

    if date_from:
        where.append("draw_date >= ?")
        params.append(date_from)
    if date_to:
        where.append("draw_date <= ?")
        params.append(date_to)

    if complete:
        where.extend([
            "white1 IS NOT NULL",
            "white2 IS NOT NULL",
            "white3 IS NOT NULL",
            "white4 IS NOT NULL",
            "white5 IS NOT NULL",
            "powerball IS NOT NULL",
        ])

    allowed_sort = {
        "draw_date": "draw_date",
        "white1": "white1",
        "white2": "white2",
        "white3": "white3",
        "white4": "white4",
        "white5": "white5",
        "powerball": "powerball",
    }
    order_col = allowed_sort.get((sort or "").lower(), "draw_date")
    dir_sql = "DESC" if (direction or "").lower() == "desc" else "ASC"

    sql = """
        SELECT draw_date, white1, white2, white3, white4, white5, powerball
        FROM draws
    """
    if where:
        sql += " WHERE " + " AND ".join(where)

    sql += f" ORDER BY {order_col} {dir_sql} LIMIT ?"
    params.append(int(limit))

    conn = sqlite3.connect(str(DB_PATH))
    try:
        cur = conn.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()

        data: List[Dict[str, Any]] = []
        lines: List[str] = []

        for r in rows:
            item = {
                "draw_date": r[0],
                "white1": r[1],
                "white2": r[2],
                "white3": r[3],
                "white4": r[4],
                "white5": r[5],
                "powerball": r[6],
            }
            data.append(item)
            lines.append(_format_lines(item))

        resp: Dict[str, Any] = {
            "status": "ok",
            "count": len(data),
            "filters": {
                "white1": white1,
                "white2": white2,
                "white3": white3,
                "white4": white4,
                "white5": white5,
                "powerball": powerball,
                "date_from": date_from,
                "date_to": date_to,
                "complete": complete,
                "sort": order_col,
                "direction": dir_sql.lower(),
                "output": output,
            },
            "meta": {"limit": int(limit), "db": str(DB_PATH)},
        }

        if (output or "").lower() == "lines":
            resp["lines"] = lines
        else:
            resp["data"] = data

        return resp
    finally:
        conn.close()


def list_draws_filtered_or(
    white1: Optional[int] = None,
    white2: Optional[int] = None,
    white3: Optional[int] = None,
    white4: Optional[int] = None,
    white5: Optional[int] = None,
    powerball: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    complete: bool = True,
    sort: str = "draw_date",
    direction: str = "asc",
    limit: int = 5000,
    output: str = "json",  # "json" | "lines"
) -> Dict[str, Any]:

    ors: List[str] = []
    params: List[Any] = []

    def add_or_eq(col: str, val: Optional[int]):
        if val is None:
            return
        ors.append(f"{col} = ?")
        params.append(int(val))

    add_or_eq("white1", white1)
    add_or_eq("white2", white2)
    add_or_eq("white3", white3)
    add_or_eq("white4", white4)
    add_or_eq("white5", white5)
    add_or_eq("powerball", powerball)

    where: List[str] = []
    if ors:
        where.append("(" + " OR ".join(ors) + ")")

    if date_from:
        where.append("draw_date >= ?")
        params.append(date_from)
    if date_to:
        where.append("draw_date <= ?")
        params.append(date_to)

    if complete:
        where.extend([
            "white1 IS NOT NULL",
            "white2 IS NOT NULL",
            "white3 IS NOT NULL",
            "white4 IS NOT NULL",
            "white5 IS NOT NULL",
            "powerball IS NOT NULL",
        ])

    allowed_sort = {
        "draw_date": "draw_date",
        "white1": "white1",
        "white2": "white2",
        "white3": "white3",
        "white4": "white4",
        "white5": "white5",
        "powerball": "powerball",
    }
    order_col = allowed_sort.get((sort or "").lower(), "draw_date")
    dir_sql = "DESC" if (direction or "").lower() == "desc" else "ASC"

    sql = """
        SELECT draw_date, white1, white2, white3, white4, white5, powerball
        FROM draws
    """
    if where:
        sql += " WHERE " + " AND ".join(where)

    sql += f" ORDER BY {order_col} {dir_sql} LIMIT ?"
    params.append(int(limit))

    conn = sqlite3.connect(str(DB_PATH))
    try:
        cur = conn.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()

        data: List[Dict[str, Any]] = []
        lines: List[str] = []

        for r in rows:
            item = {
                "draw_date": r[0],
                "white1": r[1],
                "white2": r[2],
                "white3": r[3],
                "white4": r[4],
                "white5": r[5],
                "powerball": r[6],
            }
            data.append(item)
            lines.append(_format_lines(item))

        resp: Dict[str, Any] = {
            "status": "ok",
            "count": len(data),
            "mode": "or",
            "filters": {
                "white1": white1,
                "white2": white2,
                "white3": white3,
                "white4": white4,
                "white5": white5,
                "powerball": powerball,
                "date_from": date_from,
                "date_to": date_to,
                "complete": complete,
                "sort": order_col,
                "direction": dir_sql.lower(),
                "output": output,
            },
            "meta": {"limit": int(limit), "db": str(DB_PATH)},
        }

        if (output or "").lower() == "lines":
            resp["lines"] = lines
        else:
            resp["data"] = data

        return resp
    finally:
        conn.close()


def list_draws_filtered_atleast(
    white1: Optional[int] = None,
    white2: Optional[int] = None,
    white3: Optional[int] = None,
    white4: Optional[int] = None,
    white5: Optional[int] = None,
    powerball: Optional[int] = None,
    min_match: int = 2,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    complete: bool = True,
    sort: str = "score",
    direction: str = "desc",
    limit: int = 5000,
    output: str = "json",  # "json" | "lines"
) -> Dict[str, Any]:

    checks: List[str] = []
    params_score: List[Any] = []

    def add_check(col: str, val: Optional[int]):
        if val is None:
            return
        checks.append(f"(CASE WHEN {col} = ? THEN 1 ELSE 0 END)")
        params_score.append(int(val))

    add_check("white1", white1)
    add_check("white2", white2)
    add_check("white3", white3)
    add_check("white4", white4)
    add_check("white5", white5)
    add_check("powerball", powerball)

    if not checks:
        return {"status": "error", "error": "NO_FILTERS", "message": "Debes enviar al menos un filtro."}

    score_expr = " + ".join(checks)

    where: List[str] = []
    params_where: List[Any] = []

    if date_from:
        where.append("draw_date >= ?")
        params_where.append(date_from)
    if date_to:
        where.append("draw_date <= ?")
        params_where.append(date_to)

    if complete:
        where.extend([
            "white1 IS NOT NULL",
            "white2 IS NOT NULL",
            "white3 IS NOT NULL",
            "white4 IS NOT NULL",
            "white5 IS NOT NULL",
            "powerball IS NOT NULL",
        ])

    allowed_sort = {
        "draw_date": "draw_date",
        "score": "score",
        "white1": "white1",
        "white2": "white2",
        "white3": "white3",
        "white4": "white4",
        "white5": "white5",
        "powerball": "powerball",
    }
    order_col = allowed_sort.get((sort or "").lower(), "score")
    dir_sql = "DESC" if (direction or "").lower() == "desc" else "ASC"

    inner_sql = f"""
        SELECT
            draw_date, white1, white2, white3, white4, white5, powerball,
            ({score_expr}) AS score
        FROM draws
    """
    if where:
        inner_sql += " WHERE " + " AND ".join(where)

    sql = f"""
        SELECT draw_date, white1, white2, white3, white4, white5, powerball, score
        FROM ({inner_sql})
        WHERE score >= ?
        ORDER BY {order_col} {dir_sql}
        LIMIT ?
    """

    params: List[Any] = []
    params.extend(params_score)
    params.extend(params_where)
    params.append(int(min_match))
    params.append(int(limit))

    conn = sqlite3.connect(str(DB_PATH))
    try:
        cur = conn.cursor()
        cur.execute(sql, params)
        rows = cur.fetchall()

        data: List[Dict[str, Any]] = []
        lines: List[str] = []

        for r in rows:
            item = {
                "draw_date": r[0],
                "white1": r[1],
                "white2": r[2],
                "white3": r[3],
                "white4": r[4],
                "white5": r[5],
                "powerball": r[6],
                "score": r[7],
            }
            data.append(item)
            lines.append(_format_lines(item) + f" | score:{item['score']}")

        resp: Dict[str, Any] = {
            "status": "ok",
            "count": len(data),
            "mode": "atleast",
            "filters": {
                "white1": white1,
                "white2": white2,
                "white3": white3,
                "white4": white4,
                "white5": white5,
                "powerball": powerball,
                "min_match": int(min_match),
                "date_from": date_from,
                "date_to": date_to,
                "complete": complete,
                "sort": order_col,
                "direction": dir_sql.lower(),
                "output": output,
            },
            "meta": {"limit": int(limit), "db": str(DB_PATH)},
        }

        if (output or "").lower() == "lines":
            resp["lines"] = lines
        else:
            resp["data"] = data

        return resp
    finally:
        conn.close()


# ------------------------
# FUTURE — Editable
# ------------------------

def _normalize_sort_direction(direction: str) -> str:
    d = (direction or "asc").lower()
    return "desc" if d == "desc" else "asc"


def _allowed_sort_fields_future() -> set:
    return {
        "id", "draw_date",
        "white1", "white2", "white3", "white4", "white5",
        "powerball",
        "created_at",
        "score",
    }


def create_future_quickpicks(
    n: int = 1,
    draw_date: Optional[str] = None,
    seed: Optional[int] = None,
    meta: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """
    Genera N QuickPicks:
    - 5 whites únicos [1..69]
    - 1 powerball [1..26]
    Inserta en future_draws (dedup por UNIQUE).
    """
    if n < 1:
        n = 1
    if n > 500:
        n = 500

    rng = random.Random(seed)

    inserted = 0
    skipped_duplicates = 0
    created_rows: List[Dict[str, Any]] = []

    meta_obj = meta or {"source": "quickpick"}
    meta_text = json.dumps(meta_obj, ensure_ascii=False)

    con = sqlite3.connect(str(DB_PATH))
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    for _ in range(n):
        whites = sorted(rng.sample(range(1, 70), 5))
        pb = rng.randint(1, 26)

        try:
            cur.execute(
                """
                INSERT INTO future_draws (draw_date, white1, white2, white3, white4, white5, powerball, meta)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (draw_date, whites[0], whites[1], whites[2], whites[3], whites[4], pb, meta_text),
            )
            inserted += 1
            created_rows.append({
                "draw_date": draw_date,
                "white1": whites[0], "white2": whites[1], "white3": whites[2], "white4": whites[3], "white5": whites[4],
                "powerball": pb,
            })
        except sqlite3.IntegrityError:
            skipped_duplicates += 1

    con.commit()
    con.close()

    return {
        "status": "ok",
        "requested": n,
        "inserted": inserted,
        "skipped_duplicates": skipped_duplicates,
        "data": created_rows,
    }


def list_future_filtered(
    white1: Optional[int] = None,
    white2: Optional[int] = None,
    white3: Optional[int] = None,
    white4: Optional[int] = None,
    white5: Optional[int] = None,
    powerball: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    complete: int = 0,
    output: str = "json",
    sort: str = "created_at",
    direction: str = "desc",
    limit: int = 200,
) -> Dict[str, Any]:
    direction = _normalize_sort_direction(direction)
    sort = (sort or "created_at").strip()
    if sort not in _allowed_sort_fields_future():
        sort = "created_at"

    where = []
    params: List[Any] = []

    def add_eq(col, val):
        if val is not None:
            where.append(f"{col} = ?")
            params.append(int(val))

    add_eq("white1", white1)
    add_eq("white2", white2)
    add_eq("white3", white3)
    add_eq("white4", white4)
    add_eq("white5", white5)
    add_eq("powerball", powerball)

    if date_from:
        where.append("draw_date >= ?")
        params.append(date_from)
    if date_to:
        where.append("draw_date <= ?")
        params.append(date_to)

    if int(complete) == 1:
        where.append("white1 IS NOT NULL AND white2 IS NOT NULL AND white3 IS NOT NULL AND white4 IS NOT NULL AND white5 IS NOT NULL AND powerball IS NOT NULL")

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    sql = f"""
    SELECT id, draw_date, white1, white2, white3, white4, white5, powerball, created_at, meta
    FROM future_draws
    {where_sql}
    ORDER BY {sort} {direction}
    LIMIT ?
    """
    params.append(int(limit))

    con = sqlite3.connect(str(DB_PATH))
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    rows = cur.execute(sql, params).fetchall()
    con.close()

    data = [dict(r) for r in rows]

    resp: Dict[str, Any] = {
        "status": "ok",
        "count": len(data),
        "data": data,
    }

    if (output or "").lower() == "lines":
        resp.pop("data", None)
        resp["lines"] = [
            f"{r['draw_date'] or 'NULL'} | {r['white1']}-{r['white2']}-{r['white3']}-{r['white4']}-{r['white5']} PB:{r['powerball']} | id={r['id']}"
            for r in data
        ]

    return resp


def list_future_filtered_or(
    white1: Optional[int] = None,
    white2: Optional[int] = None,
    white3: Optional[int] = None,
    white4: Optional[int] = None,
    white5: Optional[int] = None,
    powerball: Optional[int] = None,
    complete: int = 0,
    output: str = "json",
    sort: str = "created_at",
    direction: str = "desc",
    limit: int = 200,
) -> Dict[str, Any]:
    direction = _normalize_sort_direction(direction)
    sort = (sort or "created_at").strip()
    if sort not in _allowed_sort_fields_future():
        sort = "created_at"

    ors = []
    params: List[Any] = []

    def add_eq(col, val):
        if val is not None:
            ors.append(f"{col} = ?")
            params.append(int(val))

    add_eq("white1", white1)
    add_eq("white2", white2)
    add_eq("white3", white3)
    add_eq("white4", white4)
    add_eq("white5", white5)
    add_eq("powerball", powerball)

    if not ors:
        return {"status": "ok", "count": 0, "data": []}

    where = [("(" + " OR ".join(ors) + ")")]

    if int(complete) == 1:
        where.append("white1 IS NOT NULL AND white2 IS NOT NULL AND white3 IS NOT NULL AND white4 IS NOT NULL AND white5 IS NOT NULL AND powerball IS NOT NULL")

    where_sql = "WHERE " + " AND ".join(where)

    sql = f"""
    SELECT id, draw_date, white1, white2, white3, white4, white5, powerball, created_at, meta
    FROM future_draws
    {where_sql}
    ORDER BY {sort} {direction}
    LIMIT ?
    """
    params.append(int(limit))

    con = sqlite3.connect(str(DB_PATH))
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    rows = cur.execute(sql, params).fetchall()
    con.close()

    data = [dict(r) for r in rows]

    resp: Dict[str, Any] = {"status": "ok", "count": len(data), "data": data}

    if (output or "").lower() == "lines":
        resp.pop("data", None)
        resp["lines"] = [
            f"{r['draw_date'] or 'NULL'} | {r['white1']}-{r['white2']}-{r['white3']}-{r['white4']}-{r['white5']} PB:{r['powerball']} | id={r['id']}"
            for r in data
        ]

    return resp


def list_future_filtered_atleast(
    white1: Optional[int] = None,
    white2: Optional[int] = None,
    white3: Optional[int] = None,
    white4: Optional[int] = None,
    white5: Optional[int] = None,
    powerball: Optional[int] = None,
    min_match: int = 2,
    complete: int = 0,
    output: str = "json",
    sort: str = "score",
    direction: str = "desc",
    limit: int = 200,
) -> Dict[str, Any]:
    direction = _normalize_sort_direction(direction)
    sort = (sort or "score").strip()
    if sort not in _allowed_sort_fields_future():
        sort = "score"

    conds: List[Tuple[str, int]] = []
    params: List[Any] = []

    def add_cond(col, val):
        if val is not None:
            conds.append((col, int(val)))

    add_cond("white1", white1)
    add_cond("white2", white2)
    add_cond("white3", white3)
    add_cond("white4", white4)
    add_cond("white5", white5)
    add_cond("powerball", powerball)

    if not conds:
        return {"status": "ok", "count": 0, "data": []}

    if min_match < 1:
        min_match = 1
    if min_match > len(conds):
        min_match = len(conds)

    score_parts = []
    for col, val in conds:
        score_parts.append(f"(CASE WHEN {col} = ? THEN 1 ELSE 0 END)")
        params.append(val)

    score_expr = " + ".join(score_parts)

    base_where = []
    if int(complete) == 1:
        base_where.append("white1 IS NOT NULL AND white2 IS NOT NULL AND white3 IS NOT NULL AND white4 IS NOT NULL AND white5 IS NOT NULL AND powerball IS NOT NULL")

    base_where_sql = ("WHERE " + " AND ".join(base_where)) if base_where else ""

    sql = f"""
    SELECT *
    FROM (
      SELECT
        id, draw_date, white1, white2, white3, white4, white5, powerball, created_at, meta,
        ({score_expr}) AS score
      FROM future_draws
      {base_where_sql}
    ) t
    WHERE score >= ?
    ORDER BY {sort} {direction}
    LIMIT ?
    """
    params.append(int(min_match))
    params.append(int(limit))

    con = sqlite3.connect(str(DB_PATH))
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    rows = cur.execute(sql, params).fetchall()
    con.close()

    data = [dict(r) for r in rows]

    resp: Dict[str, Any] = {"status": "ok", "count": len(data), "data": data}

    if (output or "").lower() == "lines":
        resp.pop("data", None)
        resp["lines"] = [
            f"score={r['score']} | {r['draw_date'] or 'NULL'} | {r['white1']}-{r['white2']}-{r['white3']}-{r['white4']}-{r['white5']} PB:{r['powerball']} | id={r['id']}"
            for r in data
        ]

    return resp


# ------------------------
# FUTURE — Import Excel
# ------------------------

def import_future_from_excel(
    file_path: str,
    sheet_name: str | None = None,
) -> Dict[str, Any]:
    """
    Importa draws FUTUROS desde Excel.

    Formatos aceptados (headers case-insensitive; se normaliza a lower+strip):
      A) OFICIAL:
         drawdate, ball1, ball2, ball3, ball4, ball5, powerball
      B) Tu Excel actual:
         n1, n2, n3, n4, n5, powerball  (y opcional drawdate/draw_date)
      C) Canónico alterno:
         draw_date, white1, white2, white3, white4, white5, powerball

    Inserta en SQLite future_draws (esquema interno):
      draw_date, white1..white5, powerball, created_at, meta

    Reglas:
      - Si el combo existe en histórico (draws), se salta.
      - Si el combo existe en future_draws, se salta (dedup).
      - Canonicaliza whites (orden asc).
      - Reporta invalid_rows y errors_preview.
    """

    def _default_db_path() -> str:
        here = os.path.dirname(os.path.abspath(__file__))
        return os.path.join(os.path.dirname(here), "data", "powerball.db")

    def _resolve_db_path() -> str:
        db_url = os.getenv("DATABASE_URL")
        if not db_url:
            return _default_db_path()
        if db_url.startswith("sqlite:"):
            return db_url.replace("sqlite:///", "/").replace("sqlite:/", "/")
        return db_url

    def _norm_cols(cols) -> list[str]:
        return [str(c).strip().lower() for c in cols]

    def _pick_col(df_cols: set[str], candidates: list[str]) -> str | None:
        for c in candidates:
            if c in df_cols:
                return c
        return None

    def _to_int(x):
        if x is None:
            return None
        if isinstance(x, float) and pd.isna(x):
            return None
        if isinstance(x, str) and x.strip() == "":
            return None
        try:
            return int(float(x))
        except Exception:
            return None

    def _parse_date(x) -> str | None:
        if x is None:
            return None
        if isinstance(x, float) and pd.isna(x):
            return None
        if isinstance(x, pd.Timestamp):
            return x.date().isoformat()
        if isinstance(x, datetime):
            return x.date().isoformat()
        s = str(x).strip()
        if not s:
            return None
        try:
            return datetime.fromisoformat(s).date().isoformat()
        except Exception:
            pass
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except Exception:
                pass
        return None

    def _validate_whites(ws: list[int]) -> str | None:
        if len(ws) != 5:
            return "WHITES_COUNT_INVALID"
        if any(w is None for w in ws):
            return "WHITES_NULL"
        if any((w < 1 or w > 69) for w in ws):
            return "WHITES_RANGE"
        if len(set(ws)) != 5:
            return "WHITES_DUPLICATED"
        return None

    def _validate_powerball(pb: int | None) -> str | None:
        if pb is None:
            return "POWERBALL_NULL"
        if pb < 1 or pb > 26:
            return "POWERBALL_RANGE"
        return None

    # 1) read excel
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    if isinstance(df, dict):
        df = next(iter(df.values()))

    # 2) normalize cols
    df.columns = _norm_cols(df.columns)
    cols = set(df.columns)

    date_col = _pick_col(cols, ["drawdate", "draw_date", "date", "draw date"])

    whites_cols = None
    if all(c in cols for c in ["ball1", "ball2", "ball3", "ball4", "ball5"]):
        whites_cols = ["ball1", "ball2", "ball3", "ball4", "ball5"]
    elif all(c in cols for c in ["n1", "n2", "n3", "n4", "n5"]):
        whites_cols = ["n1", "n2", "n3", "n4", "n5"]
    elif all(c in cols for c in ["white1", "white2", "white3", "white4", "white5"]):
        whites_cols = ["white1", "white2", "white3", "white4", "white5"]

    pb_col = _pick_col(cols, ["powerball", "pb", "power_ball"])

    if whites_cols is None or pb_col is None:
        return {
            "status": "error",
            "error": "MISSING_COLUMNS",
            "message": (
                "No se detectó un formato válido. "
                "Se requiere (ball1..ball5 o n1..n5 o white1..white5) + powerball. "
                "Fecha opcional: drawdate/draw_date."
            ),
            "required_any_of": {
                "whites": ["ball1..ball5", "n1..n5", "white1..white5"],
                "powerball": ["powerball"],
                "date_optional": ["drawdate", "draw_date"],
            },
            "found": list(df.columns),
        }

    # 3) connect db + load existing combos
    db_path = _resolve_db_path()
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    def _combo_key(w1, w2, w3, w4, w5, pb) -> str:
        return f"{w1}-{w2}-{w3}-{w4}-{w5}|{pb}"

    existing_history: set[str] = set()
    existing_future: set[str] = set()

    try:
        for r in cur.execute("SELECT white1,white2,white3,white4,white5,powerball FROM draws"):
            existing_history.add(_combo_key(r["white1"], r["white2"], r["white3"], r["white4"], r["white5"], r["powerball"]))
    except Exception:
        pass

    try:
        for r in cur.execute("SELECT white1,white2,white3,white4,white5,powerball FROM future_draws"):
            existing_future.add(_combo_key(r["white1"], r["white2"], r["white3"], r["white4"], r["white5"], r["powerball"]))
    except Exception:
        pass

    # 4) iterate + insert
    inserted = 0
    skipped_exists_in_history = 0
    skipped_duplicates_future = 0
    invalid_rows = 0
    errors_preview: list[dict] = []

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    meta_obj = {
        "source": "excel",
        "sheet_name": sheet_name,
        "input_headers": list(df.columns),
        "detected_format": {
            "date_col": date_col,
            "whites_cols": whites_cols,
            "powerball_col": pb_col,
        },
    }
    meta_json = json.dumps(meta_obj, ensure_ascii=False)

    for idx, row in df.iterrows():
        draw_date = _parse_date(row[date_col]) if date_col else None

        ws = [_to_int(row[c]) for c in whites_cols]
        err = _validate_whites(ws)
        if err:
            invalid_rows += 1
            if len(errors_preview) < 25:
                errors_preview.append({
                    "row_index": int(idx),
                    "error": err,
                    "draw_date": draw_date,
                })
            continue

        ws_sorted = sorted(ws)
        w1, w2, w3, w4, w5 = ws_sorted

        pb = _to_int(row[pb_col])
        err_pb = _validate_powerball(pb)
        if err_pb:
            invalid_rows += 1
            if len(errors_preview) < 25:
                errors_preview.append({
                    "row_index": int(idx),
                    "error": err_pb,
                    "draw_date": draw_date,
                })
            continue

        key = _combo_key(w1, w2, w3, w4, w5, pb)

        if key in existing_history:
            skipped_exists_in_history += 1
            continue
        if key in existing_future:
            skipped_duplicates_future += 1
            continue

        try:
            cur.execute(
                """
                INSERT INTO future_draws (draw_date, white1, white2, white3, white4, white5, powerball, created_at, meta)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (draw_date, w1, w2, w3, w4, w5, pb, now_str, meta_json),
            )
            inserted += 1
            existing_future.add(key)
        except Exception as e:
            invalid_rows += 1
            if len(errors_preview) < 25:
                errors_preview.append({
                    "row_index": int(idx),
                    "error": "DB_INSERT_FAILED",
                    "detail": str(e),
                    "draw_date": draw_date,
                })
            continue

    con.commit()
    con.close()

    return {
        "status": "ok",
        "inserted": inserted,
        "skipped_exists_in_history": skipped_exists_in_history,
        "skipped_duplicates_future": skipped_duplicates_future,
        "invalid_rows": invalid_rows,
        "errors_preview": errors_preview,
        "meta": {
            "db": db_path,
            "detected_format": meta_obj["detected_format"],
        },
    }


# ------------------------
# FUTURE — Unique Quickpick
# ------------------------

def _load_existing_combos(cur) -> set:
    """
    Load existing combos from BOTH historical draws + future_draws.

    Returns:
      set of tuples: (w1,w2,w3,w4,w5,pb)
    """
    existing: set[tuple[int, int, int, int, int, int]] = set()

    def _add_rows(query: str):
        for r in cur.execute(query):
            a, b, c, d, e, pb = r
            if None in (a, b, c, d, e, pb):
                continue
            existing.add((int(a), int(b), int(c), int(d), int(e), int(pb)))

    # Canonical schema (white1..white5)
    try:
        _add_rows("SELECT white1,white2,white3,white4,white5,powerball FROM draws")
        _add_rows("SELECT white1,white2,white3,white4,white5,powerball FROM future_draws")
        return existing
    except Exception:
        existing.clear()

    # Legacy fallback (ball1..ball5)
    _add_rows("SELECT ball1,ball2,ball3,ball4,ball5,powerball FROM draws")
    _add_rows("SELECT ball1,ball2,ball3,ball4,ball5,powerball FROM future_draws")
    return existing


def create_future_quickpicks_unique(
    n: int = 1,
    draw_date: Optional[str] = None,
    seed: Optional[int] = None,
    white1: Optional[int] = None,
    white2: Optional[int] = None,
    white3: Optional[int] = None,
    white4: Optional[int] = None,
    white5: Optional[int] = None,
    powerball: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Genera N QuickPicks únicos vs:
    - draws (histórico)
    - future_draws (editable)

    Respeta constraints opcionales por posición (white1..white5) y/o powerball.
    """
    if n < 1:
        n = 1
    # IMPORTANT: contract says unique max 100
    if n > 100:
        n = 100

    rng = random.Random(seed)

    con = sqlite3.connect(str(DB_PATH))
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    existing = _load_existing_combos(cur)

    # Validate constraints
    whites_fixed = {1: white1, 2: white2, 3: white3, 4: white4, 5: white5}
    fixed_vals = [v for v in whites_fixed.values() if v is not None]

    if any((v < 1 or v > 69) for v in fixed_vals):
        con.close()
        return {"status": "error", "error": "INVALID_CONSTRAINTS", "message": "Whites deben estar en rango 1..69"}

    if len(set(fixed_vals)) != len(fixed_vals):
        con.close()
        return {"status": "error", "error": "INVALID_CONSTRAINTS", "message": "Hay whites repetidos en constraints"}

    if powerball is not None and (powerball < 1 or powerball > 26):
        con.close()
        return {"status": "error", "error": "INVALID_CONSTRAINTS", "message": "Powerball debe estar en rango 1..26"}

    # If positional constraints exist, they must be compatible with asc order once sorted.
    def violates_order() -> bool:
        vals = [whites_fixed[i] for i in range(1, 6)]
        for i in range(5):
            for j in range(i + 1, 5):
                if vals[i] is not None and vals[j] is not None and not (vals[i] < vals[j]):
                    return True
        return False

    if violates_order():
        con.close()
        return {
            "status": "error",
            "error": "INVALID_CONSTRAINTS",
            "message": "Condiciones posicionales incompatibles con orden asc (white1<white2<...<white5)",
        }

    inserted = 0
    attempts = 0
    max_attempts = max(2000, n * 2000)
    created: List[Dict[str, Any]] = []

    meta_obj = {
        "source": "quickpick_unique",
        "constraints": {
            "white1": white1, "white2": white2, "white3": white3, "white4": white4, "white5": white5,
            "powerball": powerball,
        },
    }
    meta_text = json.dumps(meta_obj, ensure_ascii=False)

    has_positional_constraints = any(v is not None for v in [white1, white2, white3, white4, white5])

    while inserted < n and attempts < max_attempts:
        attempts += 1

        if has_positional_constraints:
            # sample remaining + enforce positional matches after sort
            chosen = set([whites_fixed[i] for i in range(1, 6) if whites_fixed[i] is not None])
            remaining_needed = 5 - len(chosen)
            remaining_pool = [x for x in range(1, 70) if x not in chosen]
            extra = rng.sample(remaining_pool, remaining_needed)
            full = sorted(list(chosen) + extra)

            if white1 is not None and full[0] != white1:
                continue
            if white2 is not None and full[1] != white2:
                continue
            if white3 is not None and full[2] != white3:
                continue
            if white4 is not None and full[3] != white4:
                continue
            if white5 is not None and full[4] != white5:
                continue

            vals = full
        else:
            vals = sorted(rng.sample(range(1, 70), 5))

        pb = int(powerball) if powerball is not None else rng.randint(1, 26)

        combo = (vals[0], vals[1], vals[2], vals[3], vals[4], pb)
        if combo in existing:
            continue

        try:
            cur.execute(
                """
                INSERT INTO future_draws (draw_date, white1, white2, white3, white4, white5, powerball, meta)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (draw_date, combo[0], combo[1], combo[2], combo[3], combo[4], combo[5], meta_text),
            )
            inserted += 1
            existing.add(combo)
            created.append({
                "draw_date": draw_date,
                "white1": combo[0], "white2": combo[1], "white3": combo[2], "white4": combo[3], "white5": combo[4],
                "powerball": combo[5],
            })
        except sqlite3.IntegrityError:
            existing.add(combo)
            continue

    con.commit()
    con.close()

    return {
        "status": "ok",
        "requested": n,
        "inserted": inserted,
        "attempts": attempts,
        "exhausted": attempts >= max_attempts and inserted < n,
        "data": created,
    }


# ------------------------
# XLSX Export (Option B)
# ------------------------

_XLSX_HEADER_FONT = Font(bold=True)
_XLSX_GRAY_FILL = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")


def _xlsx_write_headers(ws, row: int, headers: List[str]) -> None:
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = _XLSX_HEADER_FONT
        cell.fill = _XLSX_GRAY_FILL


def _xlsx_set_width(ws, col: int, width: int) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def export_first_position_xlsx(
    limit: int = 69,
    source: str = "auto",
    output_path: Optional[str] = None,
) -> str:
    """
    Builds an XLSX export for First Position Frequency.
    Returns: output_path
    """
    if limit <= 0:
        limit = 69

    payload = export_by_first_position(limit=limit, source=source)

    if payload.get("status") != "ok":
        wb = Workbook()
        ws = wb.active
        ws.title = "ERROR"
        _xlsx_write_headers(ws, 1, ["status", "error", "message"])
        ws.cell(2, 1, value=str(payload.get("status")))
        ws.cell(2, 2, value=str(payload.get("error")))
        ws.cell(2, 3, value=str(payload.get("message")))
        _xlsx_set_width(ws, 1, 14)
        _xlsx_set_width(ws, 2, 18)
        _xlsx_set_width(ws, 3, 90)

        out = output_path or "powerball_first_position_export.xlsx"
        wb.save(out)
        return out

    data = payload.get("data", [])
    used_source = payload.get("source", "unknown")
    total_draws = payload.get("total_draws", 0)
    meta = payload.get("meta", {}) or {}

    wb = Workbook()

    ws = wb.active
    ws.title = "FIRST_POSITION_FREQUENCY"
    _xlsx_write_headers(ws, 1, ["Number", "Count", "Pct"])

    r = 2
    for row in data:
        ws.cell(r, 1, value=int(row.get("number", 0)))
        ws.cell(r, 2, value=int(row.get("count", 0)))
        ws.cell(r, 3, value=float(row.get("pct", 0.0)))
        r += 1

    _xlsx_set_width(ws, 1, 12)
    _xlsx_set_width(ws, 2, 12)
    _xlsx_set_width(ws, 3, 12)

    ws2 = wb.create_sheet("META")
    _xlsx_write_headers(ws2, 1, ["Key", "Value"])
    meta_rows = [
        ("source", used_source),
        ("total_draws", total_draws),
        ("limit", meta.get("limit", limit)),
        ("note", meta.get("note", "")),
    ]
    rr = 2
    for k, v in meta_rows:
        ws2.cell(rr, 1, value=str(k))
        ws2.cell(rr, 2, value=str(v))
        rr += 1

    _xlsx_set_width(ws2, 1, 22)
    _xlsx_set_width(ws2, 2, 90)

    out = output_path or "powerball_first_position_export.xlsx"
    wb.save(out)
    return out
