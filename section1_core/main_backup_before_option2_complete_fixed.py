from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Literal, Tuple
import os
import tempfile
import json
from urllib.parse import quote_plus

import pandas as pd
from fastapi import FastAPI, Depends, HTTPException, File, UploadFile, Query, Path, Body, Request, Response
from fastapi.responses import StreamingResponse, HTMLResponse, JSONResponse, RedirectResponse

from sqlalchemy.orm import Session
from sqlalchemy import asc, desc, extract, text, func

from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, ConfigDict

from starlette.middleware.gzip import GZipMiddleware
from starlette.middleware.cors import CORSMiddleware

from database import SessionLocal, engine, Base
from models import Ticket, DrawResult

from src.export_first_position import export_by_first_position

import random
from collections import Counter

import time
from typing import Tuple

from statistics import mean
from collections import defaultdict

from fastapi import FastAPI, HTTPException, Query
from pydantic import BaseModel
import pandas as pd
from collections import Counter
from typing import List, Dict, Any, Tuple

from datetime import date, datetime
from fastapi import HTTPException

def _parse_date_flexible(s: str, year: int | None = None) -> date:
    """
    Acepta:
      - YYYY-MM-DD  -> fecha exacta
      - MM-DD       -> usa year (si viene) o el año actual
    """
    s = (s or "").strip()

    # Caso 1: ISO completo YYYY-MM-DD
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(400, f"Fecha inválida (esperado YYYY-MM-DD): {s}")

    # Caso 2: MM-DD
    if len(s) == 5 and s[2] == "-":
        try:
            mm = int(s[:2])
            dd = int(s[3:])
        except ValueError:
            raise HTTPException(400, f"Fecha inválida (esperado MM-DD): {s}")

        y = int(year) if year is not None else date.today().year
        try:
            return date(y, mm, dd)
        except ValueError:
            raise HTTPException(400, f"Fecha inválida (MM-DD con year={y}): {s}")

    raise HTTPException(400, f"Formato de fecha inválido: '{s}'. Usa YYYY-MM-DD o MM-DD.")


app = FastAPI(
    title="Powerball + IA API",
    description="Backend para gestionar jugadas de Powerball.",
    version="1.2.0",
)

# -------------------------
#   DB SESSION DEPENDENCY
# -------------------------
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# -------------------------
#        HELPERS
# -------------------------


@app.get("/__test_export_route")
def __test_export_route():
    return {"ok": True}


POWERBALL_DRAWS_CSV = "powerball_draws.csv"
TICKETS_CSV = "tickets.csv"
WHITE_COLS = ["n1", "n2", "n3", "n4", "n5"]

# ---------------------------
#   EXPORT: /compare/by-date
# ---------------------------
@app.get("/compare/by-date/export.csv")
def compare_by_date_export_csv(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
    ticket: str = Query(..., description='Formato: "4,6,8,10,12|22"'),
    db: Session = Depends(get_db),
):
    results = _compare_day_df(month, day, ticket, db=db)
    safe_ticket = ticket.replace(",", "-").replace("|", "_").replace(" ", "")
    filename = f"compare_by_date_{month:02d}-{day:02d}_{safe_ticket}.csv"
    return _df_to_csv_stream(results, filename)

@app.get("/compare/by-date/export.xlsx")
def compare_by_date_export_xlsx(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
    ticket: str = Query(..., description='Formato: "4,6,8,10,12|22"'),
    db: Session = Depends(get_db),
):
    results = _compare_day_df(month, day, ticket, db=db)
    safe_ticket = ticket.replace(",", "-").replace("|", "_").replace(" ", "")
    filename = f"compare_by_date_{month:02d}-{day:02d}_{safe_ticket}.xlsx"
    return _df_to_xlsx_stream({"compare": results}, filename)


# ---------------------------
#   EXPORT: /compare/by-date/multi
# ---------------------------
@app.get("/compare/by-date/multi/export.csv")
def compare_by_date_multi_export_csv(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
    db: Session = Depends(get_db),
):
    out = _compare_multi_day_df(month, day, db=db)
    filename = f"compare_by_date_multi_{month:02d}-{day:02d}.csv"
    return _df_to_csv_stream(out, filename)

@app.get("/compare/by-date/multi/export.xlsx")
def compare_by_date_multi_export_xlsx(
    month: int = Query(..., ge=1, le=12),
    day: int = Query(..., ge=1, le=31),
    db: Session = Depends(get_db),
):
    out = _compare_multi_day_df(month, day, db=db)
    filename = f"compare_by_date_multi_{month:02d}-{day:02d}.xlsx"
    return _df_to_xlsx_stream({"multi": out}, filename)



def compute_ai_insight(
    db: Session,
    compare: str,
) -> Dict[str, Any]:
    """
    AI Insight basado en:
    - DrawResult (frecuencias reales del sorteo)
    - Ticket (comportamiento histórico)
    """

    parsed = parse_compare(compare)
    if not parsed:
        raise HTTPException(400, "compare inválido")

    base_regs, base_pb = parsed
    base_regs = sorted(list(base_regs))

    # -------------------------
    #   DRAW RESULTS ANALYSIS
    # -------------------------
    drs = db.query(DrawResult).all()
    if drs is None or len(drs) == 0:
        return {"score": 0, "notes": ["No draw results available"]}

    freq = defaultdict(int)
    sums = []
    pb_freq = defaultdict(int)

    for d in drs:
        nums = [d.wn1, d.wn2, d.wn3, d.wn4, d.wn5]
        for n in nums:
            freq[n] += 1
        sums.append(sum(nums))
        pb_freq[int(d.winning_powerball)] += 1

    avg_freq = mean(freq.values())
    avg_sum = mean(sums)
    base_sum = sum(base_regs)

    # Frecuencia promedio de la combinación
    base_freq_score = mean([freq[n] for n in base_regs])

    # -------------------------
    #   TICKETS ANALYSIS
    # -------------------------
    tickets = db.query(Ticket).all()
    ticket_freq = defaultdict(int)

    for t in tickets:
        for n in [t.n1, t.n2, t.n3, t.n4, t.n5]:
            ticket_freq[n] += 1

    avg_ticket_freq = mean(ticket_freq.values()) if ticket_freq else 0
    base_ticket_score = mean([ticket_freq.get(n, 0) for n in base_regs]) if ticket_freq else 0

    # -------------------------
    #   FEATURES
    # -------------------------
    odd = len([n for n in base_regs if n % 2 == 1])
    even = 5 - odd

    low = len([n for n in base_regs if n <= 35])
    high = 5 - low

    pb_rarity = pb_freq.get(base_pb, 0)
    pb_avg = mean(pb_freq.values())

    # -------------------------
    #   SCORING (0–100)
    # -------------------------
    score = 50

    if base_freq_score > avg_freq:
        score += 10
    else:
        score -= 5

    if base_ticket_score > avg_ticket_freq:
        score += 10

    if abs(base_sum - avg_sum) < 15:
        score += 10

    if odd in (2, 3):
        score += 5

    if low in (2, 3):
        score += 5

    if pb_rarity < pb_avg:
        score += 10
    else:
        score -= 5

    score = max(0, min(100, score))

    # -------------------------
    #   EXPLANATION
    # -------------------------
    notes = []

    notes.append(
        "Regular numbers: above-average frequency"
        if base_freq_score > avg_freq
        else "Regular numbers: below-average frequency"
    )

    notes.append(f"Parity (odd/even): {odd}/{even}")
    notes.append(f"Low / High split: {low}/{high}")
    notes.append(
        "Sum of numbers: near historical mean"
        if abs(base_sum - avg_sum) < 15
        else "Sum of numbers: far from historical mean"
    )

    notes.append(
        "Powerball: relatively rare"
        if pb_rarity < pb_avg
        else "Powerball: common"
    )

    return {
        "score": score,
        "notes": notes,
    }


# -------------------------
#   SIMPLE IN-MEMORY CACHE
# -------------------------
_COMPARE_CACHE: dict[str, Tuple[float, dict]] = {}
COMPARE_CACHE_TTL = 600  # segundos (10 minutos)


def _make_compare_cache_key(
    compare: str,
    status: Optional[str],
    type: Optional[str],
    start_date: Optional[str],
    end_date: Optional[str],
) -> str:
    return "|".join([
        compare or "",
        status or "",
        type or "",
        start_date or "",
        end_date or "",
    ])


def _get_compare_cache(key: str) -> Optional[dict]:
    item = _COMPARE_CACHE.get(key)
    if not item:
        return None

    ts, value = item
    if (time.time() - ts) > COMPARE_CACHE_TTL:
        # expirado
        _COMPARE_CACHE.pop(key, None)
        return None

    return value


def _set_compare_cache(key: str, value: dict) -> None:
    _COMPARE_CACHE[key] = (time.time(), value)


# -------------------------
#   CONFIG / CONSTANTS
# -------------------------
REG_MIN, REG_MAX = 1, 69
PB_MIN, PB_MAX = 1, 26

ALLOWED_TYPES = {"QUICK_PICK", "MANUAL"}
ALLOWED_STATUS = {"PAST", "FUTURE"}

FILL_AQUA = PatternFill(start_color="B7EDE6", end_color="B7EDE6", fill_type="solid")  # verde agua
FILL_PB = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")    # rojo suave
RED_FONT = Font(color="FF0000", bold=True)

# Crear tablas
Base.metadata.create_all(bind=engine)



# -------------------------
#   MIDDLEWARE
# -------------------------
app.add_middleware(GZipMiddleware, minimum_size=1000)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -------------------------
#        DB DEP
# -------------------------
def _norm_status(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    v = str(s).strip().upper()
    return v if v in ALLOWED_STATUS else None


def _norm_type(t: Optional[str]) -> Optional[str]:
    if t is None:
        return None
    v = str(t).strip().upper()
    if not v:
        return None
    if v in ("QP", "QUICKPICK", "QUICK-PICK"):
        v = "QUICK_PICK"
    return v if v in ALLOWED_TYPES else None


def normalize_regular_numbers(n1: int, n2: int, n3: int, n4: int, n5: int, order: str = "asc") -> List[int]:
    nums = [n1, n2, n3, n4, n5]
    nums_sorted = sorted(nums)
    if order.lower() == "desc":
        nums_sorted = list(reversed(nums_sorted))
    return nums_sorted


def validate_ticket_numbers(n1: int, n2: int, n3: int, n4: int, n5: int, pb: int) -> None:
    regs = [n1, n2, n3, n4, n5]
    if any((x < REG_MIN or x > REG_MAX) for x in regs):
        raise HTTPException(status_code=400, detail=f"N1..N5 deben estar entre {REG_MIN}-{REG_MAX}")
    if len(set(regs)) != 5:
        raise HTTPException(status_code=400, detail="N1..N5 no pueden repetirse dentro del mismo ticket")
    if pb < PB_MIN or pb > PB_MAX:
        raise HTTPException(status_code=400, detail=f"Powerball debe estar entre {PB_MIN}-{PB_MAX}")


def numbers_key(regs: List[int], pb: int) -> str:
    r = sorted(regs)
    return f"{r[0]}-{r[1]}-{r[2]}-{r[3]}-{r[4]}|{int(pb)}"


def _quote(v: Optional[str]) -> str:
    if not v:
        return ""
    return quote_plus(str(v))


def _safe_int(x: Any) -> Optional[int]:
    try:
        return int(str(x).strip())
    except Exception:
        return None


def parse_compare(compare: Optional[str]) -> Optional[Tuple[set, int]]:
    """
    Acepta formatos:
      - "10,16,29,33,69|22"
      - "10 16 29 33 69 | 22"
      - "10 16 29 33 69 22"   (último es PB)
      - "10,16,29,33,69 22"  (último es PB)
    Devuelve (set(regs), pb)
    """
    if not compare:
        return None

    raw = str(compare).strip()
    raw = raw.replace(";", ",")
    raw = raw.replace("\t", " ")
    raw = " ".join(raw.split())  # colapsa espacios

    if "|" in raw:
        left, right = raw.split("|", 1)
        left = left.strip().replace(" ", ",")
        right = right.strip()
        try:
            regs = [int(x.strip()) for x in left.split(",") if x.strip()]
            pb = int(right.strip())
        except Exception:
            raise HTTPException(status_code=400, detail='compare inválido. Ej: "10,16,29,33,69|22"')
        if len(regs) != 5:
            raise HTTPException(status_code=400, detail="compare debe tener 5 números regulares antes del |")
        validate_ticket_numbers(regs[0], regs[1], regs[2], regs[3], regs[4], pb)
        return set(regs), pb

    raw2 = raw.replace(",", " ")
    parts = [p for p in raw2.split(" ") if p.strip()]
    nums: List[int] = []
    for p in parts:
        v = _safe_int(p)
        if v is None:
            raise HTTPException(status_code=400, detail='compare inválido. Ej: "10,16,29,33,69|22"')
        nums.append(v)

    if len(nums) != 6:
        raise HTTPException(
            status_code=400,
            detail='compare inválido. Usa "10,16,29,33,69|22" o "10 16 29 33 69 22".',
        )

    regs = nums[:5]
    pb = nums[5]
    validate_ticket_numbers(regs[0], regs[1], regs[2], regs[3], regs[4], pb)
    return set(regs), pb


def _parse_date_iso(d: str) -> date:
    try:
        return date.fromisoformat(str(d).strip())
    except Exception:
        raise HTTPException(status_code=400, detail="Fecha inválida. Usa formato YYYY-MM-DD")


def _parse_date_maybe(s: Optional[str]) -> Optional[date]:
    if s is None:
        return None
    v = str(s).strip()
    if not v:
        return None
    return _parse_date_iso(v)


def _parse_int_list_csv(v: Optional[str]) -> List[int]:
    if not v:
        return []
    raw = str(v).replace(";", ",").replace("|", ",")
    items = [x.strip() for x in raw.split(",") if x.strip()]
    out: List[int] = []
    for it in items:
        n = _safe_int(it)
        if n is None:
            raise HTTPException(400, f"Lista inválida: {v}")
        out.append(int(n))
    return out


def _count_fast(q):
    """
    count() más estable y usualmente más rápido:
    SELECT count(*) FROM (subquery sin order_by)
    """
    try:
        return q.order_by(None).count()
    except Exception:
        return q.count()


# -------------------------
#   PERFORMANCE: INDICES
# -------------------------
def _ensure_indexes():
    """
    Crea índices comunes si el motor lo permite.
    - SQLite: CREATE INDEX IF NOT EXISTS ok
    """
    stmts = [
        "CREATE INDEX IF NOT EXISTS ix_ticket_status ON ticket(status)",
        "CREATE INDEX IF NOT EXISTS ix_ticket_draw_date ON ticket(draw_date)",
        "CREATE INDEX IF NOT EXISTS ix_ticket_type ON ticket(type)",
        "CREATE INDEX IF NOT EXISTS ix_drawresult_draw_date ON draw_result(draw_date)",
    ]
    try:
        with engine.begin() as conn:
            for s in stmts:
                try:
                    conn.execute(text(s))
                except Exception:
                    pass
    except Exception:
        pass


@app.on_event("startup")
def _on_startup():
    _ensure_indexes()


# -------------------------
#        PRIZES / MATCH
# -------------------------
def get_prize(matched_regular: int, matched_pb: bool) -> float:
    if matched_regular == 5 and matched_pb:
        return 0.0  # jackpot
    if matched_regular == 5 and not matched_pb:
        return 1_000_000.0
    if matched_regular == 4 and matched_pb:
        return 50_000.0
    if matched_regular == 4 and not matched_pb:
        return 100.0
    if matched_regular == 3 and matched_pb:
        return 100.0
    if matched_regular == 3 and not matched_pb:
        return 7.0
    if matched_regular == 2 and matched_pb:
        return 7.0
    if matched_regular == 1 and matched_pb:
        return 4.0
    if matched_regular == 0 and matched_pb:
        return 4.0
    return 0.0


def calculate_matches(ticket: Ticket, result: DrawResult):
    ticket_nums = {ticket.n1, ticket.n2, ticket.n3, ticket.n4, ticket.n5}
    winning_nums = {result.wn1, result.wn2, result.wn3, result.wn4, result.wn5}
    matched_regular = len(ticket_nums.intersection(winning_nums))
    matched_pb = (ticket.powerball == result.winning_powerball)
    prize = get_prize(matched_regular, matched_pb)
    ticket.matched_regular_numbers = matched_regular
    ticket.matched_powerball = matched_pb
    ticket.prize_amount = prize


# -------------------------
#        SCHEMAS
# -------------------------
class TicketCreate(BaseModel):
    draw_date: date
    status: Optional[Literal["PAST", "FUTURE"]] = None
    n1: int
    n2: int
    n3: int
    n4: int
    n5: int
    powerball: int
    type: str
    cost: float


class TicketUpdate(BaseModel):
    draw_date: Optional[date] = None
    status: Optional[Literal["PAST", "FUTURE"]] = None
    n1: Optional[int] = None
    n2: Optional[int] = None
    n3: Optional[int] = None
    n4: Optional[int] = None
    n5: Optional[int] = None
    powerball: Optional[int] = None
    type: Optional[str] = None
    cost: Optional[float] = None


class TicketOut(TicketCreate):
    id: int
    matched_regular_numbers: int = 0
    matched_powerball: bool = False
    prize_amount: float = 0.0
    model_config = ConfigDict(from_attributes=True)


class TicketNormalizedOut(BaseModel):
    id: int
    draw_date: date
    status: str
    type: str
    cost: float
    n1: int
    n2: int
    n3: int
    n4: int
    n5: int
    powerball: int
    matched_regular_numbers: int = 0
    matched_powerball: bool = False
    prize_amount: float = 0.0
    model_config = ConfigDict(from_attributes=True)


class DrawResultCreate(BaseModel):
    draw_date: date
    wn1: int
    wn2: int
    wn3: int
    wn4: int
    wn5: int
    winning_powerball: int


class DrawResultOut(DrawResultCreate):
    id: int
    model_config = ConfigDict(from_attributes=True)


class StatsSummary(BaseModel):
    total_tickets: int
    total_cost: float
    total_prize: float
    balance: float
    winning_tickets: int


# ============================================================
#   COMPARE + AI RECOMMENDER (schemas base)
# ============================================================
class CompareMatchesResponse(BaseModel):
    base: str
    total_scanned: int
    groups: Dict[str, int]
    tickets: Dict[str, List[TicketOut]]


class RecommendRequest(BaseModel):
    status: Optional[Literal["PAST", "FUTURE"]] = None
    type: Optional[str] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None

    k: int = 50
    seed: Optional[int] = None

    fixed_first: Optional[int] = None
    fixed_numbers: Optional[List[int]] = None
    exclude_numbers: Optional[List[int]] = None

    fixed_powerball: Optional[int] = None
    exclude_powerballs: Optional[List[int]] = None

    top_pool_regulars: int = 25
    top_pool_powerballs: int = 10


class RecommendResponse(BaseModel):
    generated: int
    seed: Optional[int]
    combos: List[Dict[str, Any]]


class SaveRecommendationsRequest(BaseModel):
    # Generación
    status: Optional[Literal["PAST", "FUTURE"]] = "PAST"
    type: Optional[str] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None

    k: int = 50
    seed: Optional[int] = None

    fixed_first: Optional[int] = None
    fixed_numbers: Optional[List[int]] = None
    exclude_numbers: Optional[List[int]] = None

    fixed_powerball: Optional[int] = None
    exclude_powerballs: Optional[List[int]] = None

    top_pool_regulars: int = 25
    top_pool_powerballs: int = 10

    # Guardado
    future_draw_date: date
    cost_per_ticket: float = 2.0
    save_type: Literal["QUICK_PICK", "MANUAL"] = "QUICK_PICK"
    normalize_on_save: bool = True


class SaveRecommendationsResponse(BaseModel):
    requested: int
    generated: int
    inserted: int
    skipped_duplicates: int
    future_draw_date: date
    status: str
    message: str


def _ticket_regs(t: Ticket) -> List[int]:
    return [int(t.n1), int(t.n2), int(t.n3), int(t.n4), int(t.n5)]


def _match_counts(base_regs: set, base_pb: int, t: Ticket) -> Tuple[int, bool]:
    regs = set(_ticket_regs(t))
    mr = len(regs.intersection(base_regs))
    mpb = (int(t.powerball) == int(base_pb))
    return mr, mpb


def _filter_tickets_for_ai(payload: Any, db: Session) -> List[Ticket]:
    q = db.query(Ticket)

    if getattr(payload, "status", None):
        q = q.filter(Ticket.status == payload.status)

    if getattr(payload, "type", None):
        ttype = _norm_type(payload.type)
        if ttype:
            q = q.filter(Ticket.type == ttype)

    if getattr(payload, "start_date", None):
        q = q.filter(Ticket.draw_date >= payload.start_date)

    if getattr(payload, "end_date", None):
        q = q.filter(Ticket.draw_date <= payload.end_date)

    return q.all()


def _existing_keys_for_draw(db: Session, draw_date_val: date) -> set:
    existing = db.query(Ticket).filter(Ticket.draw_date == draw_date_val).all()
    return {numbers_key([t.n1, t.n2, t.n3, t.n4, t.n5], int(t.powerball)) for t in existing}


# ============================================================
# ✅ FIX: FUNCIÓN QUE TE FALTABA (AI generator)
# ============================================================
def recommend_from_history(req: RecommendRequest, db: Session) -> RecommendResponse:
    """
    Genera combinaciones usando frecuencias históricas (tickets filtrados) + restricciones opcionales.
    Produce combos únicos dentro de la ejecución (no repite keys).
    """
    k = int(req.k or 50)
    k = max(1, min(5000, k))

    rng = random.Random(int(req.seed) if req.seed is not None else random.randrange(1, 10**9))

    tickets = _filter_tickets_for_ai(req, db)

    reg_counter = Counter()
    pb_counter = Counter()

    for t in tickets:
        reg_counter.update([int(t.n1), int(t.n2), int(t.n3), int(t.n4), int(t.n5)])
        pb_counter.update([int(t.powerball)])

    # fallback uniforme si no hay historial
    if not reg_counter:
        for n in range(REG_MIN, REG_MAX + 1):
            reg_counter[n] = 1
    if not pb_counter:
        for n in range(PB_MIN, PB_MAX + 1):
            pb_counter[n] = 1

    fixed_first = int(req.fixed_first) if req.fixed_first is not None else None
    fixed_numbers = [int(x) for x in (req.fixed_numbers or [])]
    exclude_numbers = set(int(x) for x in (req.exclude_numbers or []))

    fixed_pb = int(req.fixed_powerball) if req.fixed_powerball is not None else None
    exclude_pbs = set(int(x) for x in (req.exclude_powerballs or []))

    # valida ranges
    if fixed_first is not None and not (REG_MIN <= fixed_first <= REG_MAX):
        raise HTTPException(400, "fixed_first fuera de rango")
    for x in fixed_numbers:
        if not (REG_MIN <= x <= REG_MAX):
            raise HTTPException(400, f"fixed_numbers fuera de rango: {x}")
    if fixed_pb is not None and not (PB_MIN <= fixed_pb <= PB_MAX):
        raise HTTPException(400, "fixed_powerball fuera de rango")

    if fixed_first is not None and fixed_first in exclude_numbers:
        raise HTTPException(400, "fixed_first está en exclude_numbers")
    for x in fixed_numbers:
        if x in exclude_numbers:
            raise HTTPException(400, f"fixed_numbers contiene excluido: {x}")
    if fixed_pb is not None and fixed_pb in exclude_pbs:
        raise HTTPException(400, "fixed_powerball está en exclude_powerballs")

    # pool top
    top_pool_regulars = max(5, min(69, int(req.top_pool_regulars or 25)))
    top_regs = [n for (n, _) in reg_counter.most_common(top_pool_regulars)]
    for x in ([fixed_first] if fixed_first is not None else []) + fixed_numbers:
        if x is not None and x not in top_regs:
            top_regs.append(x)
    top_regs = [n for n in top_regs if n not in exclude_numbers]
    if len(top_regs) < 5:
        top_regs = [n for n in range(REG_MIN, REG_MAX + 1) if n not in exclude_numbers]

    if len(top_regs) < 5:
        raise HTTPException(400, "No hay suficientes números regulares disponibles (revisa excludes)")

    top_pool_powerballs = max(1, min(26, int(req.top_pool_powerballs or 10)))
    top_pbs = [n for (n, _) in pb_counter.most_common(top_pool_powerballs)]
    if fixed_pb is not None and fixed_pb not in top_pbs:
        top_pbs.append(fixed_pb)
    top_pbs = [n for n in top_pbs if n not in exclude_pbs]
    if not top_pbs:
        top_pbs = [n for n in range(PB_MIN, PB_MAX + 1) if n not in exclude_pbs]
    if not top_pbs:
        raise HTTPException(400, "No hay powerballs disponibles (revisa exclude_powerballs)")

    def _w(x: int) -> int:
        return max(1, int(reg_counter.get(x, 1)))

    def _wpb(x: int) -> int:
        return max(1, int(pb_counter.get(x, 1)))

    def pick_regs() -> List[int]:
        locked: List[int] = []
        if fixed_first is not None:
            locked.append(fixed_first)
        for x in fixed_numbers:
            if x not in locked:
                locked.append(x)
        if len(locked) > 5:
            raise HTTPException(400, "Demasiados números fijos (fixed_first + fixed_numbers > 5)")

        chosen = list(locked)
        pool = [n for n in top_regs if n not in chosen]
        while len(chosen) < 5:
            if not pool:
                break
            weights = [_w(n) for n in pool]
            n = rng.choices(pool, weights=weights, k=1)[0]
            chosen.append(n)
            pool.remove(n)

        # si no completa, rellena con universo permitido
        chosen = list(dict.fromkeys(chosen))
        if len(chosen) < 5:
            universe = [n for n in range(REG_MIN, REG_MAX + 1) if n not in exclude_numbers and n not in chosen]
            rng.shuffle(universe)
            chosen.extend(universe[: (5 - len(chosen))])

        if len(chosen) != 5 or len(set(chosen)) != 5:
            raise HTTPException(400, "No pude generar combinación válida con las restricciones dadas")
        return chosen

    def pick_pb() -> int:
        if fixed_pb is not None:
            return fixed_pb
        weights = [_wpb(n) for n in top_pbs]
        return rng.choices(top_pbs, weights=weights, k=1)[0]

    combos: List[Dict[str, Any]] = []
    seen = set()
    attempts = 0
    max_attempts = max(5000, k * 50)

    while len(combos) < k and attempts < max_attempts:
        attempts += 1
        regs = pick_regs()
        pb = pick_pb()
        if pb in exclude_pbs:
            continue

        key = numbers_key(regs, pb)
        if key in seen:
            continue
        seen.add(key)

        # output order: si fixed_first, mantenerlo primero
        if fixed_first is not None:
            rest = sorted([x for x in regs if x != fixed_first])
            out_regs = [fixed_first] + rest
        else:
            out_regs = sorted(regs)

        combos.append({
            "n1": int(out_regs[0]),
            "n2": int(out_regs[1]),
            "n3": int(out_regs[2]),
            "n4": int(out_regs[3]),
            "n5": int(out_regs[4]),
            "powerball": int(pb),
            "key": key,
        })

    return RecommendResponse(generated=len(combos), seed=req.seed, combos=combos)


# -------------------------
#   UI APP LAYOUT (PRO)
# -------------------------
APP_CSS = """
:root{
  --bg:#0b1220;
  --panel:#0f1a2e;
  --panel2:#101f38;
  --card:#0f213f;
  --muted:#9bb0d1;
  --text:#eaf1ff;
  --line:rgba(255,255,255,.08);
  --chip:rgba(255,255,255,.06);
  --accent:#7dd3fc;
  --accent2:#34d399;
  --warn:#fca5a5;
  --shadow: 0 10px 30px rgba(0,0,0,.35);
  --radius: 18px;
  --radius2: 14px;
  --mono: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace;
  --sans: ui-sans-serif, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, "Apple Color Emoji","Segoe UI Emoji";
}
*{ box-sizing:border-box; }
html,body{ height:100%; }
body{
  margin:0;
  font-family:var(--sans);
  background:
    radial-gradient(1000px 600px at 20% 10%, rgba(125,211,252,.12), transparent 55%),
    radial-gradient(900px 500px at 80% 20%, rgba(52,211,153,.10), transparent 55%),
    var(--bg);
  color:var(--text);
}
a{ color:inherit; text-decoration:none; }
small, .muted{ color:var(--muted); }
.app{
  display:grid;
  grid-template-columns: 280px 1fr;
  min-height:100vh;
}
.sidebar{
  position:sticky; top:0; height:100vh;
  padding:18px;
  border-right:1px solid var(--line);
  background: linear-gradient(180deg, rgba(255,255,255,.04), rgba(255,255,255,.02));
}
.brand{
  display:flex; align-items:center; gap:10px;
  padding:12px 12px; border:1px solid var(--line);
  border-radius: var(--radius);
  background: rgba(255,255,255,.03);
  box-shadow: var(--shadow);
}
.logo{
  width:38px; height:38px; border-radius:14px;
  background: linear-gradient(135deg, rgba(125,211,252,.9), rgba(52,211,153,.9));
  box-shadow: 0 8px 20px rgba(0,0,0,.35);
}
.brand h1{ font-size:14px; margin:0; line-height:1.1; }
.brand p{ margin:0; font-size:12px; color:var(--muted); }

.nav{ margin-top:14px; display:flex; flex-direction:column; gap:8px; }
.nav a{
  display:flex; align-items:center; gap:10px;
  padding:10px 12px;
  border:1px solid var(--line);
  border-radius: 14px;
  background: rgba(255,255,255,.02);
  transition: transform .08s ease, background .12s ease, border-color .12s ease;
}
.nav a:hover{ background: rgba(255,255,255,.05); transform: translateY(-1px); border-color: rgba(125,211,252,.25); }
.nav a.active{
  background: rgba(125,211,252,.10);
  border-color: rgba(125,211,252,.35);
}
.main{ padding:18px 18px 28px; }

.topbar{
  display:flex; align-items:center; justify-content:space-between;
  gap:10px;
  padding:12px 14px;
  border:1px solid var(--line);
  border-radius: var(--radius);
  background: rgba(255,255,255,.03);
  box-shadow: var(--shadow);
}
.topbar h2{ margin:0; font-size:16px; letter-spacing:.2px; }
.topbar .right{ display:flex; gap:10px; flex-wrap:wrap; justify-content:flex-end; }
.pill{
  display:inline-flex; align-items:center; gap:8px;
  padding:8px 10px;
  border:1px solid var(--line);
  border-radius: 999px;
  background: var(--chip);
  color: var(--muted);
  font-size:12px;
}
.grid{
  margin-top:14px;
  display:grid;
  grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
  gap:12px;
}
.card{
  border:1px solid var(--line);
  border-radius: var(--radius);
  background: rgba(255,255,255,.03);
  box-shadow: var(--shadow);
  padding:14px;
}
.card h3{ margin:0 0 6px; font-size:14px; }
.card p{ margin:0; color:var(--muted); font-size:13px; }

.btn{
  display:inline-flex;
  align-items:center;
  gap:8px;
  padding:9px 12px;
  border:1px solid var(--line);
  border-radius: 14px;
  background: rgba(255,255,255,.04);
  color:var(--text);
  cursor:pointer;
  transition: transform .08s ease, background .12s ease, border-color .12s ease;
  font-size:13px;
}
.btn:hover{ background: rgba(255,255,255,.07); transform: translateY(-1px); border-color: rgba(125,211,252,.25); }
.btn.primary{ background: rgba(125,211,252,.14); border-color: rgba(125,211,252,.35); }
.btn.good{ background: rgba(52,211,153,.12); border-color: rgba(52,211,153,.35); }
.btn.warn{ background: rgba(252,165,165,.12); border-color: rgba(252,165,165,.35); }

hr.sep{ border:0; height:1px; background: var(--line); margin:14px 0; }

.tableWrap{
  margin-top:12px;
  border:1px solid var(--line);
  border-radius: var(--radius);
  overflow:hidden;
  background: rgba(255,255,255,.02);
}
table{ border-collapse: collapse; width:100%; }
th,td{ border-bottom:1px solid var(--line); padding:10px 10px; text-align:center; font-size:13px; }
th{
  background: rgba(255,255,255,.04);
  color: var(--muted);
  font-weight:700;
  position: sticky;
  top:0;
  backdrop-filter: blur(8px);
}
tr:hover td{ background: rgba(255,255,255,.03); }
.code{ font-family:var(--mono); background: rgba(255,255,255,.06); padding:2px 6px; border-radius:10px; }

.formRow{
  display:flex; gap:10px; flex-wrap:wrap; align-items:center;
  padding:12px 12px;
  border:1px solid var(--line);
  border-radius: var(--radius);
  background: rgba(255,255,255,.02);
  margin-top:12px;
}
.field{
  padding:9px 10px;
  border:1px solid var(--line);
  border-radius: 14px;
  background: rgba(0,0,0,.18);
  color: var(--text);
  min-width: 240px;
  outline:none;
}
.fieldSmall{ min-width: 160px; }
.field:focus{ border-color: rgba(125,211,252,.45); }

.toast{
  position: fixed;
  right: 18px;
  bottom: 18px;
  background: rgba(0,0,0,.75);
  color:#fff;
  padding: 10px 12px;
  border-radius: 14px;
  display:none;
  font-size:13px;
  border:1px solid rgba(255,255,255,.12);
  box-shadow: var(--shadow);
}

.badge{
  display:inline-flex;
  align-items:center;
  gap:8px;
  font-size:12px;
  padding:6px 10px;
  border:1px solid var(--line);
  border-radius: 999px;
  background: rgba(255,255,255,.03);
  color: var(--muted);
}
"""

def render_app_page(*, title: str, active: str, body_html: str, right_pills_html: str = "") -> HTMLResponse:
    def nav_item(href: str, label: str, key: str) -> str:
        cls = "active" if key == active else ""
        return f'<a class="{cls}" href="{href}">{label}</a>'

    html = f"""
    <html>
    <head>
      <meta charset="utf-8"/>
      <meta name="viewport" content="width=device-width, initial-scale=1"/>
      <title>{title}</title>
      <style>{APP_CSS}</style>
    </head>
    <body>
      <div class="app">
        <aside class="sidebar">
          <div class="brand">
            <div class="logo"></div>
            <div>
              <h1>Powerball AI</h1>
              <p>Dashboard • SQLite • FastAPI</p>
            </div>
          </div>

          <nav class="nav">
            {nav_item("/ui", "🏠 Dashboard", "home")}
            {nav_item("/tickets/table", "📋 Tickets Table", "table")}
            {nav_item("/ui/compare", "🧩 Compare", "compare")}
            {nav_item("/ui/recommendations", "🤖 AI Recommendations", "ai")}
            {nav_item("/docs", "📘 Swagger", "docs")}
          </nav>

          <hr class="sep"/>
          <div class="muted" style="font-size:12px; line-height:1.35;">
            Tip: escribe compare como <span class="code">10,16,29,33,69|22</span>.
          </div>
        </aside>

        <main class="main">
          <div class="topbar">
            <h2>{title}</h2>
            <div class="right">{right_pills_html}</div>
          </div>

          {body_html}
        </main>
      </div>
    </body>
    </html>
    """
    return HTMLResponse(content=html)


# -------------------------
#        BASE ROUTES + UI
# -------------------------
@app.get("/")
async def read_root():
    return {"message": "API de Powerball + IA funcionando 🚀", "ui": "/ui", "docs": "/docs"}


@app.get("/apple-touch-icon.png")
@app.get("/apple-touch-icon-precomposed.png")
def apple_touch_icon():
    return Response(status_code=204)


@app.get("/ui/board")
def ui_board_redirect(request: Request):
    qs = request.url.query
    if qs:
        return RedirectResponse(url=f"/tickets/table?{qs}", status_code=307)
    return RedirectResponse(url="/tickets/table", status_code=307)


@app.get("/status")
async def get_status():
    return {"status": "ok"}


@app.get("/health")
async def health():
    return {"ok": True, "service": "powerball_ai"}


@app.get("/admin/self_test")
def admin_self_test(db: Session = Depends(get_db)):
    out: Dict[str, Any] = {"ok": True, "checks": {}}
    try:
        db.execute(text("SELECT 1"))
        out["checks"]["db"] = "ok"
    except Exception as e:
        out["ok"] = False
        out["checks"]["db"] = f"fail: {e}"

    try:
        total = db.query(func.count(Ticket.id)).scalar() or 0
        out["checks"]["tickets_total"] = int(total)
    except Exception as e:
        out["ok"] = False
        out["checks"]["tickets_total"] = f"fail: {e}"

    try:
        future = db.query(func.count(Ticket.id)).filter(Ticket.status == "FUTURE").scalar() or 0
        past = db.query(func.count(Ticket.id)).filter(Ticket.status == "PAST").scalar() or 0
        out["checks"]["tickets_future"] = int(future)
        out["checks"]["tickets_past"] = int(past)
    except Exception as e:
        out["ok"] = False
        out["checks"]["tickets_status_counts"] = f"fail: {e}"

    try:
        last = db.query(DrawResult).order_by(DrawResult.draw_date.desc()).first()
        out["checks"]["last_draw_exists"] = bool(last)
        if last:
            out["checks"]["last_draw_date"] = last.draw_date.isoformat()
    except Exception as e:
        out["ok"] = False
        out["checks"]["last_draw"] = f"fail: {e}"

    return out


@app.get("/ui", response_class=HTMLResponse)
def ui_home(db: Session = Depends(get_db)):
    try:
        total = db.query(func.count(Ticket.id)).scalar() or 0
    except Exception:
        total = 0

    try:
        future = db.query(func.count(Ticket.id)).filter(Ticket.status == "FUTURE").scalar() or 0
    except Exception:
        future = 0

    try:
        past = db.query(func.count(Ticket.id)).filter(Ticket.status == "PAST").scalar() or 0
    except Exception:
        past = 0

    last_draw = db.query(DrawResult).order_by(DrawResult.draw_date.desc()).first()
    last_draw_txt = "—"
    if last_draw:
        last_draw_txt = f"{last_draw.draw_date.isoformat()} | {last_draw.wn1},{last_draw.wn2},{last_draw.wn3},{last_draw.wn4},{last_draw.wn5} PB {last_draw.winning_powerball}"

    right = f"""
      <span class="pill">Total: <b style="color:var(--text)">{int(total)}</b></span>
      <span class="pill">PAST: <b style="color:var(--text)">{int(past)}</b></span>
      <span class="pill">FUTURE: <b style="color:var(--text)">{int(future)}</b></span>
    """

    body = f"""
      <div class="grid">
        <div class="card">
          <h3>Tickets Table</h3>
          <p>Tabla clara (1 combinación por fila), filtros, compare, matches y export.</p>
          <div style="margin-top:12px; display:flex; gap:10px; flex-wrap:wrap;">
            <a class="btn primary" href="/tickets/table">Abrir</a>
            <a class="btn" href="/tickets/table?status=FUTURE">FUTURE</a>
            <a class="btn" href="/tickets/table?status=PAST">PAST</a>
          </div>
        </div>

        <div class="card">
          <h3>Compare</h3>
          <p>Agrupa por 3/4/5/6 coincidencias y exporta Excel.</p>
          <div style="margin-top:12px;">
            <a class="btn primary" href="/ui/compare">Abrir</a>
          </div>
          <div class="muted" style="margin-top:10px;">Ej: <span class="code">10,16,29,33,69|22</span></div>
        </div>

        <div class="card">
          <h3>AI Recommendations</h3>
          <p>Genera combinaciones, exporta y guarda como FUTURE con dedupe.</p>
          <div style="margin-top:12px;">
            <a class="btn good" href="/ui/recommendations">Abrir</a>
          </div>
        </div>

        <div class="card">
          <h3>Estado</h3>
          <p>Last draw: <span class="code">{last_draw_txt}</span></p>
          <div style="margin-top:12px; display:flex; gap:10px; flex-wrap:wrap;">
            <a class="btn" href="/admin/self_test">Run self test</a>
            <a class="btn" href="/docs">Swagger</a>
          </div>
        </div>
      </div>
    """

    return render_app_page(title="Dashboard", active="home", body_html=body, right_pills_html=right)


# -------------------------
#    COMPARE VALIDATOR
# -------------------------
@app.get("/compare/parse")
def compare_parse(compare: str = Query(..., description='Ej: "10,16,29,33,69|22"')):
    parsed = parse_compare(compare)
    if not parsed:
        raise HTTPException(400, "compare inválido")
    regs, pb = parsed
    regs_sorted = sorted(list(regs))
    normalized = f"{regs_sorted[0]},{regs_sorted[1]},{regs_sorted[2]},{regs_sorted[3]},{regs_sorted[4]}|{int(pb)}"
    return {"ok": True, "normalized": normalized, "regulars": regs_sorted, "powerball": int(pb)}


# -------------------------
#        STATS
# -------------------------
@app.get("/stats/summary", response_model=StatsSummary)
def get_stats_summary(db: Session = Depends(get_db)):
    tickets = db.query(Ticket).all()
    total_tickets = len(tickets)
    total_cost = float(sum(t.cost for t in tickets))
    total_prize = float(sum((t.prize_amount or 0.0) for t in tickets))
    winning_tickets = sum(1 for t in tickets if (t.prize_amount or 0) > 0)
    balance = total_prize - total_cost
    return StatsSummary(
        total_tickets=total_tickets,
        total_cost=total_cost,
        total_prize=total_prize,
        balance=balance,
        winning_tickets=winning_tickets,
    )


# ============================================================
# ✅ /tickets/table PRO (UI principal) — Opción 2 + Extras
# ============================================================
@app.get("/tickets/table", response_class=HTMLResponse)
def tickets_table(
    status: Optional[str] = Query(default=None),
    type: Optional[str] = Query(default=None),
    order: Literal["asc", "desc"] = Query(default="asc"),
    compare: Optional[str] = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=200, ge=20, le=2000),
    key_contains: Optional[str] = Query(default=None),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),

    # ✅ EXTRAS
    only_matches: bool = Query(default=False),
    min_match: int = Query(default=0, ge=0, le=6),      # 0=all, 3,4,5,6
    sort_by: Literal["id", "draw_date", "matches"] = Query(default="id"),

    db: Session = Depends(get_db),
):
    sd = _parse_date_maybe(start_date)
    ed = _parse_date_maybe(end_date)

    query = db.query(Ticket)

    s = _norm_status(status)
    if s:
        query = query.filter(Ticket.status == s)

    ttype = _norm_type(type)
    if ttype:
        query = query.filter(Ticket.type == ttype)

    if sd:
        query = query.filter(Ticket.draw_date >= sd)
    if ed:
        query = query.filter(Ticket.draw_date <= ed)

    compare_parsed = parse_compare(compare) if compare else None
    last_draw = db.query(DrawResult).order_by(DrawResult.draw_date.desc()).first()

    if compare_parsed:
        winners_set, winning_pb = compare_parsed
        compare_label = f"Comparando con: {compare}"
    elif last_draw:
        winners_set = {last_draw.wn1, last_draw.wn2, last_draw.wn3, last_draw.wn4, last_draw.wn5}
        winning_pb = last_draw.winning_powerball
        compare_label = (
            f"Comparando con último draw ({last_draw.draw_date.isoformat()}): "
            f"{sorted(list(winners_set))} | PB {winning_pb}"
        )
    else:
        winners_set = set()
        winning_pb = None
        compare_label = "Sin comparación (no hay draw_results guardados y no enviaste compare)."

    total = _count_fast(query)

    page_size = max(20, min(2000, int(page_size)))
    offset = (page - 1) * page_size

    base_tickets = query.order_by(Ticket.id.asc()).offset(offset).limit(page_size).all()
    key_contains_norm = (key_contains or "").strip()

    def calc_match(t: Ticket) -> Tuple[int, bool, int]:
        """returns (matched_regular, matched_pb, total_balls=mr + pb_if)"""
        if not winners_set or winning_pb is None:
            return 0, False, 0
        regs = {t.n1, t.n2, t.n3, t.n4, t.n5}
        mr = len(regs.intersection(winners_set))
        mpb = int(t.powerball) == int(winning_pb)
        total_balls = mr + (1 if mpb else 0)
        return mr, mpb, total_balls

    items = []
    for t in base_tickets:
        regs_sorted = normalize_regular_numbers(t.n1, t.n2, t.n3, t.n4, t.n5, order=order)
        k = numbers_key([t.n1, t.n2, t.n3, t.n4, t.n5], int(t.powerball))

        if key_contains_norm and key_contains_norm not in k:
            continue

        mr, mpb, total_balls = calc_match(t)

        if only_matches and total_balls <= 0:
            continue
        if min_match and total_balls < int(min_match):
            continue

        items.append((t, regs_sorted, k, mr, mpb, total_balls))

    if sort_by == "draw_date":
        items.sort(key=lambda x: (x[0].draw_date, x[0].id))
    elif sort_by == "matches":
        items.sort(key=lambda x: (x[5], x[3], x[0].draw_date, x[0].id), reverse=True)
    else:
        items.sort(key=lambda x: x[0].id)

    def td(val: int, is_pb: bool = False) -> str:
        style = ""
        if is_pb and winning_pb is not None and int(val) == int(winning_pb):
            style = 'style="background:rgba(252,165,165,.22); color:#ffd1d1; font-weight:800;"'
        elif (not is_pb) and winners_set and int(val) in winners_set:
            style = 'style="background:rgba(52,211,153,.18); font-weight:800;"'
        return f"<td {style}>{val}</td>"

    def match_badge(mr: int, mpb: bool, total_balls: int) -> str:
        if not winners_set or winning_pb is None:
            return "<span class='badge'>—</span>"
        if total_balls <= 0:
            return "<span class='badge'>0</span>"
        pb_txt = " +PB" if mpb else ""
        return f"<span class='badge'><b style='color:var(--text)'>{mr}</b>{pb_txt} → <b style='color:var(--text)'>{total_balls}</b></span>"

    rows = ""
    shown = 0

    for (t, regs, k, mr, mpb, total_balls) in items:
        shown += 1
        combo_txt = f"{regs[0]} {regs[1]} {regs[2]} {regs[3]} {regs[4]} | PB {int(t.powerball)}"
        rows += "<tr>"
        rows += f"<td>{t.id}</td>"
        rows += f"<td>{t.draw_date}</td>"
        rows += f"<td><span class='badge'>{t.status}</span></td>"
        rows += f"<td><span class='badge'>{t.type}</span></td>"
        rows += td(regs[0]); rows += td(regs[1]); rows += td(regs[2]); rows += td(regs[3]); rows += td(regs[4])
        rows += td(int(t.powerball), is_pb=True)
        rows += f"<td>{match_badge(mr, mpb, total_balls)}</td>"
        rows += f"<td>${float(t.cost):.2f}</td>"
        rows += f"<td><span class='code'>{k}</span></td>"
        rows += f"<td><button class='btn' style='padding:7px 10px;' data-copy='{combo_txt}'>Copy</button></td>"
        rows += "</tr>\n"

    if shown == 0:
        rows = "<tr><td colspan='16' class='muted' style='text-align:left; padding:14px;'>No hay resultados con estos filtros.</td></tr>"

    pages = max(1, (total + page_size - 1) // page_size)

    def qparam(k: str, v: Optional[str]) -> str:
        if v is None:
            return ""
        vv = str(v).strip()
        if not vv:
            return ""
        return f"&{k}={_quote(vv)}"

    base_qs = (
        f"?page_size={page_size}"
        f"{qparam('status', s)}"
        f"{qparam('type', ttype)}"
        f"{qparam('order', order)}"
        f"{qparam('compare', compare)}"
        f"{qparam('key_contains', key_contains_norm)}"
        f"{qparam('start_date', start_date)}"
        f"{qparam('end_date', end_date)}"
        f"{qparam('sort_by', sort_by)}"
        f"&only_matches={'true' if only_matches else 'false'}"
        f"&min_match={int(min_match)}"
    )

    prev_link = f"/tickets/table{base_qs}&page={max(1, page-1)}"
    next_link = f"/tickets/table{base_qs}&page={min(pages, page+1)}"

    exp_xlsx = (
        f"/export_excel?order={order}"
        f"{qparam('status', s)}"
        f"{qparam('type', ttype)}"
        f"{qparam('compare', compare)}"
        f"{qparam('start_date', start_date)}"
        f"{qparam('end_date', end_date)}"
    )
    exp_csv = (
        f"/export_csv?order={order}"
        f"{qparam('status', s)}"
        f"{qparam('type', ttype)}"
        f"{qparam('start_date', start_date)}"
        f"{qparam('end_date', end_date)}"
    )

    compare_default = (compare or "").strip()

    body = f"""
      <div class="formRow">
        <input id="compareInput" class="field" placeholder="Compare: 10,16,29,33,69|22" value="{compare_default}"/>

        <select id="statusSel" class="field fieldSmall">
          <option value="" {"selected" if not s else ""}>Status: ALL</option>
          <option value="PAST" {"selected" if s=="PAST" else ""}>PAST</option>
          <option value="FUTURE" {"selected" if s=="FUTURE" else ""}>FUTURE</option>
        </select>

        <select id="typeSel" class="field fieldSmall">
          <option value="" {"selected" if not ttype else ""}>Type: ALL</option>
          <option value="QUICK_PICK" {"selected" if ttype=="QUICK_PICK" else ""}>QUICK_PICK</option>
          <option value="MANUAL" {"selected" if ttype=="MANUAL" else ""}>MANUAL</option>
        </select>

        <select id="orderSel" class="field fieldSmall">
          <option value="asc" {"selected" if order=="asc" else ""}>Order: ASC</option>
          <option value="desc" {"selected" if order=="desc" else ""}>Order: DESC</option>
        </select>

        <select id="sortBySel" class="field fieldSmall">
          <option value="id" {"selected" if sort_by=="id" else ""}>Sort: ID</option>
          <option value="draw_date" {"selected" if sort_by=="draw_date" else ""}>Sort: DATE</option>
          <option value="matches" {"selected" if sort_by=="matches" else ""}>Sort: MATCHES</option>
        </select>

        <input id="keyContains" class="field fieldSmall" placeholder="key contains (opcional)" value="{(key_contains_norm or "")}"/>
        <input id="startDate" class="field fieldSmall" placeholder="start YYYY-MM-DD" value="{(start_date or "")}"/>
        <input id="endDate" class="field fieldSmall" placeholder="end YYYY-MM-DD" value="{(end_date or "")}"/>

        <label class="badge" style="cursor:pointer;">
          <input id="onlyMatches" type="checkbox" style="transform:scale(1.1);" {"checked" if only_matches else ""}/>
          Only matches
        </label>

        <select id="minMatch" class="field fieldSmall">
          <option value="0" {"selected" if int(min_match)==0 else ""}>Min match: 0</option>
          <option value="3" {"selected" if int(min_match)==3 else ""}>Min match: 3</option>
          <option value="4" {"selected" if int(min_match)==4 else ""}>Min match: 4</option>
          <option value="5" {"selected" if int(min_match)==5 else ""}>Min match: 5</option>
          <option value="6" {"selected" if int(min_match)==6 else ""}>Min match: 6</option>
        </select>

        <button id="applyBtn" class="btn primary">Apply</button>
        <a class="btn" href="{exp_xlsx}">Export Excel</a>
        <a class="btn" href="{exp_csv}">Export CSV</a>
        <a class="btn" href="/ui/compare">Compare Tool</a>
      </div>

      <div class="muted" style="margin-top:10px;">{compare_label}</div>

      <div class="tableWrap">
        <table>
          <thead>
            <tr>
              <th>ID</th>
              <th>Draw Date</th>
              <th>Status</th>
              <th>Type</th>
              <th>N1</th><th>N2</th><th>N3</th><th>N4</th><th>N5</th>
              <th>PB</th>
              <th>Match</th>
              <th>Cost</th>
              <th>Key</th>
              <th>Copy</th>
            </tr>
          </thead>
          <tbody>
            {rows}
          </tbody>
        </table>
      </div>

      <div style="margin-top:12px; display:flex; gap:10px; flex-wrap:wrap; align-items:center;">
        <a class="btn" href="{prev_link}">⬅ Prev</a>
        <span class="pill">Page <b style="color:var(--text)">{page}</b> / {pages}</span>
        <span class="pill">Total <b style="color:var(--text)">{int(total)}</b></span>
        <a class="btn" href="{next_link}">Next ➡</a>
      </div>

      <div class="toast" id="toast">Copied ✅</div>

      <script>
        function toast(msg) {{
          const t = document.getElementById('toast');
          t.textContent = msg || 'OK';
          t.style.display = 'block';
          setTimeout(()=>t.style.display='none', 900);
        }}

        // Copy buttons
        document.querySelectorAll('button[data-copy]').forEach(btn => {{
          btn.addEventListener('click', async () => {{
            const txt = btn.getAttribute('data-copy');
            try {{
              await navigator.clipboard.writeText(txt);
              toast('Copied ✅');
            }} catch (e) {{
              toast('Copy failed');
            }}
          }});
        }});

        // Apply (sin editar URL a mano)
        document.getElementById('applyBtn').addEventListener('click', () => {{
          const compare = (document.getElementById('compareInput').value || '').trim();
          const status = (document.getElementById('statusSel').value || '').trim();
          const type = (document.getElementById('typeSel').value || '').trim();
          const order = (document.getElementById('orderSel').value || 'asc').trim();
          const sortBy = (document.getElementById('sortBySel').value || 'id').trim();
          const keyc = (document.getElementById('keyContains').value || '').trim();
          const sd = (document.getElementById('startDate').value || '').trim();
          const ed = (document.getElementById('endDate').value || '').trim();
          const onlyMatches = document.getElementById('onlyMatches').checked;
          const minMatch = (document.getElementById('minMatch').value || '0').trim();

          const params = new URLSearchParams();
          if (compare) params.set('compare', compare);
          if (status) params.set('status', status);
          if (type) params.set('type', type);
          if (order) params.set('order', order);
          if (sortBy) params.set('sort_by', sortBy);
          if (keyc) params.set('key_contains', keyc);
          if (sd) params.set('start_date', sd);
          if (ed) params.set('end_date', ed);
          if (onlyMatches) params.set('only_matches', 'true');
          if (minMatch && minMatch !== '0') params.set('min_match', minMatch);

          params.set('page', '1');
          params.set('page_size', '{page_size}');

          window.location.href = '/tickets/table?' + params.toString();
        }});
      </script>
    """

    right = f"""
      <span class="pill">Showing: <b style="color:var(--text)">{int(shown)}</b></span>
      <span class="pill">Page size: <b style="color:var(--text)">{int(page_size)}</b></span>
    """

    return render_app_page(title="Tickets Table", active="table", body_html=body, right_pills_html=right)


# ============================================================
#   TICKETS CRUD
# ============================================================
@app.post("/tickets", response_model=TicketOut)
def create_ticket(payload: TicketCreate, db: Session = Depends(get_db)):
    ttype = _norm_type(payload.type)
    if not ttype:
        raise HTTPException(400, "type inválido. Usa QUICK_PICK o MANUAL.")

    validate_ticket_numbers(payload.n1, payload.n2, payload.n3, payload.n4, payload.n5, payload.powerball)

    status = payload.status or ("PAST" if payload.draw_date <= date.today() else "FUTURE")

    key = numbers_key([payload.n1, payload.n2, payload.n3, payload.n4, payload.n5], int(payload.powerball))
    existing = db.query(Ticket).filter(Ticket.draw_date == payload.draw_date).all()
    existing_keys = {numbers_key([x.n1, x.n2, x.n3, x.n4, x.n5], int(x.powerball)) for x in existing}
    if key in existing_keys:
        raise HTTPException(409, "Ticket duplicado (misma combinación para ese draw_date).")

    ticket = Ticket(
        draw_date=payload.draw_date,
        status=status,
        n1=payload.n1, n2=payload.n2, n3=payload.n3, n4=payload.n4, n5=payload.n5,
        powerball=payload.powerball,
        type=ttype,
        cost=float(payload.cost or 0.0),
        matched_regular_numbers=0,
        matched_powerball=False,
        prize_amount=0.0,
    )

    res = db.query(DrawResult).filter(DrawResult.draw_date == payload.draw_date).first()
    if res:
        calculate_matches(ticket, res)

    db.add(ticket)
    db.commit()
    db.refresh(ticket)
    return ticket


@app.get("/tickets", response_model=List[TicketOut])
def list_tickets(
    status: Optional[str] = Query(default=None),
    type: Optional[str] = Query(default=None),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    limit: int = Query(default=2000, ge=1, le=20000),
    db: Session = Depends(get_db),
):
    s = _norm_status(status)
    ttype = _norm_type(type)
    sd = _parse_date_maybe(start_date)
    ed = _parse_date_maybe(end_date)

    q = db.query(Ticket)
    if s:
        q = q.filter(Ticket.status == s)
    if ttype:
        q = q.filter(Ticket.type == ttype)
    if sd:
        q = q.filter(Ticket.draw_date >= sd)
    if ed:
        q = q.filter(Ticket.draw_date <= ed)

    return q.order_by(Ticket.id.asc()).limit(limit).all()


@app.get("/tickets/{ticket_id}", response_model=TicketOut)
def get_ticket(ticket_id: int = Path(..., ge=1), db: Session = Depends(get_db)):
    t = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not t:
        raise HTTPException(404, "Ticket no encontrado")
    return t


@app.patch("/tickets/{ticket_id}", response_model=TicketOut)
def update_ticket(ticket_id: int, payload: TicketUpdate, db: Session = Depends(get_db)):
    t = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not t:
        raise HTTPException(404, "Ticket no encontrado")

    if payload.draw_date is not None:
        t.draw_date = payload.draw_date

    if payload.status is not None:
        if payload.status not in ALLOWED_STATUS:
            raise HTTPException(400, "status inválido")
        t.status = payload.status

    if payload.type is not None:
        ttype = _norm_type(payload.type)
        if not ttype:
            raise HTTPException(400, "type inválido")
        t.type = ttype

    n1 = payload.n1 if payload.n1 is not None else t.n1
    n2 = payload.n2 if payload.n2 is not None else t.n2
    n3 = payload.n3 if payload.n3 is not None else t.n3
    n4 = payload.n4 if payload.n4 is not None else t.n4
    n5 = payload.n5 if payload.n5 is not None else t.n5
    pb = payload.powerball if payload.powerball is not None else t.powerball

    validate_ticket_numbers(n1, n2, n3, n4, n5, pb)
    t.n1, t.n2, t.n3, t.n4, t.n5, t.powerball = n1, n2, n3, n4, n5, pb

    if payload.cost is not None:
        t.cost = float(payload.cost)

    key = numbers_key([t.n1, t.n2, t.n3, t.n4, t.n5], int(t.powerball))
    same_day = db.query(Ticket).filter(Ticket.draw_date == t.draw_date, Ticket.id != t.id).all()
    keys = {numbers_key([x.n1, x.n2, x.n3, x.n4, x.n5], int(x.powerball)) for x in same_day}
    if key in keys:
        raise HTTPException(409, "Update crea duplicado (misma combinación para ese draw_date).")

    res = db.query(DrawResult).filter(DrawResult.draw_date == t.draw_date).first()
    if res:
        calculate_matches(t, res)
    else:
        t.matched_regular_numbers = 0
        t.matched_powerball = False
        t.prize_amount = 0.0

    db.commit()
    db.refresh(t)
    return t


@app.delete("/tickets/{ticket_id}")
def delete_ticket(ticket_id: int, db: Session = Depends(get_db)):
    t = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not t:
        raise HTTPException(404, "Ticket no encontrado")
    db.delete(t)
    db.commit()
    return {"ok": True, "deleted": ticket_id}


# ============================================================
#   DRAW RESULTS CRUD
# ============================================================
@app.post("/draw_results", response_model=DrawResultOut)
def create_draw_result(payload: DrawResultCreate, db: Session = Depends(get_db)):
    regs = [payload.wn1, payload.wn2, payload.wn3, payload.wn4, payload.wn5]
    if any((x < REG_MIN or x > REG_MAX) for x in regs):
        raise HTTPException(400, "wn1..wn5 fuera de rango")
    if len(set(regs)) != 5:
        raise HTTPException(400, "wn1..wn5 no pueden repetirse")
    if payload.winning_powerball < PB_MIN or payload.winning_powerball > PB_MAX:
        raise HTTPException(400, "winning_powerball fuera de rango")

    existing = db.query(DrawResult).filter(DrawResult.draw_date == payload.draw_date).first()
    if existing:
        raise HTTPException(409, "Ya existe un draw_result para esa fecha")

    dr = DrawResult(
        draw_date=payload.draw_date,
        wn1=payload.wn1, wn2=payload.wn2, wn3=payload.wn3, wn4=payload.wn4, wn5=payload.wn5,
        winning_powerball=payload.winning_powerball
    )
    db.add(dr)
    db.commit()
    db.refresh(dr)

    tickets = db.query(Ticket).filter(Ticket.draw_date == payload.draw_date).all()
    for t in tickets:
        calculate_matches(t, dr)
        if t.draw_date <= date.today():
            t.status = "PAST"
    db.commit()

    return dr


@app.get("/draw_results", response_model=List[DrawResultOut])
def list_draw_results(limit: int = Query(default=2000, ge=1, le=20000), db: Session = Depends(get_db)):
    return db.query(DrawResult).order_by(DrawResult.draw_date.desc()).limit(limit).all()


@app.get("/draw_results/{draw_date}", response_model=DrawResultOut)
def get_draw_result(draw_date: str, year: int | None = None, db: Session = Depends(get_db)):
    d = _parse_date_flexible(draw_date, year=year)
    dr = db.query(DrawResult).filter(DrawResult.draw_date == d).first()
    if not dr:
        raise HTTPException(404, "No existe draw_result para esa fecha")
    return dr



@app.delete("/draw_results/{draw_date}")
def delete_draw_result(draw_date: str, year: int | None = None, db: Session = Depends(get_db)):
    d = _parse_date_flexible(draw_date, year=year)
    dr = db.query(DrawResult).filter(DrawResult.draw_date == d).first()
    if not dr:
        raise HTTPException(404, "No existe draw_result para esa fecha")

    db.delete(dr)

    tickets = db.query(Ticket).filter(Ticket.draw_date == d).all()
    for t in tickets:
        t.matched_regular_numbers = 0
        t.matched_powerball = False
        t.prize_amount = 0.0
    db.commit()
    return {"ok": True, "deleted_draw_date": d.isoformat(), "affected_tickets": len(tickets)}


# ============================================================
#   IMPORT EXCEL (Tickets)
# ============================================================
def _normalize_columns(cols: List[str]) -> Dict[str, str]:
    mapping = {}
    for c in cols:
        k = str(c).strip().lower()
        k2 = k.replace(" ", "_").replace("-", "_")
        mapping[c] = k2
    return mapping


def _coerce_status(v: Any, draw_date_val: date) -> str:
    s = _norm_status(str(v) if v is not None else "")
    if s:
        return s
    return "PAST" if draw_date_val <= date.today() else "FUTURE"


@app.post("/import_excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Sube un archivo Excel (.xlsx/.xls)")

    content = await file.read()
    try:
        df = pd.read_excel(BytesIO(content))
    except Exception as e:
        raise HTTPException(400, f"No pude leer Excel: {e}")

    if df.empty:
        return {"ok": True, "inserted": 0, "skipped": 0, "errors": 0}

    df = df.copy()
    colmap = _normalize_columns(list(df.columns))
    df.rename(columns=colmap, inplace=True)

    required_any = {"draw_date", "n1", "n2", "n3", "n4", "n5", "powerball"}
    if not required_any.issubset(set(df.columns)):
        raise HTTPException(
            400,
            f"Excel debe incluir columnas: {sorted(list(required_any))}. Columnas detectadas: {list(df.columns)}"
        )

    inserted = 0
    skipped = 0
    errors = 0

    keys_by_date: Dict[date, set] = {}

    for _, row in df.iterrows():
        try:
            dd = row.get("draw_date")
            if pd.isna(dd):
                raise ValueError("draw_date vacío")
            if isinstance(dd, datetime):
                draw_date_val = dd.date()
            elif isinstance(dd, date):
                draw_date_val = dd
            else:
                draw_date_val = _parse_date_iso(str(dd))

            n1 = int(row.get("n1")); n2 = int(row.get("n2")); n3 = int(row.get("n3"))
            n4 = int(row.get("n4")); n5 = int(row.get("n5"))
            pb = int(row.get("powerball"))

            validate_ticket_numbers(n1, n2, n3, n4, n5, pb)

            ttype = _norm_type(row.get("type"))
            if not ttype:
                ttype = "QUICK_PICK"

            cost_val = row.get("cost")
            cost = float(cost_val) if (cost_val is not None and not pd.isna(cost_val)) else 2.0

            status_val = _coerce_status(row.get("status"), draw_date_val)
            key = numbers_key([n1, n2, n3, n4, n5], pb)

            if draw_date_val not in keys_by_date:
                keys_by_date[draw_date_val] = _existing_keys_for_draw(db, draw_date_val)

            if key in keys_by_date[draw_date_val]:
                skipped += 1
                continue

            t = Ticket(
                draw_date=draw_date_val,
                status=status_val,
                n1=n1, n2=n2, n3=n3, n4=n4, n5=n5,
                powerball=pb,
                type=ttype,
                cost=cost,
                matched_regular_numbers=0,
                matched_powerball=False,
                prize_amount=0.0,
            )

            dr = db.query(DrawResult).filter(DrawResult.draw_date == draw_date_val).first()
            if dr:
                calculate_matches(t, dr)

            db.add(t)
            keys_by_date[draw_date_val].add(key)
            inserted += 1
        except Exception:
            errors += 1

    db.commit()
    return {"ok": True, "inserted": inserted, "skipped": skipped, "errors": errors}


# ============================================================
#   EXPORT CSV / EXCEL (con resaltado matches)
# ============================================================
def _filtered_ticket_query(
    db: Session,
    status: Optional[str],
    type: Optional[str],
    start_date: Optional[str],
    end_date: Optional[str],
):
    s = _norm_status(status)
    ttype = _norm_type(type)
    sd = _parse_date_maybe(start_date)
    ed = _parse_date_maybe(end_date)

    q = db.query(Ticket)
    if s:
        q = q.filter(Ticket.status == s)
    if ttype:
        q = q.filter(Ticket.type == ttype)
    if sd:
        q = q.filter(Ticket.draw_date >= sd)
    if ed:
        q = q.filter(Ticket.draw_date <= ed)
    return q


@app.get("/export_csv")
def export_csv(
    status: Optional[str] = Query(default=None),
    type: Optional[str] = Query(default=None),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    order: Literal["asc", "desc"] = Query(default="asc"),
    db: Session = Depends(get_db),
):
    q = _filtered_ticket_query(db, status, type, start_date, end_date)
    tickets = q.order_by(Ticket.id.asc()).all()

    rows = []
    for t in tickets:
        regs = normalize_regular_numbers(t.n1, t.n2, t.n3, t.n4, t.n5, order=order)
        rows.append({
            "id": t.id,
            "draw_date": t.draw_date.isoformat(),
            "status": t.status,
            "type": t.type,
            "n1": regs[0], "n2": regs[1], "n3": regs[2], "n4": regs[3], "n5": regs[4],
            "powerball": int(t.powerball),
            "matched_regular_numbers": int(getattr(t, "matched_regular_numbers", 0) or 0),
            "matched_powerball": bool(getattr(t, "matched_powerball", False) or False),
            "prize_amount": float(getattr(t, "prize_amount", 0.0) or 0.0),
            "cost": float(t.cost or 0.0),
            "key": numbers_key([t.n1, t.n2, t.n3, t.n4, t.n5], int(t.powerball)),
        })

    df = pd.DataFrame(rows)
    buf = BytesIO()
    df.to_csv(buf, index=False)
    buf.seek(0)

    filename = "tickets_export.csv"
    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )



@app.get("/export_compare_group")
def export_compare_group(
    compare: str = Query(...),
    group: Literal["3", "4", "5", "6"] = Query(...),
    status: Optional[str] = Query(default=None),
    type: Optional[str] = Query(default=None),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    order: Literal["asc", "desc"] = Query(default="asc"),
    db: Session = Depends(get_db),
):
    parsed = parse_compare(compare)
    if not parsed:
        raise HTTPException(400, "compare inválido")

    winners_set, winning_pb = parsed  # base regs + pb
    sd = _parse_date_maybe(start_date)
    ed = _parse_date_maybe(end_date)

    q = db.query(Ticket)

    s = _norm_status(status)
    if s:
        q = q.filter(Ticket.status == s)

    ttype = _norm_type(type)
    if ttype:
        q = q.filter(Ticket.type == ttype)

    if sd:
        q = q.filter(Ticket.draw_date >= sd)
    if ed:
        q = q.filter(Ticket.draw_date <= ed)

    rows = []
    tickets = q.order_by(Ticket.id.asc()).all()

    for t in tickets:
        regs = normalize_regular_numbers(t.n1, t.n2, t.n3, t.n4, t.n5, order=order)
        regs_set = set(regs)

        mr = len(regs_set.intersection(winners_set)) if winners_set else 0
        mpb = (winning_pb is not None and int(t.powerball) == int(winning_pb)) if winners_set else False

        ok = (
            (group == "6" and mr == 5 and mpb) or
            (group == "5" and mr == 5 and not mpb) or
            (group == "4" and mr == 4) or
            (group == "3" and mr == 3)
        )
        if not ok:
            continue

        total_balls = mr + (1 if mpb else 0)

        rows.append({
            "ID": t.id,
            "Draw Date": t.draw_date.isoformat() if getattr(t, "draw_date", None) else "",
            "Status": t.status,
            "Type": t.type,
            "N1": regs[0], "N2": regs[1], "N3": regs[2], "N4": regs[3], "N5": regs[4],
            "PB": int(t.powerball),
            "Match Regular": mr,
            "Match PB": "YES" if mpb else "NO",
            "Total Balls": total_balls,
            "Prize": float(getattr(t, "prize_amount", 0.0) or 0.0),
            "Cost": float(getattr(t, "cost", 0.0) or 0.0),
            "Key": numbers_key([t.n1, t.n2, t.n3, t.n4, t.n5], int(t.powerball)),
        })

    if not rows:
        raise HTTPException(404, "No hay resultados para este grupo")

    df = pd.DataFrame(rows)

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sheet = f"group_{group}"
        df.to_excel(writer, index=False, sheet_name=sheet)
        ws = writer.sheets[sheet]

        # Header bold
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        # Colorear coincidencias con el "compare" como base
        if winners_set and winning_pb is not None:
            # Column indexes (1-based): ajustados a nuestro df
            # ID(1), Draw Date(2), Status(3), Type(4), N1(5) ... N5(9), PB(10)
            col_n = {"N1": 5, "N2": 6, "N3": 7, "N4": 8, "N5": 9, "PB": 10}
            max_row = ws.max_row

            for r in range(2, max_row + 1):
                for c in [col_n["N1"], col_n["N2"], col_n["N3"], col_n["N4"], col_n["N5"]]:
                    v = ws.cell(row=r, column=c).value
                    try:
                        if int(v) in winners_set:
                            ws.cell(row=r, column=c).fill = FILL_AQUA
                    except Exception:
                        pass

                pbv = ws.cell(row=r, column=col_n["PB"]).value
                try:
                    if int(pbv) == int(winning_pb):
                        ws.cell(row=r, column=col_n["PB"]).fill = FILL_PB
                        ws.cell(row=r, column=col_n["PB"]).font = RED_FONT
                except Exception:
                    pass

        # Autosize columnas
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = "" if cell.value is None else str(cell.value)
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(40, max(10, max_len + 2))

    buf.seek(0)
    filename = f"compare_group_{group}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ============================================================
#   EXPORT FIRST POSITION (ya lo tenías) — se mantiene
# ============================================================
@app.get("/export_first_position")
def export_first_position_endpoint(db: Session = Depends(get_db)):
    """Exporta Excel agrupado por 'first position'.
    Intentamos llamar a export_by_first_position de forma compatible con varias firmas.
    """
    try:
        import inspect
        import traceback as _tb

        sig = inspect.signature(export_by_first_position)
        kwargs = {}

        # Compatibilidad con firmas comunes
        for name in ("db", "session", "db_session"):
            if name in sig.parameters:
                kwargs[name] = db

        for name in ("base_dir", "output_dir", "out_dir", "folder"):
            if name in sig.parameters:
                kwargs[name] = BASE_DIR

        for name in ("tickets_csv", "tickets_path", "tickets_file"):
            if name in sig.parameters:
                kwargs[name] = _abs_path(TICKETS_CSV)

        for name in ("draws_csv", "draws_path", "draws_file"):
            if name in sig.parameters:
                kwargs[name] = _abs_path(POWERBALL_DRAWS_CSV)

        result = export_by_first_position(**kwargs)
        return {"ok": True, "result": result}

    except HTTPException:
        raise
    except Exception as e:
        detail = f"{e}\\n" + _tb.format_exc()
        raise HTTPException(status_code=500, detail=detail)


@app.get("/compare/matches", response_model=CompareMatchesResponse)
def compare_matches(
    compare: str = Query(..., description='Ej: "10,16,29,33,69|22"'),
    status: Optional[str] = Query(default=None),
    type: Optional[str] = Query(default=None),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    # 🔥 CACHE KEY (compare + filtros)
    cache_key = _make_compare_cache_key(compare, status, type, start_date, end_date)

    cached = _get_compare_cache(cache_key)
    if cached:
        return CompareMatchesResponse(**cached)

    parsed = parse_compare(compare)
    if not parsed:
        raise HTTPException(400, "compare inválido")

    base_regs, base_pb = parsed
    sd = _parse_date_maybe(start_date)
    ed = _parse_date_maybe(end_date)

    q = db.query(Ticket)

    s = _norm_status(status)
    if s:
        q = q.filter(Ticket.status == s)

    ttype = _norm_type(type)
    if ttype:
        q = q.filter(Ticket.type == ttype)

    if sd:
        q = q.filter(Ticket.draw_date >= sd)
    if ed:
        q = q.filter(Ticket.draw_date <= ed)

    tickets = q.order_by(Ticket.id.asc()).all()
    buckets: Dict[str, List[TicketOut]] = {"3": [], "4": [], "5": [], "6": []}

    for t in tickets:
        mr, mpb = _match_counts(base_regs, base_pb, t)
        if mr == 5 and mpb:
            buckets["6"].append(t)
        elif mr == 5:
            buckets["5"].append(t)
        elif mr == 4:
            buckets["4"].append(t)
        elif mr == 3:
            buckets["3"].append(t)

    groups = {k: len(v) for k, v in buckets.items()}
    base_sorted = sorted(list(base_regs))
    base_str = f"{base_sorted[0]},{base_sorted[1]},{base_sorted[2]},{base_sorted[3]},{base_sorted[4]}|{int(base_pb)}"

    payload = {
        "base": base_str,
        "total_scanned": len(tickets),
        "groups": groups,
        "tickets": buckets,
    }

    # ✅ Guardar en cache
    _set_compare_cache(cache_key, payload)

    return CompareMatchesResponse(**payload)


@app.get("/ui/compare", response_class=HTMLResponse)
def ui_compare(
    compare: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    type: Optional[str] = Query(default=None),
    start_date: Optional[str] = Query(default=None),
    end_date: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    base_value = (compare or "").strip()
    base_label = ""
    if not base_value:
        last = db.query(DrawResult).order_by(DrawResult.draw_date.desc()).first()
        if last:
            regs = sorted([last.wn1, last.wn2, last.wn3, last.wn4, last.wn5])
            base_value = f"{regs[0]},{regs[1]},{regs[2]},{regs[3]},{regs[4]}|{int(last.winning_powerball)}"
            base_label = f"Usando último draw: {last.draw_date.isoformat()} • {regs} | PB {int(last.winning_powerball)}"
        else:
            base_label = "No hay draw_results. Escribe un compare y presiona Apply."
    else:
        base_label = f"Comparando con: {base_value}"

    s = _norm_status(status) or ""
    ttype = _norm_type(type) or ""

    body = f"""
      <div class="card">
        <h3>Compare Tool</h3>
        <p class="muted">{base_label}</p>

        <div class="formRow" style="margin-top:12px;">
          <input id="cmp" class="field"
                 placeholder="10,16,29,33,69|22  (o: 4 6 8 10 12 pb 22)"
                 value="{base_value}"/>

          <select id="st" class="field fieldSmall">
            <option value="" {"selected" if not s else ""}>Status: ALL</option>
            <option value="PAST" {"selected" if s=="PAST" else ""}>PAST</option>
            <option value="FUTURE" {"selected" if s=="FUTURE" else ""}>FUTURE</option>
          </select>

          <select id="tp" class="field fieldSmall">
            <option value="" {"selected" if not ttype else ""}>Type: ALL</option>
            <option value="QUICK_PICK" {"selected" if ttype=="QUICK_PICK" else ""}>QUICK_PICK</option>
            <option value="MANUAL" {"selected" if ttype=="MANUAL" else ""}>MANUAL</option>
          </select>

          <input id="sd" type="date" class="field fieldSmall" value="{(start_date or '')}"/>
          <input id="ed" type="date" class="field fieldSmall" value="{(end_date or '')}"/>

          <button id="run" class="btn primary">Apply</button>
          <a id="xlsx" class="btn" href="#" style="pointer-events:none; opacity:.6;">Export Excel</a>
          <button id="copyAll" class="btn">Copy all</button>
        </div>

        <div id="out" style="margin-top:12px;"></div>
      </div>

      <div class="toast" id="toast">OK</div>

      <script>
        function toast(msg) {{
          const t = document.getElementById('toast');
          t.textContent = msg || 'OK';
          t.style.display = 'block';
          setTimeout(()=>t.style.display='none', 1100);
        }}

        function esc(s){{
          return (s||'')
            .replaceAll('&','&amp;')
            .replaceAll('<','&lt;')
            .replaceAll('>','&gt;');
        }}

        function normalizeCompare(raw) {{
          let v = (raw || '').trim();
          if (!v) return '';
          if (v.includes('|')) return v.replace(/\\s+/g,'');
          v = v.toLowerCase()
               .replace(/powerball/g,' ')
               .replace(/\\bpb\\b/g,' ')
               .replace(/[^0-9,\\s]+/g,' ')
               .replace(/\\s+/g,' ')
               .trim();
          const parts = v.split(/[\\s,]+/).filter(Boolean);
          if (parts.length === 6) {{
            return `${{parts.slice(0,5).join(',')}}|${{parts[5]}}`;
          }}
          return parts.join(',');
        }}

        function groupBars(groups) {{
          const g = groups || {{}};
          const maxVal = Math.max(g['6']||0, g['5']||0, g['4']||0, g['3']||0, 1);
          const bar = (label, val) => {{
            const w = Math.round((val / maxVal) * 100);
            return `
              <div style="margin:6px 0;">
                <div style="display:flex; justify-content:space-between; font-size:13px;">
                  <span>${{label}}</span><b>${{val}}</b>
                </div>
                <div style="background:#1f2a44; height:8px; border-radius:6px;">
                  <div style="width:${{w}}%; height:8px; background:var(--accent2); border-radius:6px;"></div>
                </div>
              </div>
            `;
          }};
          return `
            <div class="card">
              <h3>Groups</h3>
              ${{bar('6 matches (5 + PB)', g['6']||0)}}
              ${{bar('5 matches', g['5']||0)}}
              ${{bar('4 matches', g['4']||0)}}
              ${{bar('3 matches', g['3']||0)}}
            </div>
          `;
        }}

        function exportLinks(params) {{
          const exp = (g) => {{
            const p = new URLSearchParams(params.toString());
            p.set('group', g);
            p.set('order', 'asc');

            return '/export_compare_group?' + p.toString();
          }};
          return `
            <div class="card">
              <h3>Export by Group</h3>
              <div style="display:flex; gap:8px; flex-wrap:wrap;">
                <a class="btn" href="${{exp('6')}}">Export 6</a>
                <a class="btn" href="${{exp('5')}}">Export 5</a>
                <a class="btn" href="${{exp('4')}}">Export 4</a>
                <a class="btn" href="${{exp('3')}}">Export 3</a>
              </div>
              <p class="muted" style="margin-top:8px;">Exporta solo el grupo seleccionado con los mismos filtros.</p>
            </div>
          `;
        }}

        async function run(){{
          const raw = document.getElementById('cmp').value || '';
          const compare = normalizeCompare(raw);
          let sd = document.getElementById('sd').value || '';
          let ed = document.getElementById('ed').value || '';
          const status = document.getElementById('st').value || '';
          const type = document.getElementById('tp').value || '';

          if (sd && ed && ed < sd) {{
            const tmp = sd; sd = ed; ed = tmp;
            document.getElementById('sd').value = sd;
            document.getElementById('ed').value = ed;
          }}

          if (!compare) {{
            alert('Escribe compare');
            return;
          }}

          const params = new URLSearchParams();
          params.set('compare', compare);
          if (status) params.set('status', status);
          if (type) params.set('type', type);
          if (sd) params.set('start_date', sd);
          if (ed) params.set('end_date', ed);

          // 🔗 Guardar estado en la URL (shareable)
          const url = new URL(window.location.href);
          // Limpia primero claves conocidas para evitar residuos
          ['compare','status','type','start_date','end_date'].forEach(k => url.searchParams.delete(k));
          for (const [k,v] of params.entries()) url.searchParams.set(k,v);
          window.history.replaceState({{}}, '', url.toString());

          const res = await fetch('/compare/matches?' + params.toString());
          const data = await res.json();
          if (!res.ok) {{
            alert(data.detail || 'Error');
            return;
          }}

          const out = document.getElementById('out');
          out.innerHTML = `
            <div class="grid">
              <div class="card"><h3>Base</h3><p><span class="code">${{esc(data.base)}}</span></p></div>
              <div class="card"><h3>Scanned</h3><p><span class="code">${{data.total_scanned}}</span></p></div>
              ${{groupBars(data.groups)}}
              ${{exportLinks(params)}}
            </div>
          `;

          // Mantengo tu export existente (si lo tienes en /export_excel)
          const xlsx = document.getElementById('xlsx');
          const ex = new URLSearchParams();
          ex.set('order','asc');
          ex.set('compare', compare);
          if (status) ex.set('status', status);
          if (type) ex.set('type', type);
          if (sd) ex.set('start_date', sd);
          if (ed) ex.set('end_date', ed);
          xlsx.href = '/export_excel?' + ex.toString();
          xlsx.style.pointerEvents = 'auto';
          xlsx.style.opacity = '1';

          window.__lastCompareLines = [];
          const g = data.tickets || {{}};
          ['6','5','4','3'].forEach(k => {{
            (g[k]||[]).forEach(t => {{
              const nums = [t.n1,t.n2,t.n3,t.n4,t.n5].sort((a,b)=>a-b);
              window.__lastCompareLines.push(`${{nums.join(' ')}} | PB ${{t.powerball}}`);
            }});
          }});
        }}

        document.getElementById('run').addEventListener('click', run);
        document.getElementById('cmp').addEventListener('keydown', e => {{
          if (e.key === 'Enter') {{ e.preventDefault(); run(); }}
        }});

        document.getElementById('copyAll').addEventListener('click', async () => {{
          const lines = window.__lastCompareLines || [];
          if (!lines.length) {{
            toast('Nada para copiar');
            return;
          }}
          await navigator.clipboard.writeText(lines.join('\\n'));
          toast('Copied all ✅');
        }});

        if ((document.getElementById('cmp').value || '').trim()) {{
          run();
        }}
      </script>
    """

    return render_app_page(title="Compare", active="compare", body_html=body)

 
 
 
#=========================
#   COMPARE - AI INSIGHT   
# =========================
@app.get("/compare/insight")
def compare_insight(
    compare: str = Query(...),
    db: Session = Depends(get_db),
):
    insight = compute_ai_insight(db, compare)
    return insight



# ============================================================
#   AI RECOMMENDATIONS (API + UI) — ÚNICA VERSIÓN
# ============================================================
@app.post("/ai/recommend", response_model=RecommendResponse)
def ai_recommend(req: RecommendRequest, db: Session = Depends(get_db)):
    return recommend_from_history(req, db)


@app.post("/ai/save_recommendations", response_model=SaveRecommendationsResponse)
def ai_save_recommendations(req: SaveRecommendationsRequest, db: Session = Depends(get_db)):
    gen_req = RecommendRequest(
        status=req.status,
        type=req.type,
        start_date=req.start_date,
        end_date=req.end_date,
        k=req.k,
        seed=req.seed,
        fixed_first=req.fixed_first,
        fixed_numbers=req.fixed_numbers,
        exclude_numbers=req.exclude_numbers,
        fixed_powerball=req.fixed_powerball,
        exclude_powerballs=req.exclude_powerballs,
        top_pool_regulars=req.top_pool_regulars,
        top_pool_powerballs=req.top_pool_powerballs,
    )
    rec = recommend_from_history(gen_req, db)

    existing_keys = _existing_keys_for_draw(db, req.future_draw_date)

    inserted = 0
    skipped = 0

    for c in rec.combos:
        regs = [int(c["n1"]), int(c["n2"]), int(c["n3"]), int(c["n4"]), int(c["n5"])]
        pb = int(c["powerball"])
        key = numbers_key(regs, pb)

        if key in existing_keys:
            skipped += 1
            continue

        if req.normalize_on_save:
            regs = sorted(regs)

        t = Ticket(
            draw_date=req.future_draw_date,
            n1=int(regs[0]), n2=int(regs[1]), n3=int(regs[2]), n4=int(regs[3]), n5=int(regs[4]),
            powerball=int(pb),
            type=_norm_type(req.save_type) or "QUICK_PICK",
            cost=float(req.cost_per_ticket),
            status="FUTURE",
            matched_regular_numbers=0,
            matched_powerball=False,
            prize_amount=0.0,
        )
        db.add(t)
        inserted += 1
        existing_keys.add(key)

    db.commit()

    return SaveRecommendationsResponse(
        requested=int(req.k),
        generated=int(rec.generated),
        inserted=int(inserted),
        skipped_duplicates=int(skipped),
        future_draw_date=req.future_draw_date,
        status="ok",
        message=f"Guardados {inserted} tickets FUTURE. Duplicados evitados: {skipped}.",
    )


@app.get("/ui/recommendations", response_class=HTMLResponse)
def ui_recommendations():
    body = """
      <div class="card">
        <h3>AI Recommendations</h3>
        <p class="muted">Genera combinaciones por frecuencia histórica + restricciones opcionales. Guarda como FUTURE con dedupe.</p>

        <div class="formRow" style="margin-top:12px;">
          <select id="st" class="field fieldSmall">
            <option value="">History Status: ALL</option>
            <option value="PAST">PAST</option>
            <option value="FUTURE">FUTURE</option>
          </select>

          <select id="tp" class="field fieldSmall">
            <option value="">History Type: ALL</option>
            <option value="QUICK_PICK">QUICK_PICK</option>
            <option value="MANUAL">MANUAL</option>
          </select>

          <input id="k" class="field fieldSmall" value="50" />
          <input id="seed" class="field fieldSmall" placeholder="seed (opcional)" />
          <input id="fixedFirst" class="field fieldSmall" placeholder="fixed_first (opcional)" />
          <input id="fixedNums" class="field" placeholder="fixed_numbers CSV (opcional) ej: 10,16" />
          <input id="excludeNums" class="field" placeholder="exclude_numbers CSV (opcional) ej: 1,2,3" />

          <input id="fixedPB" class="field fieldSmall" placeholder="fixed_powerball (opcional)" />
          <input id="excludePB" class="field" placeholder="exclude_powerballs CSV (opcional) ej: 1,2" />

          <button id="gen" class="btn good">Generate</button>
          <button id="copyAll" class="btn">Copy all</button>
        </div>

        <div class="formRow" style="margin-top:10px;">
          <input id="futureDate" class="field fieldSmall" placeholder="future_draw_date YYYY-MM-DD" />
          <input id="cost" class="field fieldSmall" placeholder="cost_per_ticket" value="2.0" />
          <select id="saveType" class="field fieldSmall">
            <option value="QUICK_PICK" selected>Save Type: QUICK_PICK</option>
            <option value="MANUAL">Save Type: MANUAL</option>
          </select>
          <button id="save" class="btn primary">Save as FUTURE</button>
          <button id="dlcsv" class="btn">Download CSV</button>
        </div>

        <div id="out" style="margin-top:12px;"></div>
      </div>

      <div class="toast" id="toast">OK</div>

      <script>
        function toast(msg) {
          const t = document.getElementById('toast');
          t.textContent = msg || 'OK';
          t.style.display = 'block';
          setTimeout(()=>t.style.display='none', 1100);
        }

        function esc(s){
          return (s||'')
            .replaceAll('&','&amp;')
            .replaceAll('<','&lt;')
            .replaceAll('>','&gt;');
        }

        function parseCsvInts(s){
          s = (s||'').trim();
          if(!s) return [];
          return s
            .replaceAll(';',',')
            .replaceAll('|',',')
            .split(',')
            .map(x => x.trim())
            .filter(Boolean)
            .map(x => parseInt(x,10))
            .filter(x => Number.isFinite(x));
        }

        function nextDrawDateISO(){
          // Powerball: Mon/Wed/Sat. Elegimos el próximo de esos días.
          const d = new Date();
          d.setHours(0,0,0,0);
          const want = new Set([1,3,6]); // 0=Sun,1=Mon,3=Wed,6=Sat
          for(let i=0;i<14;i++){
            const dd = new Date(d.getTime() + i*86400000);
            if(want.has(dd.getDay()) && i>0){
              const y = dd.getFullYear();
              const m = String(dd.getMonth()+1).padStart(2,'0');
              const da = String(dd.getDate()).padStart(2,'0');
              return `${y}-${m}-${da}`;
            }
          }
          // fallback: mañana
          const dd = new Date(d.getTime() + 86400000);
          const y = dd.getFullYear();
          const m = String(dd.getMonth()+1).padStart(2,'0');
          const da = String(dd.getDate()).padStart(2,'0');
          return `${y}-${m}-${da}`;
        }

        function renderTable(combos){
          if(!combos || !combos.length){
            return `<div class="muted" style="margin-top:10px;">No combos generated.</div>`;
          }

          let rows = combos.map((c, idx) => {
            const nums = [c.n1,c.n2,c.n3,c.n4,c.n5].map(x=>parseInt(x,10));
            const pb = parseInt(c.powerball,10);
            const line = `${nums.join(' ')} | PB ${pb}`;
            return `
              <tr>
                <td>${idx+1}</td>
                <td>${nums[0]}</td><td>${nums[1]}</td><td>${nums[2]}</td><td>${nums[3]}</td><td>${nums[4]}</td>
                <td>${pb}</td>
                <td><span class="code">${esc(c.key||'')}</span></td>
                <td><button class="btn" style="padding:7px 10px;" data-copy="${esc(line)}">Copy</button></td>
              </tr>
            `;
          }).join('');

          return `
            <div class="tableWrap" style="margin-top:12px;">
              <table>
                <thead>
                  <tr>
                    <th>#</th>
                    <th>N1</th><th>N2</th><th>N3</th><th>N4</th><th>N5</th>
                    <th>PB</th>
                    <th>Key</th>
                    <th>Action</th>
                  </tr>
                </thead>
                <tbody>${rows}</tbody>
              </table>
            </div>
          `;
        }

        function downloadCsvFromCombos(combos){
          if(!combos || !combos.length){
            toast('Nothing to download');
            return;
          }
          const header = ['n1','n2','n3','n4','n5','powerball','key'];
          const lines = [header.join(',')].concat(
            combos.map(c => {
              const vals = [c.n1,c.n2,c.n3,c.n4,c.n5,c.powerball,(c.key||'')];
              return vals.map(v => String(v).replaceAll('"','""')).map(v => `"${v}"`).join(',');
            })
          );
          const blob = new Blob([lines.join('\\n')], {type:'text/csv;charset=utf-8;'});
          const url = URL.createObjectURL(blob);
          const a = document.createElement('a');
          a.href = url;
          a.download = 'ai_recommendations.csv';
          document.body.appendChild(a);
          a.click();
          a.remove();
          URL.revokeObjectURL(url);
        }

        window.__lastRecs = [];

        async function generate(){
          const status = (document.getElementById('st').value||'').trim();
          const type = (document.getElementById('tp').value||'').trim();

          const k = parseInt((document.getElementById('k').value||'50').trim(),10) || 50;
          const seedRaw = (document.getElementById('seed').value||'').trim();
          const seed = seedRaw ? parseInt(seedRaw,10) : null;

          const fixedFirstRaw = (document.getElementById('fixedFirst').value||'').trim();
          const fixed_first = fixedFirstRaw ? parseInt(fixedFirstRaw,10) : null;

          const fixed_numbers = parseCsvInts(document.getElementById('fixedNums').value||'');
          const exclude_numbers = parseCsvInts(document.getElementById('excludeNums').value||'');

          const fixedPBRaw = (document.getElementById('fixedPB').value||'').trim();
          const fixed_powerball = fixedPBRaw ? parseInt(fixedPBRaw,10) : null;

          const exclude_powerballs = parseCsvInts(document.getElementById('excludePB').value||'');

          const payload = {
            k: k,
            seed: seed,
            status: status || null,
            type: type || null,
            fixed_first: fixed_first,
            fixed_numbers: fixed_numbers.length ? fixed_numbers : null,
            exclude_numbers: exclude_numbers.length ? exclude_numbers : null,
            fixed_powerball: fixed_powerball,
            exclude_powerballs: exclude_powerballs.length ? exclude_powerballs : null
          };

          const res = await fetch('/ai/recommend', {
            method:'POST',
            headers:{'Content-Type':'application/json'},
            body: JSON.stringify(payload)
          });

          const data = await res.json();
          if(!res.ok){
            alert(data.detail || 'Error');
            return;
          }

          window.__lastRecs = (data.combos || []);
          const out = document.getElementById('out');

          out.innerHTML = `
            <div class="grid">
              <div class="card"><h3>Generated</h3><p><span class="code">${data.generated}</span></p></div>
              <div class="card"><h3>Seed</h3><p><span class="code">${esc(String(data.seed ?? 'auto'))}</span></p></div>
              <div class="card"><h3>Actions</h3><p class="muted">Copy / Download / Save as FUTURE</p></div>
            </div>
            ${renderTable(window.__lastRecs)}
          `;

          document.querySelectorAll('button[data-copy]').forEach(btn => {
            btn.addEventListener('click', async () => {
              const txt = btn.getAttribute('data-copy');
              try {
                await navigator.clipboard.writeText(txt);
                toast('Copied ✅');
              } catch(e){
                toast('Copy failed');
              }
            });
          });

          toast('Generated ✅');
        }

        document.getElementById('gen').addEventListener('click', generate);

        document.getElementById('copyAll').addEventListener('click', async ()=>{
          const combos = window.__lastRecs || [];
          if(!combos.length){ toast('Nothing to copy'); return; }
          const lines = combos.map(c => {
            const nums = [c.n1,c.n2,c.n3,c.n4,c.n5].map(x=>parseInt(x,10));
            const pb = parseInt(c.powerball,10);
            return `${nums.join(' ')} | PB ${pb}`;
          });
          try{
            await navigator.clipboard.writeText(lines.join('\\n'));
            toast('Copied all ✅');
          }catch(e){
            toast('Copy failed');
          }
        });

        document.getElementById('dlcsv').addEventListener('click', ()=>{
          downloadCsvFromCombos(window.__lastRecs || []);
        });

        document.getElementById('save').addEventListener('click', async ()=>{
          const combos = window.__lastRecs || [];
          if(!combos.length){ toast('Generate first'); return; }

          const status = (document.getElementById('st').value||'').trim();
          const type = (document.getElementById('tp').value||'').trim();

          const k = parseInt((document.getElementById('k').value||'50').trim(),10) || 50;
          const seedRaw = (document.getElementById('seed').value||'').trim();
          const seed = seedRaw ? parseInt(seedRaw,10) : null;

          const fixedFirstRaw = (document.getElementById('fixedFirst').value||'').trim();
          const fixed_first = fixedFirstRaw ? parseInt(fixedFirstRaw,10) : null;

          const fixed_numbers = parseCsvInts(document.getElementById('fixedNums').value||'');
          const exclude_numbers = parseCsvInts(document.getElementById('excludeNums').value||'');

          const fixedPBRaw = (document.getElementById('fixedPB').value||'').trim();
          const fixed_powerball = fixedPBRaw ? parseInt(fixedPBRaw,10) : null;

          const exclude_powerballs = parseCsvInts(document.getElementById('excludePB').value||'');

          const future_draw_date = (document.getElementById('futureDate').value||'').trim();
          if(!future_draw_date){
            alert('future_draw_date is required (YYYY-MM-DD)');
            return;
          }

          const cost = parseFloat((document.getElementById('cost').value||'2.0').trim());
          const saveType = (document.getElementById('saveType').value||'QUICK_PICK').trim();

          const payload = {
            status: status || null,
            type: type || null,
            k: k,
            seed: seed,
            fixed_first: fixed_first,
            fixed_numbers: fixed_numbers.length ? fixed_numbers : null,
            exclude_numbers: exclude_numbers.length ? exclude_numbers : null,
            fixed_powerball: fixed_powerball,
            exclude_powerballs: exclude_powerballs.length ? exclude_powerballs : null,
            future_draw_date: future_draw_date,
            cost_per_ticket: Number.isFinite(cost) ? cost : 2.0,
            save_type: saveType || 'QUICK_PICK',
            normalize_on_save: true
          };

          const res = await fetch('/ai/save_recommendations', {
            method:'POST',
            headers:{'Content-Type':'application/json'},
            body: JSON.stringify(payload)
          });

          const data = await res.json();
          if(!res.ok){
            alert(data.detail || 'Error saving');
            return;
          }
          toast(data.message || 'Saved ✅');
        });

        // set default future date
        (function init(){
          const fd = document.getElementById('futureDate');
          if(fd && !fd.value){
            fd.value = nextDrawDateISO();
          }
        })();
      </script>
    """
    return render_app_page(title="AI Recommendations", active="ai", body_html=body)


# ============================================================
#   MAINTENANCE / ADMIN UTILITIES (opcionales pero útiles)
# ============================================================
@app.post("/admin/recompute_matches")
def admin_recompute_matches(db: Session = Depends(get_db)):
    """
    Recalcula matches/prize para tickets que tengan draw_result en la misma fecha.
    Útil si importaste draw_results después.
    """
    draws = db.query(DrawResult).all()
    draws_by_date = {d.draw_date: d for d in draws}

    tickets = db.query(Ticket).all()
    updated = 0
    for t in tickets:
        dr = draws_by_date.get(t.draw_date)
        if dr:
            calculate_matches(t, dr)
            if t.draw_date <= date.today():
                t.status = "PAST"
            updated += 1
        else:
            # sin draw_result, limpio
            t.matched_regular_numbers = 0
            t.matched_powerball = False
            t.prize_amount = 0.0

    db.commit()
    return {"ok": True, "tickets_processed": len(tickets), "tickets_updated_with_draw": updated}


@app.post("/admin/normalize_all_ticket_numbers")
def admin_normalize_all_ticket_numbers(db: Session = Depends(get_db)):
    """
    Ordena N1..N5 asc en todos los tickets (sin cambiar PB).
    Mantiene la integridad de que N1..N5 sigan siendo 5 únicos.
    """
    tickets = db.query(Ticket).all()
    changed = 0
    for t in tickets:
        regs = [int(t.n1), int(t.n2), int(t.n3), int(t.n4), int(t.n5)]
        regs_sorted = sorted(regs)
        if regs_sorted != regs:
            t.n1, t.n2, t.n3, t.n4, t.n5 = regs_sorted
            changed += 1
    db.commit()
    return {"ok": True, "tickets_total": len(tickets), "changed": changed}


@app.get("/admin/duplicates")
def admin_duplicates(db: Session = Depends(get_db)):
    """
    Devuelve duplicados por (draw_date + key). No borra nada; solo reporta.
    """
    tickets = db.query(Ticket).all()
    seen = {}
    dups = []
    for t in tickets:
        k = numbers_key([t.n1, t.n2, t.n3, t.n4, t.n5], int(t.powerball))
        kk = f"{t.draw_date.isoformat()}|{k}"
        if kk in seen:
            dups.append({"first_id": seen[kk], "dup_id": t.id, "draw_date": t.draw_date.isoformat(), "key": k})
        else:
            seen[kk] = t.id
    return {"ok": True, "duplicates": dups, "count": len(dups)}