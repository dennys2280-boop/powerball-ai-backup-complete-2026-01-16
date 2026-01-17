from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(bold=True)
GRAY_FILL = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")


def export_by_first_position():
    return {"status": "ok", "message": "export_by_first_position funcionando"}


def _autosize(ws, max_col=20, min_w=10, max_w=22):
    for c in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(min_w, min(max_w, ws.column_dimensions[get_column_letter(c)].width or min_w))

def _write_headers(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = GRAY_FILL

def export_by_first_position(
    tickets,  # list[list[int]] : [n1..n5,pb]
    output_path="powerball_first_position.xlsx",
    first_numbers=None,  # None => todos los que existan en tickets; o [1,2,3]
):
    # 1) Agrupar por primera posición (n1)
    groups = defaultdict(list)
    for t in tickets:
        if not t or len(t) != 6:
            continue
        n1 = int(t[0])
        groups[n1].append([int(x) for x in t])

    if first_numbers is None:
        first_numbers = sorted(groups.keys())
    else:
        first_numbers = [int(x) for x in first_numbers if int(x) in groups]

    wb = Workbook()
    wb.remove(wb.active)

    # 2) Hoja tabla en blanco para tus combinaciones
    ws_blank = wb.create_sheet("MY_COMBOS")
    _write_headers(ws_blank, 1, ["Group_First", "N1", "N2", "N3", "N4", "N5", "Powerball", "Notes"])
    for r in range(2, 202):  # 200 filas en blanco
        ws_blank.cell(r, 1, value="")  # Group_First
    _autosize(ws_blank, max_col=8)

    # 3) Crear hoja por grupo
    for first in first_numbers:
        sheet_name = f"FIRST_{first:02d}"
        ws = wb.create_sheet(sheet_name)

        g = groups[first]

        # --- Sección: Orden creciente / decreciente
        ws.cell(1, 1, value=f"GROUP: First position = {first}").font = HEADER_FONT

        # Orden creciente: por la tupla completa (n1..n5,pb)
        asc = sorted(g, key=lambda x: (x[0], x[1], x[2], x[3], x[4], x[5]))
        desc = sorted(g, key=lambda x: (x[0], x[1], x[2], x[3], x[4], x[5]), reverse=True)

        _write_headers(ws, 3, ["N1","N2","N3","N4","N5","PB"])
        for i, t in enumerate(asc, start=4):
            for c, v in enumerate(t, start=1):
                ws.cell(i, c, value=v)

        start_desc = 4 + len(asc) + 2
        ws.cell(start_desc - 1, 1, value="ORDER: DESC").font = HEADER_FONT
        _write_headers(ws, start_desc, ["N1","N2","N3","N4","N5","PB"])
        for i, t in enumerate(desc, start=start_desc + 1):
            for c, v in enumerate(t, start=1):
                ws.cell(i, c, value=v)

        # --- Sección: Frecuencias 1..69 (en el grupo)
        # Cuenta apariciones de números 1..69 en las 5 posiciones regulares (incluye el 1er número también).
        freq = Counter()
        for t in g:
            for v in t[:5]:
                freq[v] += 1

        freq_row = start_desc + 1 + len(desc) + 3
        ws.cell(freq_row, 1, value="FREQUENCY (1..69) in this group (regular balls only)").font = HEADER_FONT
        _write_headers(ws, freq_row + 1, ["Number", "Count"])

        r = freq_row + 2
        for num in range(1, 70):
            ws.cell(r, 1, value=num)
            ws.cell(r, 2, value=freq.get(num, 0))
            r += 1

        # --- Sección: Descomposición por posiciones (quitar 1, quitar 1-2, ..., hasta PB)
        # Aquí hacemos “stages” y contamos frecuencias en lo que queda.
        stage_start = r + 2
        ws.cell(stage_start, 1, value="DECOMPOSITION STAGES (remove from left)").font = HEADER_FONT

        # stages: remove k regular positions (0..4), dejando 5-k regulares + PB
        row_cursor = stage_start + 2
        for k in range(1, 6):  # 1..5 (quitar N1, luego N1-N2, ..., hasta quitar N1..N5)
            # cuando k=5, solo queda PB
            ws.cell(row_cursor, 1, value=f"Stage: remove first {k} position(s)").font = HEADER_FONT
            row_cursor += 1

            if k < 5:
                _write_headers(ws, row_cursor, [f"N{i}" for i in range(k+1, 6)] + ["PB"])
                row_cursor += 1

                # escribir combinaciones “recortadas”
                trimmed = []
                for t in g:
                    trimmed_row = t[k:5] + [t[5]]
                    trimmed.append(trimmed_row)

                for tr in trimmed:
                    for c, v in enumerate(tr, start=1):
                        ws.cell(row_cursor, c, value=v)
                    row_cursor += 1

                # frecuencia de números en el tramo regular restante (1..69)
                freq2 = Counter()
                for tr in trimmed:
                    for v in tr[:-1]:
                        freq2[v] += 1

                row_cursor += 1
                ws.cell(row_cursor, 1, value=f"Counts after removing first {k} position(s) (1..69)").font = HEADER_FONT
                row_cursor += 1
                _write_headers(ws, row_cursor, ["Number", "Count"])
                row_cursor += 1
                for num in range(1, 70):
                    ws.cell(row_cursor, 1, value=num)
                    ws.cell(row_cursor, 2, value=freq2.get(num, 0))
                    row_cursor += 1
                row_cursor += 2
            else:
                # k == 5: solo Powerball
                _write_headers(ws, row_cursor, ["PB"])
                row_cursor += 1
                pb_counts = Counter([t[5] for t in g])
                for pb, cnt in sorted(pb_counts.items()):
                    ws.cell(row_cursor, 1, value=pb)
                    ws.cell(row_cursor, 2, value=cnt)
                    row_cursor += 1
                row_cursor += 2

        _autosize(ws, max_col=12)

    wb.save(output_path)
    return output_path
