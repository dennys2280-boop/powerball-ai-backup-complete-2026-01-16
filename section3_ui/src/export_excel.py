from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# =====================
# CONFIGURACIÓN
# =====================
TARGET_NUMBERS = {8, 32, 52, 56, 64}
TARGET_POWERBALL = 23

USE_GREEN_BACKGROUND = True

RED_FONT = Font(color="FF0000")
HEADER_FONT = Font(bold=True)
GREEN_FILL = PatternFill(
    start_color="A7EDE7",
    end_color="A7EDE7",
    fill_type="solid"
)

HEADERS = [
    "N1", "N2", "N3", "N4", "N5",
    "Powerball", "Match_Count", "PB_Match"
]

# =====================
# DATOS DE PRUEBA
# =====================
PAST_TICKETS = [
    [8, 32, 52, 10, 11, 23],
    [1, 2, 3, 4, 5, 10],
]

FUTURE_TICKETS = [
    [8, 32, 40, 56, 64, 23],
    [9, 19, 29, 39, 49, 12],
]

# =====================
# FUNCIÓN PARA CREAR HOJAS
# =====================
def create_sheet(workbook, sheet_name, tickets):
    ws = workbook.create_sheet(title=sheet_name)

    # Encabezados
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Filas
    for row_idx, ticket in enumerate(tickets, start=2):
        numbers = ticket[:5]
        powerball = ticket[5]

        match_count = len(set(numbers) & TARGET_NUMBERS)
        pb_match = powerball == TARGET_POWERBALL

        row_data = numbers + [
            powerball,
            match_count,
            "YES" if pb_match else "NO"
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Números regulares coincidentes
            if col_idx <= 5 and value in TARGET_NUMBERS:
                cell.font = RED_FONT
                if USE_GREEN_BACKGROUND:
                    cell.fill = GREEN_FILL

            # Powerball coincidente
            if col_idx == 6 and pb_match:
                cell.font = RED_FONT
                if USE_GREEN_BACKGROUND:
                    cell.fill = GREEN_FILL

# =====================
# EJECUCIÓN
# =====================
if __name__ == "__main__":
    wb = Workbook()
    wb.remove(wb.active)

    create_sheet(wb, "PAST", PAST_TICKETS)
    create_sheet(wb, "FUTURE", FUTURE_TICKETS)

    wb.save("powerball_final.xlsx")
    print("✅ Archivo creado: powerball_final.xlsx")
